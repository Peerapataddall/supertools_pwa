from datetime import date, datetime, timedelta , time 
from functools import wraps
from sqlite3 import Connection as SQLite3Connection

import os  # ← ใช้สำหรับอัปโหลดโลโก้
import io  # ← ใช้ทำไฟล์ Excel ในหน่วยความจำ
from typing import List, Dict, Optional
from flask import Flask, render_template, redirect, url_for, request, flash, abort, jsonify, Blueprint, render_template_string, send_file
import re
from flask_sqlalchemy import SQLAlchemy
from collections import defaultdict
from sqlalchemy import func, case
from sqlalchemy.exc import OperationalError
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user, current_user, login_required
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename  # ← ใช้บันทึกไฟล์โลโก้
from flask_migrate import Migrate
from sqlalchemy import event, select, or_, CheckConstraint, UniqueConstraint, inspect as sa_inspect

from sqlalchemy.engine import Engine
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload, selectinload, relationship, Mapped, mapped_column, foreign, aliased

from urllib.parse import urlparse
from decimal import Decimal
from sqlalchemy.types import Numeric
from sqlalchemy import Enum as SAEnum
from werkzeug.routing import BuildError
from types import SimpleNamespace
from sqlalchemy import text
import types
from enum import Enum
from flask import current_app
import sys



# ================== App & DB ==================
app = Flask(__name__)

# อ่าน SECRET_KEY จาก env ถ้ามี ถ้าไม่มีใช้ค่าเดิม
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "change-me")

# เลือก DB จาก DATABASE_URL ถ้ามี ไม่งั้นใช้ sqlite เดิม (สำหรับ dev)
db_url = os.getenv("DATABASE_URL")

if not db_url:
    # โหมด dev ในเครื่อง → ใช้ sqlite
    db_url = "sqlite:///supertools.db"
else:
    # บางผู้ให้บริการให้ prefix postgres:// ต้องแปลงเป็น postgresql+psycopg2://
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql+psycopg2://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024  # 5MB เผื่อไฟล์โลโก้ / รูป


db = SQLAlchemy(app)

@app.context_processor
def inject_global_helpers():
    return {
        "th_status": th_status,
        "th_credit_term": th_credit_term,
    }

migrate = Migrate(app, db)



def _is_migration_command() -> bool:
    """
    คืนค่า True ถ้ากำลังรันคำสั่งเกี่ยวกับ flask db / alembic
    เพื่อไม่ให้ create_all / seed ทำงาน
    """
    argv = " ".join(sys.argv).lower()
    return (
        ("flask" in argv and " db " in f" {argv} ")
        or "alembic" in argv
    )

def _xlsx_response(workbook, filename: str):
    """
    แปลง openpyxl Workbook -> response สำหรับ download xlsx
    """
    bio = BytesIO()
    workbook.save(bio)
    bio.seek(0)

    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.context_processor
def inject_template_helpers():
    def safe_href(endpoint: str, **kwargs) -> str:
        """คืนลิงก์ของ endpoint ถ้ามีจริง ไม่งั้นคืน '#'"""
        try:
            # ไม่ใช้ current_app ในเทมเพลต แต่ใช้ที่นี่ได้ ปลอดภัยใน app context
            if endpoint and endpoint in current_app.view_functions:
                return url_for(endpoint, **kwargs)
        except BuildError:
            pass
        except Exception:
            pass
        return "#"
    return dict(safe_href=safe_href)


@app.template_filter("strftime")
def jinja_strftime(value, fmt="%d/%m/%Y"):
    """ใช้ใน Jinja: {{ some_date|strftime('%d/%m/%Y') }}"""
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime(fmt)
    s = str(value).strip()
    try:
        dt = datetime.fromisoformat(s)
        return dt.strftime(fmt)
    except Exception:
        pass
    try:
        from dateutil import parser
        return parser.parse(s).strftime(fmt)
    except Exception:
        return s

# ================== Models ==================
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, index=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(120))
    is_active = db.Column(db.Boolean, default=True)

# ===== Claims (งานเคลม) =====
ClaimStatusEnum = SAEnum(
    "DRAFT", "SUBMITTED", "APPROVED", "CLOSED", name="claim_status_enum"
)

class Claim(db.Model):
    __tablename__ = "claims"
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(32), unique=True, index=True)           # CLM20251110xxxx
    date = db.Column(db.Date, nullable=False, default=date.today)
    status = db.Column(ClaimStatusEnum, nullable=False, default="DRAFT")

    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id"), nullable=False)
    quote_id = db.Column(db.Integer, db.ForeignKey("sales_doc.id"), nullable=False)  # อ้าง QU
    remark = db.Column(db.Text, default="")

    # ---- Contract / Installment billing ----
    billing_mode = db.Column(db.String(12), default="ONCE")  # ONCE / INSTALLMENT
    contract_start = db.Column(db.Date)  # วันที่เริ่มสัญญา (ใช้กับ QU/CT)
    contract_end = db.Column(db.Date)    # วันที่สิ้นสุดสัญญา
    installment_count = db.Column(db.Integer, default=0)     # จำนวนงวด (เช่น 12)

    customer = db.relationship("Customer", lazy="joined")
    quote = db.relationship("SalesDoc", lazy="joined", foreign_keys=[quote_id])
    items = db.relationship("ClaimItem", backref="claim", cascade="all, delete-orphan")

class ClaimItem(db.Model):
    __tablename__ = "claim_items"
    id = db.Column(db.Integer, primary_key=True)
    claim_id = db.Column(db.Integer, db.ForeignKey("claims.id"), nullable=False)
    sales_item_id = db.Column(db.Integer, db.ForeignKey("sales_item.id"), nullable=False)
    qty_claim = db.Column(db.Float, nullable=False, default=1.0)
    replacement_equipment_id = db.Column(db.Integer, db.ForeignKey("equipment.id"), nullable=True)

    sales_item = db.relationship("SalesItem", lazy="joined")
    replacement_equipment = db.relationship("Equipment", lazy="joined")

# --- RBAC ---
class Role(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)

class Permission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(80), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)

class UserRole(db.Model):
    __table_args__ = (db.UniqueConstraint("user_id", "role_id", name="uq_user_role"), )
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), index=True, nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey("role.id"), index=True, nullable=False)

class RolePermission(db.Model):
    __table_args__ = (db.UniqueConstraint("role_id", "perm_id", name="uq_role_perm"), )
    id = db.Column(db.Integer, primary_key=True)
    role_id = db.Column(db.Integer, db.ForeignKey("role.id"), index=True, nullable=False)
    perm_id = db.Column(db.Integer, db.ForeignKey("permission.id"), index=True, nullable=False)

class UserPermission(db.Model):
    __table_args__ = (db.UniqueConstraint("user_id", "perm_id", name="uq_user_perm"), )
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), index=True, nullable=False)
    perm_id = db.Column(db.Integer, db.ForeignKey("permission.id"), index=True, nullable=False)

# ---------- Purchases Models ----------
class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True)
    tax_id = db.Column(db.String(32))
    phone = db.Column(db.String(64))
    address = db.Column(db.Text)
    district = db.Column(db.String(100))
    amphoe = db.Column(db.String(100))
    province = db.Column(db.String(100))
    postcode = db.Column(db.String(10))

class PurchaseOrder(db.Model):
    __tablename__ = "purchase_order"
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(32), unique=True, index=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey("supplier.id"), nullable=False)
    po_date = db.Column(db.Date, nullable=False, default=date.today)
    status = db.Column(db.String(20), nullable=False, default="DRAFT")  # DRAFT, APPROVED, ORDERED

    supplier = db.relationship(Supplier, lazy="joined")
    items = db.relationship("POItem", backref="po", cascade="all, delete-orphan")

    @property
    def amount_subtotal(self):
        return sum((it.qty or 0) * (it.unit_cost or 0) for it in self.items)

    @property
    def amount_discount(self):
        return sum(((it.qty or 0) * (it.unit_cost or 0)) * ((it.discount_pct or 0)/100) for it in self.items)

    @property
    def amount_total(self):
        return self.amount_subtotal - self.amount_discount

class POItem(db.Model):
    __tablename__ = "po_item"
    id = db.Column(db.Integer, primary_key=True)
    po_id = db.Column(db.Integer, db.ForeignKey("purchase_order.id"), index=True, nullable=False)
    sku = db.Column(db.String(80))
    name = db.Column(db.String(255), nullable=False)

    brand = db.Column(db.String(120), default="")  # NEW: ยี่ห้อ
    qty = db.Column(db.Float, nullable=False, default=1.0)
    unit = db.Column(db.String(32), default="ชิ้น")
    unit_cost = db.Column(db.Float, nullable=False, default=0.0)
    discount_pct = db.Column(db.Float, nullable=False, default=0.0)

    @property
    def line_subtotal(self):
        return (self.qty or 0) * (self.unit_cost or 0)

    @property
    def line_discount(self):
        return self.line_subtotal * ((self.discount_pct or 0)/100)

    @property
    def line_total(self):
        return self.line_subtotal - self.line_discount

class GRNItem(db.Model):
    __tablename__ = "grn_item"
    id = db.Column(db.Integer, primary_key=True)
    grn_id = db.Column(db.Integer, db.ForeignKey("goods_receipt.id"), index=True, nullable=False)
    sku = db.Column(db.String(80))
    name = db.Column(db.String(255), nullable=False)

    brand = db.Column(db.String(120), default="")  # NEW: ยี่ห้อ
    qty = db.Column(db.Float, nullable=False, default=0.0)
    unit = db.Column(db.String(32), default="ชิ้น")
    unit_cost = db.Column(db.Float, nullable=False, default=0.0)


# ---------- Incoming Equipments (Pending to add into system) ----------
class IncomingEquipment(db.Model):
    """
    สเตจอุปกรณ์ที่ 'รับสินค้าเข้า (GRN)' แล้ว แต่ยังไม่ได้สร้างเป็น Equipment จริง
    เพื่อให้ผู้ใช้เข้ามาหน้าเพิ่มอุปกรณ์แล้วกด "เพิ่ม" และให้ระบบเติมข้อมูลให้อัตโนมัติ
    """
    __tablename__ = "incoming_equipment"
    id = db.Column(db.Integer, primary_key=True)

    grn_id = db.Column(db.Integer, db.ForeignKey("goods_receipt.id"), index=True, nullable=False)
    grn_item_id = db.Column(db.Integer, db.ForeignKey("grn_item.id"), index=True, nullable=False)

    name = db.Column(db.String(255), nullable=False)
    brand = db.Column(db.String(120), default="")
    unit_cost = db.Column(db.Float, nullable=False, default=0.0)
    received_date = db.Column(db.Date, nullable=False, default=date.today)

    status = db.Column(db.String(20), nullable=False, default="PENDING")  # PENDING / DONE
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    grn = db.relationship("GoodsReceipt", lazy="joined")
    grn_item = db.relationship("GRNItem", lazy="joined")


# ---------- Company Profile ----------
class CompanyProfile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, default="ชื่อบริษัทของคุณ")
    address = db.Column(db.Text, default="")
    district = db.Column(db.String(100), default="")
    amphoe = db.Column(db.String(100), default="")
    province = db.Column(db.String(100), default="")
    postcode = db.Column(db.String(10), default="")
    phone = db.Column(db.String(50), default="")
    tax_id = db.Column(db.String(32), default="")
    logo_path = db.Column(db.String(255), default="")  # eg. uploads/company/logo.png

def get_company() -> CompanyProfile:
    row = db.session.get(CompanyProfile, 1)
    if not row:
        row = CompanyProfile(id=1)
        db.session.add(row)
        db.session.commit()
    return row

class GoodsReceipt(db.Model):
    __tablename__ = "goods_receipt"
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(32), unique=True, index=True)
    po_id = db.Column(db.Integer, db.ForeignKey("purchase_order.id"), nullable=False)
    grn_date = db.Column(db.Date, nullable=False, default=date.today)
    status = db.Column(db.String(20), nullable=False, default="RECEIVED")

    po = db.relationship(PurchaseOrder, lazy="joined")
    items = db.relationship("GRNItem", backref="grn", cascade="all, delete-orphan")

    @property
    def amount_subtotal(self):
        return sum((it.qty or 0) * (it.unit_cost or 0) for it in self.items)

    @property
    def amount_total(self):
        return self.amount_subtotal

# ---------- Customers ----------
class Customer(db.Model):
    __tablename__ = "customer"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False, index=True)
    address = db.Column(db.Text, default="")
    district = db.Column(db.String(120), default="")
    amphoe = db.Column(db.String(120), default="")
    province = db.Column(db.String(120), default="")
    postcode = db.Column(db.String(10), default="")
    phone = db.Column(db.String(64), default="")
    tax_id = db.Column(db.String(32), default="")
    contact_name = db.Column(db.String(120), default="")
    contact_phone = db.Column(db.String(64), default="")
    credit_term_days = db.Column(db.Integer, default=0)  # เครดิตลูกค้า (วัน) / 0=เงินสด
    payment_terms = db.Column(db.String(120), default="")  # เงื่อนไขชำระเงิน (ข้อความ)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

# ---------- Equipment Module ----------
EQUIP_STATUS = ("READY", "RENTED", "REPAIR")
EQUIP_STATUS_THAI = {
    "READY": "พร้อมให้เช่า",
    "RESERVED": "จองไว้",
    "RENTED": "ถูกเช่า",
    "REPAIR": "รอซ่อม",
    "LOST": "สูญหาย",
}


# ---------- Global Status Thai Mapping ----------
DOC_STATUS_THAI = {
    # SalesDoc / DeliveryDoc / Repair / etc.
    "DRAFT": "ร่าง",
    "PENDING": "รอดำเนินการ",
    "APPROVED": "อนุมัติแล้ว",
    "CANCELLED": "ยกเลิก",
    "CLOSED": "ปิดงาน",
    "OPEN": "เปิดงาน",
    "PAID": "ชำระแล้ว",
    "UNPAID": "ค้างชำระ",
    "OVERDUE": "เกินกำหนด",
    "PLANNED": "วางแผน",
    "READY": "พร้อม",
    "INVOICED": "ออกใบกำกับแล้ว",
    "RECEIPTED": "ออกใบเสร็จแล้ว",
    "ACTIVE": "ใช้งาน",
    "RELEASED": "คืนแล้ว",
}

def th_status(value: str) -> str:
    v = (value or "").strip()
    if not v:
        return ""
    if v in EQUIP_STATUS_THAI:
        return EQUIP_STATUS_THAI.get(v, v)
    return DOC_STATUS_THAI.get(v, v)


def th_credit_term(days: int | None, terms: str = "") -> str:
    """แสดงเครดิต/เงื่อนไขชำระเงินภาษาไทยแบบสั้น ๆ"""
    try:
        d = int(days or 0)
    except Exception:
        d = 0
    t = (terms or "").strip()
    if d <= 0:
        return t if t else "เงินสด"
    base = f"เครดิต {d} วัน"
    return f"{base} • {t}" if t else base

class Category(db.Model):
    __tablename__ = "category"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False, unique=True)
    prefix_sku = db.Column(db.String(20), nullable=False, unique=True)

class Equipment(db.Model):
    __tablename__ = "equipment"
    id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(40), nullable=False, unique=True, index=True)
    name = db.Column(db.String(255), nullable=False, index=True)

    brand = db.Column(db.String(120), default="")  # NEW: ยี่ห้อ

    warehouse = db.Column(db.String(60), default="MAIN", index=True)  # NEW: คลัง/สาขา

    category_id = db.Column(db.Integer, db.ForeignKey("category.id"), nullable=False, index=True)
    category = db.relationship(Category, lazy="joined")

    received_date = db.Column(db.Date, nullable=False)
    cost = db.Column(db.Float, nullable=False, default=0.0)

    life_years = db.Column(db.Integer, default=0)
    life_months = db.Column(db.Integer, default=0)
    life_days = db.Column(db.Integer, default=0)

    image_path = db.Column(db.String(255), default="")
    status = db.Column(db.String(12), nullable=False, default="READY")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    @property
    def status_th(self) -> str:
        return EQUIP_STATUS_THAI.get(self.status, self.status)

    @property
    def lifetime_days(self) -> int:
        y = (self.life_years or 0) * 365
        m = (self.life_months or 0) * 30
        d = (self.life_days or 0)
        return max(1, y + m + d)

    @property
    def price_per_day_break_even(self) -> float:
        return round((self.cost or 0) / self.lifetime_days, 2)

    @property
    def price_per_month_break_even(self) -> float:
        return round(self.price_per_day_break_even * 30, 2)

    @property
    def price_per_year_break_even(self) -> float:
        return round(self.price_per_day_break_even * 365, 2)

    @property
    def days_used(self) -> int:
        if not self.received_date:
            return 0
        return max(0, (date.today() - self.received_date).days)

    @property
    def depreciation_per_day(self) -> float:
        return round((self.cost or 0) / self.lifetime_days, 2)

    @property
    def depreciation_per_month(self) -> float:
        return round(self.depreciation_per_day * 30, 2)

    @property
    def depreciation_per_year(self) -> float:
        return round(self.depreciation_per_day * 365, 2)

    @property
    def accumulated_depr(self) -> float:
        return round(min((self.cost or 0), self.depreciation_per_day * self.days_used), 2)

    @property
    def current_value(self) -> float:
        return round(max(0.0, (self.cost or 0) - self.accumulated_depr), 2)

class EquipmentLog(db.Model):
    __tablename__ = "equipment_log"
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey("equipment.id"), index=True, nullable=False)
    action = db.Column(db.String(30), nullable=False)  # ADD, EDIT, STATUS, RENT_OUT, RETURN, CLAIM_SEND, CLAIM_DONE ...
    note = db.Column(db.Text, default="")
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    customer_name = db.Column(db.String(200), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    equipment = db.relationship(Equipment, lazy="joined")
    user = db.relationship(User, lazy="joined")

def _equip_log(equipment, action: str, note: str = "", ref_model: str = "Claim", ref_id: int | None = None):
    try:
        EquipmentLogModel = EquipmentLog
    except NameError:
        EquipmentLogModel = None
    if EquipmentLogModel is None or equipment is None:
        return
    db.session.add(EquipmentLogModel(
        equipment_id=equipment.id,
        action=action,
        note=note,
        user_id=(current_user.id if current_user.is_authenticated else None),
    ))

class Promotion(db.Model):
    __tablename__ = "promotion"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    active = db.Column(db.Boolean, default=True)
    start_date = db.Column(db.Date, nullable=True)
    end_date   = db.Column(db.Date, nullable=True)
    min_items = db.Column(db.Integer, default=0)
    rental_unit = db.Column(db.String(6), default="DAY")
    min_duration = db.Column(db.Integer, default=0)
    discount_type  = db.Column(db.String(3), default="PCT")   # PCT | AMT
    discount_value = db.Column(db.Float, default=0.0)
    cheapest_units_to_discount = db.Column(db.Integer, default=1)
    note = db.Column(db.Text, default="")
    def is_in_effect(self, on_date: date) -> bool:
        if not self.active: return False
        if self.start_date and on_date < self.start_date: return False
        if self.end_date   and on_date > self.end_date:   return False
        return True

# ---------- Sales Documents ----------
SALE_TYPES = ("QU","BL","IV","RC","DN","RN")
TAX_MODE = ("EXC","INC","NONE")
WHT_CHOICES = (0,1,2,3,5)

class SalesDoc(db.Model):
    __tablename__ = "sales_doc"
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(32), unique=True, index=True)
    doc_type = db.Column(db.String(2), nullable=False)
    status = db.Column(db.String(20), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id"), nullable=False)
    warehouse = db.Column(db.String(60), default="MAIN", index=True)  # NEW: คลัง/สาขา
    project_name = db.Column(db.String(255), default="")
    po_customer = db.Column(db.String(64), default="")
    credit_days = db.Column(db.Integer, default=0)
    tax_mode = db.Column(db.String(4), default="EXC")
    wht_pct = db.Column(db.Integer, default=0)
    date = db.Column(db.Date, nullable=False, default=date.today)
    remark = db.Column(db.Text, default="")

    # ---- Contract / Installment billing ----
    # ONCE = ออกเอกสารปกติชุดเดียว (เดิม)
    # INSTALLMENT = สร้างสัญญา/PO ใหญ่ (CT) + ตารางงวดรายเดือน แล้วออก BL/IV/RC แยกตามงวด
    billing_mode = db.Column(db.String(12), default="ONCE")  # ONCE / INSTALLMENT
    contract_start = db.Column(db.Date)  # วันที่เริ่มสัญญา (ใช้กับ QU/CT)
    contract_end = db.Column(db.Date)    # วันที่สิ้นสุดสัญญา
    installment_count = db.Column(db.Integer, default=0)     # จำนวนงวด (เช่น 12)
    amount_subtotal = db.Column(db.Float, default=0.0)
    amount_vat = db.Column(db.Float, default=0.0)
    amount_total = db.Column(db.Float, default=0.0)
    amount_wht = db.Column(db.Float, default=0.0)
    amount_grand = db.Column(db.Float, default=0.0)
    parent_id = db.Column(db.Integer, db.ForeignKey("sales_doc.id"))
    customer = db.relationship(Customer, lazy="joined")
    parent = db.relationship("SalesDoc", remote_side=[id])
    items = db.relationship("SalesItem", backref="doc", cascade="all, delete-orphan")


    @property
    def status_th(self) -> str:
        return th_status(self.status)

class SalesItem(db.Model):
    __tablename__ = "sales_item"
    id = db.Column(db.Integer, primary_key=True)
    doc_id = db.Column(db.Integer, db.ForeignKey("sales_doc.id"), index=True, nullable=False)
    # เลือกเป็น "หมวดหมู่" (Category) ตอนทำใบเสนอราคา; เลือก "ตัวอุปกรณ์จริง" ตอนทำใบจอง (BK)
    category_id = db.Column(db.Integer, db.ForeignKey("category.id"), nullable=True)
    category_prefix = db.Column(db.String(50), nullable=True)  # เก็บ prefix_sku ไว้เผื่อหมวดถูกแก้ชื่อ
    brand = db.Column(db.String(120), nullable=True)  # NEW: ยี่ห้อ/รุ่น (ใช้สำหรับจองสต็อก)
    allocated_skus = db.Column(db.Text, nullable=True)  # เก็บ SKU อุปกรณ์จริงที่ถูกเลือกตอนทำใบจอง (คั่นด้วย ,)
    line_status = db.Column(db.String(12), nullable=False, default="APPROVED", index=True)  # NEW: สถานะระดับรายการ (APPROVED/REJECTED)
    source_qu_item_id = db.Column(db.Integer, nullable=True, index=True)  # อ้างอิง SalesItem.id ของ QU แม่ (ใช้ sync SKU กลับไป QU)

    image_path = db.Column(db.String(255), default="")
    name = db.Column(db.String(255), nullable=False)
    qty = db.Column(db.Float, default=1.0)
    rent_unit = db.Column(db.String(6), default="DAY")  # HOUR/DAY/MONTH/YEAR
    rent_duration = db.Column(db.Integer, default=1)
    unit_price = db.Column(db.Float, default=0.0)
    discount_pct = db.Column(db.Float, default=0.0)
    line_subtotal = db.Column(db.Float, default=0.0)
    line_total = db.Column(db.Float, default=0.0)



class StockReservation(db.Model):
    __tablename__ = "stock_reservation"
    id = db.Column(db.Integer, primary_key=True)
    doc_type = db.Column(db.String(4), nullable=False)   # QU/BK/...
    doc_id = db.Column(db.Integer, nullable=False, index=True)
    sales_item_id = db.Column(db.Integer, nullable=True) # อ้างอิง SalesItem.id (ถ้ามี)
    category_id = db.Column(db.Integer, db.ForeignKey("category.id"), nullable=False, index=True)
    brand = db.Column(db.String(120), nullable=True, index=True)
    warehouse = db.Column(db.String(60), default="MAIN", index=True)  # NEW: คลัง/สาขา
    qty = db.Column(db.Float, nullable=False, default=0.0)
    status = db.Column(db.String(12), nullable=False, default="ACTIVE")  # ACTIVE/RELEASED
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

class SalesInstallment(db.Model):
    """
    ตารางงวดรายเดือน (ลูกของสัญญา/PO ใหญ่)
    contract = SalesDoc(doc_type="CT")
    """
    __tablename__ = "sales_installment"
    id = db.Column(db.Integer, primary_key=True)

    contract_id = db.Column(db.Integer, db.ForeignKey("sales_doc.id"), index=True, nullable=False)
    installment_no = db.Column(db.Integer, nullable=False)  # 1..N

    period_start = db.Column(db.Date, nullable=False)
    period_end = db.Column(db.Date, nullable=False)
    bill_date = db.Column(db.Date, nullable=False)
    due_date = db.Column(db.Date, nullable=False)

    po_customer_sub = db.Column(db.String(64), default="")   # PO ย่อยของลูกค้า (ถ้ามี)

    status = db.Column(db.String(20), default="PLANNED")     # PLANNED/READY/INVOICED/RECEIPTED/OVERDUE/CANCELLED

    amount_subtotal = db.Column(db.Float, default=0.0)
    amount_vat = db.Column(db.Float, default=0.0)
    amount_total = db.Column(db.Float, default=0.0)
    amount_wht = db.Column(db.Float, default=0.0)
    amount_grand = db.Column(db.Float, default=0.0)

    bill_id = db.Column(db.Integer, db.ForeignKey("sales_doc.id"))
    invoice_id = db.Column(db.Integer, db.ForeignKey("sales_doc.id"))
    receipt_id = db.Column(db.Integer, db.ForeignKey("sales_doc.id"))

    contract = db.relationship("SalesDoc", foreign_keys=[contract_id], lazy="joined")
    bill = db.relationship("SalesDoc", foreign_keys=[bill_id], lazy="joined")
    invoice = db.relationship("SalesDoc", foreign_keys=[invoice_id], lazy="joined")
    receipt = db.relationship("SalesDoc", foreign_keys=[receipt_id], lazy="joined")

    __table_args__ = (
        db.UniqueConstraint("contract_id", "installment_no", name="uq_sales_installment_contract_no"),
    )

# ---------- Spare parts (อะไหล่) ----------
class SparePart(db.Model):
    __tablename__ = "spare_parts"
    id: Mapped[int] = mapped_column(primary_key=True)
    code: Mapped[str] = mapped_column(db.String(32), unique=True, nullable=False, index=True)
    name: Mapped[str] = mapped_column(db.String(255), nullable=False, index=True)
    unit: Mapped[str] = mapped_column(db.String(32), default="ชิ้น", nullable=False)
    unit_cost: Mapped[Decimal] = mapped_column(Numeric(12,2), default=0, nullable=False)
    stock_qty: Mapped[Decimal] = mapped_column(Numeric(12,2), default=0, nullable=False)
    is_active: Mapped[bool] = mapped_column(db.Boolean, default=True, nullable=False)
    notes: Mapped[str | None] = mapped_column(db.Text)
    __table_args__ = (
        CheckConstraint("unit_cost >= 0", name="ck_spare_parts_unit_cost_nonneg"),
        CheckConstraint("stock_qty >= 0", name="ck_spare_parts_stock_qty_nonneg"),
    )


# ================== Repairs Models ==================


class RepairJob(db.Model):
    __tablename__ = "repair_jobs"
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(24), unique=True, nullable=False)
    status = db.Column(db.String(16), nullable=False, default="OPEN")

    claim_id = db.Column(db.Integer, db.ForeignKey(f"{Claim.__tablename__}.id"))
    claim_item_id = db.Column(db.Integer, db.ForeignKey(f"{ClaimItem.__tablename__}.id"))
    equipment_id = db.Column(db.Integer, db.ForeignKey(f"{Equipment.__tablename__}.id"), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey(f"{Customer.__tablename__}.id"))

    symptom = db.Column(db.Text)                      # สรุปอาการเสีย
    labor_cost = db.Column(db.Numeric(12,2), default=0)
    parts_total = db.Column(db.Numeric(12,2), default=0)
    total_cost = db.Column(db.Numeric(12,2), default=0)

    opened_at = db.Column(db.DateTime, default=datetime.utcnow)
    closed_at = db.Column(db.DateTime)

    items = db.relationship("RepairItem", backref="job", cascade="all, delete-orphan")

class RepairItem(db.Model):
    __tablename__ = "repair_items"
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey(f"{RepairJob.__tablename__}.id"), nullable=False)
    part_id = db.Column(db.Integer, db.ForeignKey(f"{SparePart.__tablename__}.id"), nullable=False)

    # อะไหล่
    part_code = db.Column(db.String(64))
    part_name = db.Column(db.String(255))

    qty = db.Column(db.Numeric(12,2), default=1)
    unit_price = db.Column(db.Numeric(12,2), default=0)
    line_total = db.Column(db.Numeric(12,2), default=0)


# ==================== Transport / Delivery Models ====================



class DeliveryStatus(str, Enum):
    PENDING = "PENDING"     # รอจัดส่ง
    ONGOING = "ONGOING"     # กำลังจัดส่ง
    DONE = "DONE"           # จัดส่งสำเร็จ
    CANCELLED = "CANCELLED" # ยกเลิกการส่ง

class DeliveryType(str, Enum):
    DL  = "DL"   # ส่งปกติ (จากใบเสนอราคาหรือเอกสารขายของคุณ)
    DLC = "DLC"  # ส่งอุปกรณ์เคลม

class DeliveryVehicle(db.Model):
    __tablename__ = "delivery_vehicles"
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)  # เช่น TRK-001
    name = db.Column(db.String(120), nullable=False)              # ชื่อเล่นรถ / รุ่น
    plate_no = db.Column(db.String(50))                           # ป้ายทะเบียน
    capacity_note = db.Column(db.String(200))                     # หมายเหตุความจุ
    is_active = db.Column(db.Boolean, default=True)

class Driver(db.Model):
    __tablename__ = "drivers"
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)  # เช่น DRV-001
    full_name = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(50))
    license_no = db.Column(db.String(80))
    is_active = db.Column(db.Boolean, default=True)

# ==== Delivery Photos ========================================================

class DeliveryPhoto(db.Model):
    __tablename__ = "delivery_photos"

    id = db.Column(db.Integer, primary_key=True)
    doc_id = db.Column(db.Integer, db.ForeignKey("delivery_docs.id"), nullable=False, index=True)
    kind = db.Column(db.String(20), nullable=False)  # 'BEFORE' หรือ 'AFTER'
    filename = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    doc = db.relationship("DeliveryDoc", back_populates="photos")


class DeliveryDoc(db.Model):
    __tablename__ = "delivery_docs"
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(30), unique=True, nullable=False)      # รหัสเอกสารส่ง เช่น DL-YYYYMM-0001 / DLC-...
    d_type = db.Column(db.Enum(DeliveryType), nullable=False)           # DL / DLC
    status = db.Column(db.Enum(DeliveryStatus), default=DeliveryStatus.PENDING, nullable=False)
    delivery_date = db.Column(db.Date, nullable=True)
    # อ้างอิงเอกสารต้นทาง
    source_type = db.Column(db.String(20), nullable=False)              # 'QUOTATION' หรือ 'CLAIM'
    source_id = db.Column(db.Integer, nullable=False)                    # id ของใบเสนอราคา / ใบเคลม

    # ข้อมูลสถานที่ส่ง/ผู้รับ
    ship_to_name = db.Column(db.String(200))
    ship_to_phone = db.Column(db.String(80))
    ship_to_address = db.Column(db.Text)
    ship_to_note = db.Column(db.String(255))

    # จัดสายรถ
    vehicle_id = db.Column(db.Integer, db.ForeignKey("delivery_vehicles.id"))
    driver_id  = db.Column(db.Integer, db.ForeignKey("drivers.id"))
    schedule_at = db.Column(db.DateTime)        # วันที่-เวลาที่วางแผนจัดส่ง
    started_at  = db.Column(db.DateTime)        # เริ่มจัดส่งจริง
    finished_at = db.Column(db.DateTime)        # สำเร็จจริง

    # ยกเลิก & นัดส่งใหม่
    cancel_reason_code = db.Column(db.String(20))   # 'ADDR_CHANGED', 'DATE_CHANGED', 'AREA_CHANGED', 'ACCIDENT', 'OTHER'
    cancel_reason_text = db.Column(db.String(255))  # ถ้า OTHER ให้ระบุข้อความ
    reschedule_at = db.Column(db.DateTime)          # วันที่นัดส่งใหม่ (เมื่อตั้งค่านี้แล้วจะกลับสู่ PENDING)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    vehicle = db.relationship("DeliveryVehicle", lazy="joined")
    driver  = db.relationship("Driver", lazy="joined")
    items   = db.relationship("DeliveryItem", cascade="all, delete-orphan", backref="doc", lazy="selectin")

    photos = db.relationship("DeliveryPhoto", back_populates="doc", lazy="selectin")

    @property
    def photos_before(self):
        return [p for p in self.photos if p.kind == "BEFORE"]

    @property
    def photos_after(self):
        return [p for p in self.photos if p.kind == "AFTER"]

    __table_args__ = (
        db.Index("ix_delivery_unique_src", "source_type", "source_id", unique=True),
    )

class DeliveryItem(db.Model):
    __tablename__ = "delivery_items"
    id = db.Column(db.Integer, primary_key=True)
    doc_id = db.Column(db.Integer, db.ForeignKey("delivery_docs.id"), nullable=False)
    # อ้างอิง item ต้นทาง (ถ้าต้องการย้อนกลับไปหาออเดอร์/เคลมไอเท็ม)
    source_item_id = db.Column(db.Integer)
    product_name   = db.Column(db.String(200), nullable=False)
    qty            = db.Column(db.Float, default=1)
    unit           = db.Column(db.String(30), default="ชิ้น")
    note           = db.Column(db.String(200))



# ================== GIFT / LOYALTY MODELS ==================

class GiftCampaign(db.Model):
    __tablename__ = "gift_campaigns"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)        # ชื่อแคมเปญ เช่น "ของขวัญรอบ 1/2568"
    description = db.Column(db.Text)                        # รายละเอียดเพิ่มเติม (ถ้ามี)

    # ช่วงวันที่ของแคมเปญ (ใช้คำนวณยอด)
    period_start = db.Column(db.Date, nullable=False)
    period_end = db.Column(db.Date, nullable=False)

    # ข้อมูลรอบ (เพื่อให้รู้ว่าออกของทุกกี่เดือน เช่น 4 เดือน)
    cycle_months = db.Column(db.Integer, nullable=False, default=4)
    anchor_month = db.Column(db.Integer, nullable=False, default=1)  # ปกติ 1 = มกราคม

    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    tiers = db.relationship("GiftTier", backref="campaign", lazy="selectin",
                            cascade="all, delete-orphan")
    results = db.relationship(
        "GiftResult",
        back_populates="campaign",
        lazy="selectin",
        cascade="all, delete-orphan",   
    )


class GiftTier(db.Model):
    __tablename__ = "gift_tiers"

    id = db.Column(db.Integer, primary_key=True)
    campaign_id = db.Column(db.Integer, db.ForeignKey("gift_campaigns.id"), nullable=False)

    code = db.Column(db.String(50), nullable=False)         # เช่น "A", "B", "C"
    name = db.Column(db.String(200), nullable=False)        # ชื่อแสดง เช่น "เกรด A"
    min_amount = db.Column(Numeric(12, 2), nullable=False)  # ยอดขั้นต่ำที่เข้าเกรดนี้

    sort_order = db.Column(db.Integer, nullable=False, default=0)  # เอาไว้เรียงเกรด
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint("campaign_id", "code", name="uq_gift_tier_campaign_code"),
    )


class GiftResult(db.Model):
    __tablename__ = "gift_results"

    id = db.Column(db.Integer, primary_key=True)

    campaign_id = db.Column(
        db.Integer,
        db.ForeignKey("gift_campaigns.id"),
        nullable=False,
        index=True,
    )

    customer_id = db.Column(
        db.Integer,
        db.ForeignKey("customer.id"),
        nullable=False,
        index=True,
    )

    total_amount = db.Column(Numeric(14, 2), nullable=False, default=0)
    tier_name = db.Column(db.String(50), nullable=True)
    times_achieved = db.Column(db.Integer, nullable=False, default=0)
    last_achieved_at = db.Column(db.DateTime, nullable=True)

    status = db.Column(db.String(16), nullable=False, default="PENDING")

    # map python field = given_at -> DB column = last_given_at
    given_at = db.Column("last_given_at", db.DateTime, nullable=True)

    campaign = db.relationship("GiftCampaign", back_populates="results")
    customer = db.relationship("Customer", backref="gift_results")

# ================== RETURN NOTES (ใบคืนสินค้า) ==================

class ReturnDoc(db.Model):
    """
    เอกสารใบคืนสินค้า

    - คืน "ขาย/เช่า" ได้ 2 แบบ
      1) อ้างอิงใบเสนอราคา (QU)  -> quote_id
      2) อ้างอิงใบจอง (BK)        -> booking_id
    """
    __tablename__ = "return_docs"

    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(32), unique=True, nullable=False)  # รูปแบบ RTYYYYMMDD001
    date = db.Column(db.Date, nullable=False, default=date.today)

    # FK ไปยัง customer และเอกสารอ้างอิง (QU/BK)
    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id"), nullable=False)
    quote_id    = db.Column(db.Integer, db.ForeignKey("sales_doc.id"), nullable=True)
    booking_id  = db.Column(db.Integer, db.ForeignKey("sales_doc.id"), nullable=True)

    ref_type = db.Column(db.String(2), default="QU")  # QU/BK

    remark = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer)  # user.id ที่สร้างใบคืน (ยังไม่ทำ FK ก็ได้)

    # relation เอาไว้ดึงชื่อไปโชว์
    customer = db.relationship("Customer", lazy="joined")
    quote    = db.relationship("SalesDoc", foreign_keys=[quote_id], lazy="joined")
    booking  = db.relationship("SalesDoc", foreign_keys=[booking_id], lazy="joined")

    items = db.relationship(
        "ReturnItem",
        back_populates="doc",
        cascade="all, delete-orphan",
        lazy="joined",
    )

    def __repr__(self) -> str:
        return f"<ReturnDoc {self.number}>"



class ReturnItem(db.Model):
    __tablename__ = "return_items"

    id = db.Column(db.Integer, primary_key=True)
    doc_id = db.Column(db.Integer, db.ForeignKey("return_docs.id"), nullable=False)
    # ✅ ใส่ ForeignKey มาที่ equipment.id
    equipment_id = db.Column(
        db.Integer,
        db.ForeignKey("equipment.id"),
        nullable=False,
    )
    qty = db.Column(db.Integer, nullable=False, default=1)

    # NEW: สภาพ/ผลลัพธ์หลังคืน
    condition = db.Column(db.String(16), default="GOOD")  # GOOD/REPAIR/LOST
    damage_note = db.Column(db.Text, default="")
    damage_cost = db.Column(db.Float, default=0.0)

    # ความสัมพันธ์
    doc = db.relationship("ReturnDoc", back_populates="items", lazy="joined")
    # ✅ ไม่ต้องกำหนด primaryjoin เอง ปล่อยให้ SQLAlchemy ใช้ FK
    equipment = db.relationship("Equipment", lazy="joined")

    def __repr__(self) -> str:
        return f"<ReturnItem doc={self.doc_id} eq={self.equipment_id}>"


# ---------- helpers ----------

def _find_spare_model():
    # ลองเดาชื่อคลาสที่เป็นไปได้
    for name in ("SparePart", "Spare", "SpareParts", "Sparepart", "Spares", "Part"):
        if name in globals():
            return globals()[name]
    # ลองเดาจาก __tablename__
    try:
        for cls in list(db.Model._decl_class_registry.values()):
            if hasattr(cls, "__tablename__") and cls.__tablename__:
                if "spare" in cls.__tablename__.lower():
                    return cls
    except Exception:
        pass
    return None

def _load_spares():
    """คืนลิสต์อะไหล่ที่มีฟิลด์มาตรฐาน: id, code, name, unit_price"""
    Model = _find_spare_model()
    if Model:
        q = Model.query
        if hasattr(Model, "is_active"):
            q = q.filter(Model.is_active == True)
        if hasattr(Model, "code"):
            q = q.order_by(Model.code.asc())

        rows = q.all()
        out = []
        for p in rows:
            # map ราคามาเป็น unit_price เสมอ (ถ้าไม่มี unit_price ให้ใช้ unit_cost)
            price = getattr(p, "unit_price", None)
            if price is None:
                price = getattr(p, "unit_cost", 0)
            out.append(
                types.SimpleNamespace(
                    id=getattr(p, "id"),
                    code=getattr(p, "code", ""),   # ถ้าในอนาคตใช้ชื่ออื่น ค่อยขยายตรงนี้
                    name=getattr(p, "name", ""),
                    unit_price=price,
                )
            )
        return out



    # 2) Fallback raw SQL: ลองหลายชื่อ table + รองรับชื่อคอลัมน์ price ต่างกัน
    table_candidates = ["spare_parts", "spares", "spare", "parts"]
    for tbl in table_candidates:
        try:
            rows = db.session.execute(text(f"""
                SELECT id,
                       COALESCE(code, sku, part_code)          AS code,
                       COALESCE(name, part_name, title)         AS name,
                       COALESCE(unit_price, price, unitcost, 0) AS unit_price,
                       COALESCE(is_active, 1)                   AS is_active
                FROM {tbl}
                WHERE COALESCE(is_active, 1)=1
                ORDER BY code
            """)).mappings().all()
            if rows:
                return [types.SimpleNamespace(
                    id=r.get("id"),
                    code=r.get("code"),
                    name=r.get("name"),
                    unit_price=r.get("unit_price"),
                ) for r in rows]
        except Exception:
            pass
    return []



# ---------- helpers for SKU resolving ----------
_INVIS = ["\u200b", "\ufeff", "\u2060", "\u00a0"]

def _norm_sku(s: str | None) -> str | None:
    """ตัดอักขระล่องหน/ช่องว่าง เก็บไว้เป็น SKU สะอาดๆ"""
    if not s: 
        return None
    s = str(s).strip()
    for ch in _INVIS:
        s = s.replace(ch, "")
    return s

def _extract_tokens_from_text(text: str | None) -> list[str]:
    """ดึงข้อความในวงเล็บเหลี่ยม [TOKEN] ทั้งหมด"""
    if not text:
        return []
    # อนุญาตตัวอักษร/ตัวเลข/._- ภายในวงเล็บ
    import re
    tokens = re.findall(r"\[([A-Za-z0-9_.\-]+)\]", text)
    return [_norm_sku(t) for t in tokens if _norm_sku(t)]


def _dec(x):
    try:
        return Decimal(str(x)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _sync_child_from_parent(parent: SalesDoc, child: SalesDoc):
    """Sync child doc items/totals from parent (used after BK allocation/approval).
    This fixes the case where BL/IV/RC were created before allocated_skus existed."""
    if not child:
        return
    # ล้างรายการเดิมของลูก
    SalesItem.query.filter_by(doc_id=child.id).delete(synchronize_session=False)
    db.session.flush()
    # คัดลอกรายการใหม่จาก parent -> child
    _clone_items(parent, child)
    # คัดลอกยอดรวม
    for f in ["amount_subtotal", "amount_vat", "amount_total", "amount_wht", "amount_grand"]:
        if hasattr(child, f) and hasattr(parent, f):
            setattr(child, f, getattr(parent, f) or 0.0)


def _ensure_children_for_booking(bk: SalesDoc):
    children = {c.doc_type: c for c in SalesDoc.query.filter_by(parent_id=bk.id).all()}

    # สร้างถ้ายังไม่มี
    if "BL" not in children:
        children["BL"] = _create_child_doc(bk, "BL", "UNPAID")
    if "IV" not in children:
        children["IV"] = _create_child_doc(bk, "IV", "UNISSUED")
    if "RC" not in children:
        children["RC"] = _create_child_doc(bk, "RC", "UNISSUED")

    # สำคัญ: sync รายการทุกครั้ง (เพราะ BK อาจเพิ่งถูก allocate/approve)
    for t in ["BL", "IV", "RC"]:
        _sync_child_from_parent(bk, children.get(t))



# relations
User.roles = db.relationship(Role, secondary="user_role", lazy="joined")
Role.perms = db.relationship(Permission, secondary="role_permission", lazy="joined")
User.perms = db.relationship(Permission, secondary="user_permission", lazy="joined")

# ================== Login manager ==================
login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "โปรดเข้าสู่ระบบก่อนใช้งาน"

@login_manager.user_loader
def load_user(user_id: str):
    return db.session.get(User, int(user_id))

# ================== Permission helpers ==================
def user_has_perm(user, perm_code: str) -> bool:
    if not user or not user.is_authenticated:
        return False

    # ----- ให้ admin เป็น superuser -----
    # 1) ถ้าชื่อผู้ใช้เป็น admin
    if getattr(user, "username", None) == "admin":
        return True

    # 2) ถ้ามี field user.role แล้วเป็น admin (เผื่อใช้แบบเก่า)
    if getattr(user, "role", None) == "admin":
        return True

    # 3) ถ้ามี role object ที่ code == 'admin'
    if any(r.code == "admin" for r in getattr(user, "roles", [])):
        return True

    # ----- ตรวจสิทธิ์ตาม perm ปกติ -----
    # perm ตรง ๆ ผูกกับ user
    if any(p.code == perm_code for p in getattr(user, "perms", [])):
        return True

    # perm ผ่าน role ต่าง ๆ
    for r in getattr(user, "roles", []):
        if any(p.code == perm_code for p in getattr(r, "perms", [])):
            return True

    return False


def permission_required(perm_code: str):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))
            if not user_has_perm(current_user, perm_code):
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco

def _booking_flow_ready(bk: "SalesDoc") -> bool:
    """พร้อมปิดงานหรือยัง (ใช้กับ UI)
    เงื่อนไข: ต้องมี BL/IV/RC ลูก และ BL=PAID, IV=ISSUED, RC=ISSUED
    """
    try:
        if not bk or (bk.doc_type or "").upper() != "BK":
            return False
        kids = SalesDoc.query.filter_by(parent_id=bk.id).all()
        m = { (c.doc_type or "").upper(): c for c in kids }
        bl = m.get("BL")
        iv = m.get("IV")
        rc = m.get("RC")
        if not (bl and iv and rc):
            return False
        ok_bl = (bl.status or "").upper() == "PAID"
        ok_iv = (iv.status or "").upper() == "ISSUED"
        ok_rc = (rc.status or "").upper() == "ISSUED"
        return bool(ok_bl and ok_iv and ok_rc)
    except Exception:
        return False


def _update_booking_flow_state(bk: "SalesDoc") -> None:
    """อัพเดทสถานะ flow ของ BK (optional)
    - ไม่ปิดงานอัตโนมัติ (เพราะเราทำ manual close แล้ว)
    - แต่สามารถใช้ตั้ง flag/field เพิ่มในอนาคตได้
    ตอนนี้ให้ทำแบบ safe no-op + future-proof
    """
    try:
        # ยังไม่บังคับเปลี่ยน status ของ BK อัตโนมัติ
        # แค่อ่านเพื่อให้แน่ใจว่าเด็กครบและสถานะถูกต้อง
        _ = _booking_flow_ready(bk)
    except Exception:
        pass


@app.context_processor
def inject_perms():
    def _can(code):
        # ถ้ายังไม่ล็อกอิน → ไม่มีสิทธิ์
        if not current_user.is_authenticated:
            return False

        # ให้ admin เห็นทุกเมนู (superuser)
        if getattr(current_user, "username", None) == "admin":
            return True

        # ปกติใช้ฟังก์ชันเดิมเช็คสิทธิ์
        return user_has_perm(current_user, code)

    return {"can": _can}

def _unit_to_days(unit: str, n: int | float) -> int:
    """แปลงหน่วยเช่า + จำนวนหน่วย → จำนวนวัน (ปัดขึ้นอย่างปลอดภัย)"""
    u = (unit or "DAY").upper()
    n = max(0, float(n or 0))
    if u == "MONTH":
        return int(round(n * 30))
    if u == "YEAR":
        return int(round(n * 365))
    # ถ้าจะรองรับ HOUR ภายหลัง คูณ 1/24 เพิ่มได้
    return int(round(n))  # DAY


def compute_promo_discount(items: List[Dict], rental_days: int | None, promo: Promotion) -> float:
    """
    items:
      - name
      - qty
      - unit_price_per_day  (ถ้าไม่มี จะคำนวณจาก unit_price + rent_unit + rent_duration)
      - unit_price          (ราคา/หน่วยเช่า 1 ชิ้น)
      - rent_unit           ("DAY"/"MONTH"/"YEAR")
      - rent_duration       (จำนวนหน่วยเช่า)
    promo: ใช้ promo.rental_unit & promo.min_duration เป็นเงื่อนไขขั้นต่ำ
    """
    if not promo or not promo.is_in_effect(date.today()):
        return 0.0

    # ---- ผูกโปรกับหน่วยเช่าให้ตรงกัน (รายวัน / รายเดือน / ปี) ----
    promo_unit = (promo.rental_unit or "DAY").upper()
    bill_units: set[str] = set()

    for it in (items or []):
        qty = int(it.get("qty", 0) or 0)
        if qty <= 0:
            continue
        ru_item = (it.get("rent_unit") or "DAY").upper()
        bill_units.add(ru_item)

    # ถ้าบิลมีหน่วยเช่าจริง แต่ไม่มีหน่วยที่ตรงกับโปรเลย → ไม่ให้ใช้โปรนี้
    if bill_units and promo_unit not in bill_units:
        return 0.0

    # ------------- เตรียมข้อมูล -------------
    # แปลงราคา/วัน และจำนวนวันรวมจากรายการจริง
    normalized = []
    days_each_row = []

    for it in (items or []):
        qty = int(it.get("qty", 0) or 0)
        if qty <= 0:
            continue

        # ราคา/วัน
        if it.get("unit_price_per_day") is not None:
            ppd = float(it["unit_price_per_day"] or 0.0)
        else:
            unit_price = float(it.get("unit_price", 0.0) or 0.0)
            ru = (it.get("rent_unit") or "DAY").upper()
            rd = int(it.get("rent_duration", 1) or 1)
            days = _unit_to_days(ru, 1) or 1
            ppd = unit_price / days  # ราคา/วัน (ของ 1 หน่วยเช่า)

        # จำนวนวัน “ที่เช่าจริง” ของแถวนี้
        ru = (it.get("rent_unit") or "DAY").upper()
        rd = it.get("rent_duration", 1) or 1
        d_this = max(1, _unit_to_days(ru, rd))

        normalized.append({"ppd": ppd, "qty": qty, "days": d_this})
        days_each_row.append(d_this)

    if not normalized:
        return 0.0

    # ------------- เงื่อนไขขั้นต่ำ -------------
    total_qty = sum(r["qty"] for r in normalized)
    if total_qty < (promo.min_items or 0):
        return 0.0

    # จำนวนวันจริงของบิล (ถ้า caller ไม่ส่งมา ให้ยึด “น้อยที่สุด” ของทุกแถว)
    # เหตุผล: ถ้าแถวใดเช่าสั้นกว่า ก็ไม่ควรใช้วันของแถวที่ยาวกว่ามา unlock โปรฯ
    if rental_days is None:
        rental_days = min(days_each_row)

    # แปลงเงื่อนไขขั้นต่ำของโปรฯ → วัน
    min_days_required = _unit_to_days(promo.rental_unit, promo.min_duration or 0)
    if rental_days < min_days_required:
        return 0.0

    # ------------- คำนวณส่วนลด -------------
    # โปรฯ “ลดชิ้นถูกสุด K ชิ้น” *ตามจำนวนวันจริง*
    units = []
    for r in normalized:
        for _ in range(r["qty"]):
            units.append(r["ppd"])
    units.sort()  # จากถูก → แพง

    k = max(1, int(promo.cheapest_units_to_discount or 1))
    k = min(k, len(units))
    base_amount = sum(units[:k]) * rental_days

    if promo.discount_type == "PCT":
        disc = base_amount * (float(promo.discount_value or 0) / 100.0)
    else:
        disc = float(promo.discount_value or 0)

    return max(0.0, min(disc, base_amount))




def _items_from_doc(d: SalesDoc) -> list[dict]:
    rows = []
    for it in d.items:
        rows.append({
            "qty": int(it.qty or 0),
            "unit_price": float(it.unit_price or 0),
            "rent_unit": (it.rent_unit or "DAY").upper(),
            "rent_duration": int(it.rent_duration or 1),
        })
    return rows


# ---------- helpers กันพลาดตอนลบ ----------
def _has_other_admin(exclude_uid: int) -> bool:
    q = (
        select(UserRole)
        .join(Role, UserRole.role_id == Role.id)
        .where(Role.code == "admin", UserRole.user_id != exclude_uid)
        .limit(1)
    )
    return db.session.execute(q).first() is not None

# ================== Bootstrap (Flask 3.x compatible) ==================
def bootstrap():
    with app.app_context():
        db.create_all()
        os.makedirs(os.path.join(app.static_folder, "uploads", "company"), exist_ok=True)
        get_company()

        # ---- seed roles ----
        role_defs = [
            ("admin", "ผู้ดูแลระบบ"),
            ("manager", "ผู้จัดการ"),
            ("sales", "เซลส์"),
            ("purchasing", "จัดซื้อ"),
            ("warehouse", "คลัง/สโตร์"),
            ("delivery", "ขนส่ง"),
            ("accounting", "บัญชี"),
        ]
        codes_to_role = {}
        for code, name in role_defs:
            r = Role.query.filter_by(code=code).first()
            if not r:
                r = Role(code=code, name=name)
                db.session.add(r)
            codes_to_role[code] = r
        db.session.commit()

        # ---- seed permissions (ตัด maintenance.*, repairs.* ออก) ----
        perm_defs = [
    ("dashboard.view", "ดูแดชบอร์ด"),
    ("users.manage", "จัดการผู้ใช้/สิทธิ์"),

    ("purchases.view",   "ดูเอกสารซื้อ"),
    ("purchases.create", "สร้าง/แก้ไขใบสั่งซื้อ"),
    ("goods.receive",    "รับสินค้า (GRN)"),

    ("company.manage", "ตั้งค่าบริษัท"),

    ("customers.view",   "ดูรายชื่อลูกค้า"),
    ("customers.manage", "จัดการลูกค้า"),

    ("equipment.view",   "ดูอุปกรณ์"),
    ("equipment.manage", "จัดการอุปกรณ์/หมวดหมู่"),

    ("promos.view",   "ดูโปรโมชั่น"),
    ("promos.manage", "จัดการโปรโมชั่น"),

    ("sales.view",   "ดูเอกสารขาย"),
    ("sales.manage", "สร้าง/แก้ไข/อนุมัติ เอกสารขาย"),

    ("claims.view",   "ดูงานเคลม"),
    ("claims.manage", "จัดการงานเคลม"),

    # Spares
    ("spares.view",   "ดูรายการอะไหล่"),
    ("spares.create", "เพิ่มอะไหล่"),
    ("spares.edit",   "แก้ไขอะไหล่"),
    ("spares.delete", "ลบอะไหล่"),

    ("repairs.view",   "ดูงานซ่อม"),
    ("repairs.manage", "จัดการงานซ่อม"),

    ("transport.access",        "เข้าถึงเมนูงานขนส่ง"),
    ("transport.manage",        "บริหารรถ/คนขับ/จัดสายรถ"),
    ("transport.update_status", "อัปเดตสถานะเอกสารขนส่ง"),

    ("gifts.view",   "ดูเมนูของขวัญ"),
    ("gifts.manage", "จัดการแคมเปญของขวัญ"),
]

        codes_to_perm = {}
        for code, name in perm_defs:
            p = Permission.query.filter_by(code=code).first()
            if not p:
                p = Permission(code=code, name=name)
                db.session.add(p)
            codes_to_perm[code] = p
        db.session.commit()

        # ---- create default admin user ----
        admin_u = User.query.filter_by(username="admin").first()
        if not admin_u:
            admin_u = User(
                username="admin",
                full_name="Administrator",
                password_hash=generate_password_hash("admin123"),
                is_active=True,
            )
            db.session.add(admin_u)
            db.session.commit()

        # bind admin role
        admin_role = codes_to_role["admin"]
        if admin_role not in admin_u.roles:
            db.session.add(UserRole(user_id=admin_u.id, role_id=admin_role.id))
            db.session.commit()

        def _grant_role_perm(role_code: str, perm_code: str):
            r = codes_to_role.get(role_code)
            p = codes_to_perm.get(perm_code)
            if not r or not p:
                return
            exists = RolePermission.query.filter_by(role_id=r.id, perm_id=p.id).first()
            if not exists:
                db.session.add(RolePermission(role_id=r.id, perm_id=p.id))

        # ---- GRANTS ----
        # Purchases / Warehouse
        _grant_role_perm("purchasing", "purchases.view")
        _grant_role_perm("purchasing", "purchases.create")
        _grant_role_perm("warehouse", "purchases.view")
        _grant_role_perm("warehouse", "goods.receive")
        _grant_role_perm("manager", "purchases.view")

        # Company / Customers
        _grant_role_perm("admin", "company.manage")
        _grant_role_perm("manager", "company.manage")
        for rc in ("admin","manager","sales","accounting"):
            _grant_role_perm(rc, "customers.view")
        for rc in ("admin","manager","sales"):
            _grant_role_perm(rc, "customers.manage")

        # Equipment
        for rc in ("admin","manager","warehouse"):
            _grant_role_perm(rc, "equipment.view")
            _grant_role_perm(rc, "equipment.manage")
        _grant_role_perm("sales", "equipment.view")

        # Dashboard
        for rc in ("admin","manager","sales","purchasing","warehouse","delivery","accounting"):
            _grant_role_perm(rc, "dashboard.view")

        # Promotions
        for rc in ("admin","manager","sales"):
            _grant_role_perm(rc, "promos.view")
        for rc in ("admin","manager"):
            _grant_role_perm(rc, "promos.manage")

        # Sales
        for rc in ("admin","manager","sales","accounting"):
            _grant_role_perm(rc, "sales.view")
        for rc in ("admin","manager","sales"):
            _grant_role_perm(rc, "sales.manage")

        # Claims
        for rc in ("admin","manager","sales","warehouse","accounting"):
            _grant_role_perm(rc, "claims.view")
        for rc in ("admin","manager","sales"):
            _grant_role_perm(rc, "claims.manage")

        # Spares
        for rc in ("admin","manager","warehouse"):
            _grant_role_perm(rc, "spares.view")
        for rc in ("admin","manager"):
            _grant_role_perm(rc, "spares.create")
            _grant_role_perm(rc, "spares.edit")
            _grant_role_perm(rc, "spares.delete")

        # Repairs permissions
        for rc in ("admin","manager","warehouse"):
            _grant_role_perm(rc, "repairs.view")
        for rc in ("admin","manager"):
            _grant_role_perm(rc, "repairs.manage")

        # Transport
        for rc in ("admin","manager","delivery"):
            _grant_role_perm(rc, "transport.access")
        for rc in ("admin","manager"):
            _grant_role_perm(rc, "transport.manage")
        for rc in ("admin","manager","delivery"):
            _grant_role_perm(rc, "transport.update_status")

         # Gifts / ของขวัญ  👇 เพิ่มส่วนนี้
        for rc in ("admin", "manager", "sales"):
            _grant_role_perm(rc, "gifts.view")
        for rc in ("admin", "manager"):
            _grant_role_perm(rc, "gifts.manage")


        db.session.commit()
    


# ---------- Uploads: Equipment ----------
UPLOAD_EQUIP_DIR = os.path.join(app.static_folder, "uploads", "equipment")
ALLOWED_IMG = {".png", ".jpg", ".jpeg", ".webp"}
MAX_IMG_MB = 5

def _save_image(file_storage, filename_stub: str) -> str:
    os.makedirs(UPLOAD_EQUIP_DIR, exist_ok=True)
    ext = os.path.splitext(file_storage.filename.lower())[1]
    if ext not in ALLOWED_IMG:
        raise ValueError("รองรับเฉพาะ PNG/JPG/JPEG/WEBP")
    file_storage.seek(0, os.SEEK_END)
    mb = file_storage.tell()/(1024*1024)
    file_storage.seek(0)
    if mb > MAX_IMG_MB:
        raise ValueError(f"ไฟล์ใหญ่เกิน {MAX_IMG_MB}MB")
    fname = secure_filename(f"{filename_stub}{ext}")
    file_storage.save(os.path.join(UPLOAD_EQUIP_DIR, fname))
    return f"uploads/equipment/{fname}"

def gen_sku(prefix: str, dt: date) -> str:
    """Generate Equipment SKU (Requirement 2.7)

    รูปแบบ:
      SPT{CAT}{YY}{MM}-{SEQ3}
      - CAT = รหัสหมวด (Category.prefix_sku)
      - YY  = ปี 2 หลัก
      - MM  = เดือน 2 หลัก
      - SEQ = running number 3 หลัก ภายในกลุ่ม (SPT{CAT}{YY}{MM})

    ตัวอย่าง:
      SPTE6901-001

    หมายเหตุ:
      - ถ้า prefix_sku เดิมกรอกมาเป็น 'SPTE' หรือมีขีด/ช่องว่าง เราจะ normalize ให้เป็น 'SPTE'
      - SKU เก่าที่มีอยู่แล้วในระบบจะไม่ถูกเปลี่ยน
    """
    raw = (prefix or "").strip().upper()
    raw = re.sub(r"[^A-Z0-9]", "", raw)

    if not raw.startswith("SPT"):
        raw = "SPT" + raw

    base = f"{raw}{dt.strftime('%y%m')}"  # SPT{CAT}{YY}{MM}
    like = f"{base}-%"

    last = (Equipment.query
            .filter(Equipment.sku.like(like))
            .order_by(Equipment.sku.desc())
            .first())

    seq = 1
    if last and last.sku:
        mm = re.search(r"-(\d{3,})$", last.sku)
        if mm:
            try:
                seq = int(mm.group(1)) + 1
            except Exception:
                seq = 1

    return f"{base}-{seq:03d}"

def _gen_sales_running(prefix: str) -> str:
    today_s = date.today().strftime("%Y%m%d")
    like = f"{prefix}{today_s}%"
    last = SalesDoc.query.filter(
        SalesDoc.number.like(like)
    ).order_by(SalesDoc.number.desc()).first()
    seq = 1
    if last and last.number[-4:].isdigit():
        seq = int(last.number[-4:]) + 1
    return f"{prefix}{today_s}{seq:04d}"

def _calc_sales_totals(doc: SalesDoc):
    sub = 0.0
    for it in doc.items:
        gross = (it.qty or 0) * (it.rent_duration or 0) * (it.unit_price or 0)
        disc = gross * (max(0.0, it.discount_pct or 0)/100.0)
        it.line_subtotal = round(gross, 2)
        it.line_total = round(gross - disc, 2)
        sub += it.line_total
    doc.amount_subtotal = round(sub, 2)
    if doc.tax_mode == "EXC":
        vat = round(sub * 0.07, 2)
        total = round(sub + vat, 2)
    elif doc.tax_mode == "INC":
        vat = round(sub * (7/107), 2)
        total = round(sub, 2)
    else:
        vat = 0.0
        total = round(sub, 2)
    doc.amount_vat = vat
    doc.amount_total = total
    # ✅ WHT ต้องคิดจากฐาน "ก่อน VAT" (Requirement)
    wht_base = sub
    if doc.tax_mode == "INC":
        # sub/total เป็นยอดรวม VAT แล้ว → ฐานก่อน VAT = total - vat
        wht_base = max(0.0, total - vat)
    wht = round(wht_base * (max(0, doc.wht_pct or 0)/100.0), 2)
    doc.amount_wht = wht
    # ยอดสุทธิหลังหัก ณ ที่จ่าย (รวม VAT แล้วค่อยหัก WHT)
    doc.amount_grand = round(total - wht, 2)

def _first_nonempty(obj, names):
    for n in names:
        if not obj:
            continue
        val = getattr(obj, n, None)
        if isinstance(val, str) and val.strip():
            return val.strip()
        if isinstance(val, (list, tuple)) and val:
            for v in val:
                if isinstance(v, str) and v.strip():
                    return v.strip()
    return None

def _as_static_url(path: str) -> str:
    if not path:
        return None
    scheme = urlparse(path).scheme.lower()
    if scheme in ("http", "https", "data"):
        return path
    p = path.lstrip("/")
    if p.startswith("static/"):
        p = p[7:]
    return url_for("static", filename=p)

def _gen_running(prefix: str, Model):
    today = datetime.utcnow()
    yyyymm = today.strftime("%Y%m")
    like = f"{prefix}-{yyyymm}-%"
    last = (
        db.session.query(Model)
        .filter(Model.number.like(like))
        .order_by(Model.id.desc())
        .first()
    )
    seq = 1
    if last:
        try:
            seq = int(last.number.split("-")[-1]) + 1
        except Exception:
            seq = last.id + 1
    return f"{prefix}-{yyyymm}-{seq:03d}"



# ===== helper: ดึงรายการไอเท็มของใบเคลมแบบยืดหยุ่น =====
def _claim_items_of(c):
    """
    คืน list ของไอเท็มในใบเคลม รองรับหลายชื่อความสัมพันธ์:
    - c.items
    - c.claim_items
    - c.lines
    ถ้าไม่มีสักอย่าง จะ fallback ไป query ClaimItem โดย claim_id
    """
    for attr in ("items", "claim_items", "lines"):
        if hasattr(c, attr):
            items = getattr(c, attr) or []
            if items:
                return items
    try:
        return ClaimItem.query.filter_by(claim_id=c.id).all()
    except Exception:
        return []

def _get_num(x, *names, default=1):
    for n in names:
        if hasattr(x, n):
            v = getattr(x, n)
            if v is not None:
                return v
        if isinstance(x, dict) and n in x:
            return x[n]
    return default

def _get_str(x, *names):
    for n in names:
        if hasattr(x, n):
            v = getattr(x, n)
            if v:
                return str(v)
        if isinstance(x, dict) and n in x and x[n]:
            return str(x[n])
    return ""




# โฟลเดอร์เก็บรูปใบส่ง: static/uploads/delivery/
DELIVERY_PHOTO_SUBDIR = os.path.join("uploads", "delivery")


def _save_delivery_photos(files, doc, kind: str) -> int:
    """
    บันทึกรูปใบส่งสินค้า
    kind: 'BEFORE' หรือ 'AFTER'
    return: จำนวนรูปที่เซฟได้
    """
    from werkzeug.utils import secure_filename

    base_dir = os.path.join(app.root_path, "static", DELIVERY_PHOTO_SUBDIR)
    os.makedirs(base_dir, exist_ok=True)

    saved = 0
    for f in files:
        if not f or not getattr(f, "filename", None):
            continue
        fname = secure_filename(f.filename)
        if not fname:
            continue

        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
        fname = f"{doc.number}_{kind.lower()}_{ts}_{fname}"
        full_path = os.path.join(base_dir, fname)

        f.save(full_path)
        db.session.add(DeliveryPhoto(doc_id=doc.id, kind=kind, filename=fname))
        saved += 1

    return saved



def _doc_amount(d):
    """
    คืนยอดรวมของเอกสารขายแบบปลอดภัย
    ใช้ลำดับเดียวกับ dashboard
    """
    for attr in (
        "amount_grand",
        "amount_total",
        "grand_total",
        "total_amount",
        "amount_subtotal",
    ):
        if hasattr(d, attr):
            try:
                return float(getattr(d, attr) or 0.0)
            except Exception:
                continue
    return 0.0




def recalc_gift_results(campaign: "GiftCampaign"):
    """
    คำนวณ GiftResult สำหรับแคมเปญที่กำหนด จากยอดขายลูกค้าในช่วง period_start–period_end
    - ใช้เอกสาร RC เป็นหลัก (กันยอดซ้ำ)
    - นับสถานะเดียวกับที่ Dashboard ใช้ (PAID/DONE/RECEIPTED/ISSUED ฯลฯ)
    - ถ้าไม่พบ RC เลยในช่วงนั้น ค่อย fallback ไปนับ BL ที่ PAID
    """
    from collections import defaultdict
    from sqlalchemy.orm import joinedload

    # 1) ดึง tier มาจัดเรียงจาก min_amount มาก -> น้อย
    tiers = sorted(campaign.tiers, key=lambda t: t.min_amount, reverse=True)
    if not tiers:
        return  # ไม่มี tier ก็ไม่ต้องทำอะไร

    # 2) สถานะ RC ที่ถือว่า "ยอดขายนับได้" (อิง Dashboard)
    rc_statuses = [
        "PAID", "Paid", "paid",
        "PAID_FULL", "Paid_Full", "paid_full",
        "DONE", "Done", "done",
        "RECEIPTED", "Receipted", "receipted",
        "ISSUED", "Issued", "issued",
    ]

    # 3) ดึงเอกสาร RC ในช่วงแคมเปญ
    rc_docs = (
        SalesDoc.query
        .options(joinedload(SalesDoc.customer))
        .filter(
            SalesDoc.doc_type == "RC",
            SalesDoc.status.in_(rc_statuses),
            SalesDoc.date.between(campaign.period_start, campaign.period_end),
        )
        .all()
    )

    # 4) รวมยอดตามลูกค้า
    customer_totals = defaultdict(float)

    if rc_docs:
        # ใช้ RC เป็นหลัก
        for d in rc_docs:
            if not d.customer_id:
                continue
            customer_totals[d.customer_id] += float(_doc_amount(d) or 0.0)
    else:
        # fallback: ถ้าไม่มี RC เลย ให้ใช้ BL ที่ PAID
        bl_statuses = ["PAID", "Paid", "paid", "DONE", "Done", "done"]
        bl_docs = (
            SalesDoc.query
            .options(joinedload(SalesDoc.customer))
            .filter(
                SalesDoc.doc_type == "BL",
                SalesDoc.status.in_(bl_statuses),
                SalesDoc.date.between(campaign.period_start, campaign.period_end),
            )
            .all()
        )
        for d in bl_docs:
            if not d.customer_id:
                continue
            customer_totals[d.customer_id] += float(_doc_amount(d) or 0.0)

    # 5) โหลดผลเดิมไว้ เพื่อไม่รีเซ็ตคนที่กด "ให้ของขวัญแล้ว"
    existing = {
        r.customer_id: r
        for r in GiftResult.query.filter_by(campaign_id=campaign.id).all()
    }

    # 6) สร้าง/อัปเดต GiftResult ตาม tier
    for cust_id, total in customer_totals.items():
        total = round(float(total or 0.0), 2)

        matched_tier = None
        for t in tiers:
            if total >= float(t.min_amount or 0.0):
                matched_tier = t
                break
        if not matched_tier:
            continue  # ยอดไม่ถึงเกณฑ์ใดเลย

        gr = existing.get(cust_id)
        if not gr:
            gr = GiftResult(
                campaign_id=campaign.id,
                customer_id=cust_id,
                status="PENDING",
            )
            db.session.add(gr)

        gr.total_amount = total
        gr.tier_code = matched_tier.code
        gr.tier_name = matched_tier.name
        # ไม่แตะ status/given_at เพื่อไม่รีเซ็ตคนที่ติ๊กว่าให้ของขวัญแล้ว

    db.session.commit()


def _create_booking_from_quote(qu: SalesDoc) -> SalesDoc:
    bk = SalesDoc(
        number=_gen_running("BK", SalesDoc),
        doc_type="BK",
        status="DRAFT",
        customer_id=qu.customer_id,
        project_name=qu.project_name,
        po_customer=qu.po_customer,
        credit_days=qu.credit_days or 0,
        tax_mode=qu.tax_mode,
        wht_pct=qu.wht_pct or 0,
        date=date.today(),
        remark=qu.remark,
        parent=qu,
        amount_subtotal=qu.amount_subtotal or 0.0,
        amount_vat=qu.amount_vat or 0.0,
        amount_total=qu.amount_total or 0.0,
        amount_wht=qu.amount_wht or 0.0,
        amount_grand=qu.amount_grand or 0.0,
    )
    db.session.add(bk)
    db.session.flush()

    # clone items จาก QU มาลง BK ก่อน (คงหมวด + brand เพื่อใช้ reserve/allocate)
    for it in qu.items:
        # ✅ สร้าง BK เฉพาะรายการที่อนุมัติ (Requirement 3.4)
        if ((getattr(it, 'line_status', 'APPROVED') or 'APPROVED').upper() == 'REJECTED'):
            continue
        db.session.add(SalesItem(
            doc_id=bk.id,
            source_qu_item_id=it.id,
            image_path=it.image_path,
            name=it.name,
            category_id=it.category_id,
            category_prefix=it.category_prefix,
            brand=it.brand,
            qty=it.qty,
            rent_unit=it.rent_unit,
            rent_duration=it.rent_duration,
            unit_price=it.unit_price,
            discount_pct=it.discount_pct,
            line_subtotal=it.line_subtotal,
            line_total=it.line_total,
        ))

    
    # โอน RESERVED จาก QU -> BK (กันค้างและกันจองซ้ำ)
    # - ลบ reservation ของ QU
    # - สร้าง reservation ใหม่ของ BK ตามรายการที่ clone มา
    _release_reservations_for_doc("QU", qu.id)
    db.session.flush()
    for it in SalesItem.query.filter_by(doc_id=bk.id).all():
        if it.category_id:
            _check_and_reserve("BK", bk.id, it.id, it.category_id, it.brand, bk.warehouse, it.qty or 0)

    return bk






# ================== Contract / Installment Helpers ==================

def _add_months(dt: date, months: int) -> date:
    """เพิ่มเดือนแบบปลอดภัย (ไม่ใช้ external lib)"""
    y = dt.year + (dt.month - 1 + months) // 12
    m = (dt.month - 1 + months) % 12 + 1
    # clamp day
    d = dt.day
    # days in month
    import calendar
    last_day = calendar.monthrange(y, m)[1]
    d = min(d, last_day)
    return date(y, m, d)

def _split_amount(total: float, n: int, decimals: int = 2) -> list[float]:
    """แบ่งยอด total ออกเป็น n งวด โดยปรับเศษให้ไปอยู่ที่งวดสุดท้าย"""
    if n <= 0:
        return []
    q = round(float(total) / n, decimals)
    arr = [q for _ in range(n)]
    s = round(sum(arr), decimals)
    diff = round(float(total) - s, decimals)
    if arr:
        arr[-1] = round(arr[-1] + diff, decimals)
    return arr

def _create_contract_from_quote(qu: SalesDoc) -> SalesDoc:
    """สร้างเอกสารสัญญา/PO ใหญ่ (CT) จาก QU ที่อนุมัติแล้ว"""
    # กันซ้ำ
    existing = SalesDoc.query.filter_by(parent_id=qu.id, doc_type="CT").first()
    if existing:
        return existing

    start_dt = qu.contract_start or qu.date or date.today()
    n = int(qu.installment_count or 0)
    if n <= 0:
        # infer: ถ้ามีรายการเดือน/ปี ให้แปลงเป็นจำนวนเดือน
        n = 12
        try:
            for it in (qu.items or []):
                if (it.rent_unit or "").upper() == "MONTH":
                    n = max(n, int(it.rent_duration or 1))
                elif (it.rent_unit or "").upper() == "YEAR":
                    n = max(n, int(it.rent_duration or 1) * 12)
        except Exception:
            pass
    end_dt = qu.contract_end
    if not end_dt:
        # end = start + n months - 1 day
        end_dt = _add_months(start_dt, n) - timedelta(days=1)

    ct = SalesDoc(
        number=_gen_sales_running("CT"),
        doc_type="CT",
        status="ACTIVE",
        customer_id=qu.customer_id,
        project_name=qu.project_name,
        po_customer=qu.po_customer,
        credit_days=qu.credit_days or 0,
        tax_mode=qu.tax_mode,
        wht_pct=qu.wht_pct or 0,
        date=start_dt,
        remark=(qu.remark or "").strip(),
        parent_id=qu.id,
        billing_mode="INSTALLMENT",
        contract_start=start_dt,
        contract_end=end_dt,
        installment_count=n,
    )
    db.session.add(ct)
    db.session.flush()

    # สร้างตารางงวด
    _ensure_installments_for_contract(ct, qu)

    return ct

def _ensure_installments_for_contract(ct: SalesDoc, qu: SalesDoc | None = None):
    """สร้างงวด (ถ้ายังไม่มี)"""
    existing = SalesInstallment.query.filter_by(contract_id=ct.id).count()
    if existing > 0:
        return

    n = int(ct.installment_count or 0) or 12
    start_dt = ct.contract_start or ct.date or date.today()

    # split amounts
    if qu is None:
        qu = SalesDoc.query.options(joinedload(SalesDoc.items)).get(ct.parent_id) if ct.parent_id else None

    sub_list = _split_amount(qu.amount_subtotal if qu else ct.amount_subtotal, n)
    vat_list = _split_amount(qu.amount_vat if qu else ct.amount_vat, n)
    total_list = _split_amount(qu.amount_total if qu else ct.amount_total, n)
    wht_list = _split_amount(qu.amount_wht if qu else ct.amount_wht, n)
    grand_list = _split_amount(qu.amount_grand if qu else ct.amount_grand, n)

    for i in range(1, n + 1):
        period_start = _add_months(start_dt, i - 1)
        period_end = _add_months(start_dt, i) - timedelta(days=1)
        bill_date = period_start
        due_date = bill_date + timedelta(days=int(ct.credit_days or 0))

        inst = SalesInstallment(
            contract_id=ct.id,
            installment_no=i,
            period_start=period_start,
            period_end=period_end,
            bill_date=bill_date,
            due_date=due_date,
            status="PLANNED",
            amount_subtotal=sub_list[i-1] if i-1 < len(sub_list) else 0.0,
            amount_vat=vat_list[i-1] if i-1 < len(vat_list) else 0.0,
            amount_total=total_list[i-1] if i-1 < len(total_list) else 0.0,
            amount_wht=wht_list[i-1] if i-1 < len(wht_list) else 0.0,
            amount_grand=grand_list[i-1] if i-1 < len(grand_list) else 0.0,
        )
        db.session.add(inst)

def _create_docs_for_installment(ct: SalesDoc, inst: SalesInstallment) -> tuple[SalesDoc, SalesDoc, SalesDoc]:
    """สร้าง BL/IV/RC จากงวด (กันซ้ำ)"""
    # ถ้ามีแล้ว คืนของเดิม
    if inst.bill_id and inst.invoice_id and inst.receipt_id:
        bl = SalesDoc.query.get(inst.bill_id)
        iv = SalesDoc.query.get(inst.invoice_id)
        rc = SalesDoc.query.get(inst.receipt_id)
        return bl, iv, rc

    # ใช้โครงสร้าง child docs ที่มีอยู่แล้ว (BL/IV/RC) แต่สร้างจาก CT เป็น parent
    # หมายเหตุ: ใช้ _create_child_doc เพื่อ clone items เหมือน QU
    bl = None
    iv = None
    rc = None

    def _child_number(prefix: str) -> str:
        return _gen_sales_running(prefix)

    # สร้าง BL
    if not inst.bill_id:
        bl = SalesDoc(
            number=_child_number("BL"),
            doc_type="BL",
            status="UNPAID",
            customer_id=ct.customer_id,
            project_name=ct.project_name,
            po_customer=(inst.po_customer_sub or ct.po_customer or ""),
            credit_days=ct.credit_days or 0,
            tax_mode=ct.tax_mode,
            wht_pct=ct.wht_pct or 0,
            date=inst.bill_date,
            remark=f"(งวดที่ {inst.installment_no}) {ct.number} : {inst.period_start.strftime('%d/%m/%Y')} - {inst.period_end.strftime('%d/%m/%Y')}",
            parent_id=ct.id,
            billing_mode="ONCE",
        )
        db.session.add(bl)
        db.session.flush()

        # clone items from CT parent quote (CT has no items) -> clone from QU
        src = None
        if ct.parent_id:
            # ถ้ามีใบจอง (BK) ให้เอารายการจาก BK เพื่อพก allocated_skus / หมวดอุปกรณ์
            src = (SalesDoc.query.options(joinedload(SalesDoc.items))
                   .filter_by(parent_id=ct.parent_id, doc_type='BK')
                   .first())
            if not src:
                # fallback: เอารายการจากใบเสนอราคา (QU)
                src = SalesDoc.query.options(joinedload(SalesDoc.items)).get(ct.parent_id)
        if src:
            _clone_items(src, bl)
        # override amounts to installment
        bl.amount_subtotal = inst.amount_subtotal
        bl.amount_vat = inst.amount_vat
        bl.amount_total = inst.amount_total
        bl.amount_wht = inst.amount_wht
        bl.amount_grand = inst.amount_grand
        inst.bill_id = bl.id

    # สร้าง IV
    if not inst.invoice_id:
        # ต้อง flush ก่อนเพื่อมี bl
        if not bl and inst.bill_id:
            bl = SalesDoc.query.get(inst.bill_id)
        iv = SalesDoc(
            number=_child_number("IV"),
            doc_type="IV",
            status="UNISSUED",
            customer_id=ct.customer_id,
            project_name=ct.project_name,
            po_customer=(inst.po_customer_sub or ct.po_customer or ""),
            credit_days=ct.credit_days or 0,
            tax_mode=ct.tax_mode,
            wht_pct=ct.wht_pct or 0,
            date=inst.bill_date,
            remark=f"(งวดที่ {inst.installment_no}) {ct.number} : {inst.period_start.strftime('%d/%m/%Y')} - {inst.period_end.strftime('%d/%m/%Y')}",
            parent_id=ct.id,
            billing_mode="ONCE",
        )
        db.session.add(iv)
        db.session.flush()
        src = None
        if ct.parent_id:
            # ถ้ามีใบจอง (BK) ให้เอารายการจาก BK เพื่อพก allocated_skus / หมวดอุปกรณ์
            src = (SalesDoc.query.options(joinedload(SalesDoc.items))
                   .filter_by(parent_id=ct.parent_id, doc_type='BK')
                   .first())
            if not src:
                # fallback: เอารายการจากใบเสนอราคา (QU)
                src = SalesDoc.query.options(joinedload(SalesDoc.items)).get(ct.parent_id)
        if src:
            _clone_items(src, iv)
        iv.amount_subtotal = inst.amount_subtotal
        iv.amount_vat = inst.amount_vat
        iv.amount_total = inst.amount_total
        iv.amount_wht = inst.amount_wht
        iv.amount_grand = inst.amount_grand
        inst.invoice_id = iv.id

    # สร้าง RC
    if not inst.receipt_id:
        rc = SalesDoc(
            number=_child_number("RC"),
            doc_type="RC",
            status="UNISSUED",
            customer_id=ct.customer_id,
            project_name=ct.project_name,
            po_customer=(inst.po_customer_sub or ct.po_customer or ""),
            credit_days=ct.credit_days or 0,
            tax_mode=ct.tax_mode,
            wht_pct=ct.wht_pct or 0,
            date=inst.bill_date,
            remark=f"(งวดที่ {inst.installment_no}) {ct.number} : {inst.period_start.strftime('%d/%m/%Y')} - {inst.period_end.strftime('%d/%m/%Y')}",
            parent_id=ct.id,
            billing_mode="ONCE",
        )
        db.session.add(rc)
        db.session.flush()
        src = None
        if ct.parent_id:
            # ถ้ามีใบจอง (BK) ให้เอารายการจาก BK เพื่อพก allocated_skus / หมวดอุปกรณ์
            src = (SalesDoc.query.options(joinedload(SalesDoc.items))
                   .filter_by(parent_id=ct.parent_id, doc_type='BK')
                   .first())
            if not src:
                # fallback: เอารายการจากใบเสนอราคา (QU)
                src = SalesDoc.query.options(joinedload(SalesDoc.items)).get(ct.parent_id)
        if src:
            _clone_items(src, rc)
        rc.amount_subtotal = inst.amount_subtotal
        rc.amount_vat = inst.amount_vat
        rc.amount_total = inst.amount_total
        rc.amount_wht = inst.amount_wht
        rc.amount_grand = inst.amount_grand
        inst.receipt_id = rc.id

    # update status
    inst.status = "INVOICED"
    return bl, iv, rc


def _next_return_number_by_date_with_prefix(prefix: str = "RT",
                                            dt: date | None = None) -> str:
    """
    gen เลขที่ใบคืนสินค้าแบบ RTYYYYMMDD001 คล้าย ๆ กับใบเคลม
    """
    dt = dt or date.today()
    yyyymmdd = dt.strftime("%Y%m%d")
    prefix_today = f"{prefix}{yyyymmdd}"
    like_prefix = f"{prefix_today}%"

    last = (
        db.session.query(ReturnDoc)
        .filter(ReturnDoc.number.like(like_prefix))
        .order_by(ReturnDoc.number.desc())
        .first()
    )

    if not last or not (last.number or "").startswith(prefix_today):
        return f"{prefix_today}001"

    m = re.match(rf"^{prefix_today}(\d{{3}})$", last.number or "")
    if not m:
        return f"{prefix_today}001"

    seq = int(m.group(1)) + 1
    return f"{prefix_today}{seq:03d}"


def _build_item_image_map(doc: SalesDoc) -> dict[int, str]:
    img_map: dict[int, str] = {}

    for it in doc.items:
        # ดึง path จาก item ก่อน
        img_path = getattr(it, "image_path", None)

        # เผื่อในอนาคตมี it.equipment ก็ลองดึงต่อ แต่จะไม่ error ถ้าไม่มี attribute นี้
        if not img_path:
            eq = getattr(it, "equipment", None)
            if eq is not None:
                img_path = getattr(eq, "image_path", None)

        if not img_path:
            continue

        # แปลง path เป็น URL
        if "://" in img_path or img_path.startswith("data:"):
            url = img_path
        else:
            rel = img_path.lstrip("/")
            if rel.startswith("static/"):
                rel = rel[7:]
            url = url_for("static", filename=rel)

        img_map[it.id] = url

    return img_map


# ================== SQLite PRAGMA ==================
@event.listens_for(Engine, "connect")
def set_sqlite_pragma(dbapi_conn, conn_record):
    """Enable foreign key constraint only when using SQLite."""
    # ถ้าเป็น SQLite เท่านั้นค่อยสั่ง PRAGMA
    if isinstance(dbapi_conn, SQLite3Connection):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

# ================== Routes ==================
@app.route("/")
def home():
    if not current_user.is_authenticated:
        return redirect(url_for("login"))
    return redirect(url_for("dashboard"))

# ---- Dashboard ----
# ---- Dashboard ----
@app.route("/dashboard")
@permission_required("dashboard.view")
def dashboard():
    from collections import defaultdict
    from datetime import datetime, time
    from sqlalchemy.orm import joinedload

    today = date.today()
    rng = (request.args.get("range") or "7d").lower()

    # ---------- 1) คำนวณช่วงวันที่ ----------
    def _parse_range(rng_key: str):
        nonlocal today
        if rng_key == "today":
            return today, today, "today"
        elif rng_key == "7d":
            return today - timedelta(days=6), today, "7d"
        elif rng_key == "30d":
            return today - timedelta(days=29), today, "30d"
        elif rng_key == "1y":
            return today - timedelta(days=365), today, "1y"
        elif rng_key == "custom":
            s = request.args.get("start") or ""
            e = request.args.get("end") or ""
            try:
                start_d = datetime.strptime(s, "%Y-%m-%d").date() if s else today - timedelta(days=6)
                end_d = datetime.strptime(e, "%Y-%m-%d").date() if e else today
                if start_d > end_d:
                    start_d, end_d = end_d, start_d
                return start_d, end_d, "custom"
            except ValueError:
                # ถ้าพิมพ์วันที่ผิด format ให้ fallback เป็น 7 วัน
                return today - timedelta(days=6), today, "7d"
        # ค่าอื่น ๆ ให้ fallback เป็น 7 วัน
        return today - timedelta(days=6), today, "7d"

    start, end, rng = _parse_range(rng)

    # เตรียม list ของทุกวันในช่วง
    day_count = (end - start).days + 1
    days = [start + timedelta(days=i) for i in range(max(day_count, 1))]

    # helper แปลงจำนวนเงินจาก SalesDoc ให้ปลอดภัย
    def _doc_amount(d: "SalesDoc") -> float:
        for attr in ("amount_grand", "amount_total", "amount_subtotal"):
            if hasattr(d, attr):
                val = getattr(d, attr) or 0.0
                try:
                    return float(val)
                except Exception:
                    continue
        return 0.0

    def _safe_num(x) -> float:
        try:
            return float(x or 0)
        except Exception:
            return 0.0

    # ---------- 2) รายรับจากใบเสร็จรับเงิน (RC) ----------
    # รองรับสถานะได้หลายแบบ (ระบบคุณใช้คำว่า "ชำระแล้ว" บน UI)
    # ถ้าต้องการ "นับเฉพาะจ่ายแล้วจริง" ให้เหลือแค่ ["PAID","RECEIPTED"]
    rc_statuses = [
        "PAID", "Paid", "paid",
        "RECEIPTED", "Receipted", "receipted",
        "ISSUED", "Issued", "issued",
    ]

    rc_docs = (
        SalesDoc.query
        .options(joinedload(SalesDoc.items), joinedload(SalesDoc.customer))
        .filter(
            SalesDoc.doc_type == "RC",
            SalesDoc.status.in_(rc_statuses),
            SalesDoc.date.between(start, end),
        )
        .all()
    )

    income_by_day = defaultdict(float)
    for d in rc_docs:
        amt = _doc_amount(d)
        if d.date:
            income_by_day[d.date] += amt
    total_income = round(sum(income_by_day.values()), 2)

    # ---------- 3) รายจ่าย: GRN + งานซ่อม + ค่าเสื่อม ----------
    # 3.1 ใบรับสินค้า (GoodsReceipt) ในช่วง
    grn_list = (
        GoodsReceipt.query
        .filter(
            GoodsReceipt.status == "RECEIVED",
            GoodsReceipt.grn_date.between(start, end),
        )
        .all()
    )
    grn_total = 0.0
    expense_by_day = defaultdict(float)

    for g in grn_list:
        amt = _safe_num(getattr(g, "amount_subtotal", 0.0))
        grn_total += amt
        if g.grn_date:
            expense_by_day[g.grn_date] += amt

    # 3.2 ค่าซ่อมจากงานซ่อม (RepairJob) ที่ปิดงานในช่วง
    start_dt = datetime.combine(start, time.min)
    end_dt = datetime.combine(end, time.max)

    repair_jobs = (
        RepairJob.query
        .filter(
            RepairJob.closed_at.isnot(None),
            RepairJob.closed_at >= start_dt,
            RepairJob.closed_at <= end_dt,
        )
        .all()
    )

    repairs_total = 0.0
    for job in repair_jobs:
        cost = _safe_num(job.total_cost)
        repairs_total += cost
        if job.closed_at:
            d = job.closed_at.date()
            expense_by_day[d] += cost

    # 3.3 ค่าเสื่อม (คำนวณแบบ straight-line จาก Equipment ทุกตัว)
    equipments = Equipment.query.all()
    depreciation_total = 0.0
    for eq in equipments:
        if not eq.received_date:
            continue
        # ช่วงที่นับค่าเสื่อมของตัวนี้จริง ๆ
        eq_start = max(start, eq.received_date)
        eq_end_life = eq.received_date + timedelta(days=eq.lifetime_days - 1)
        eq_end = min(end, eq_end_life)
        if eq_start > eq_end:
            continue
        days_eq = (eq_end - eq_start).days + 1
        daily_dep = _safe_num(eq.depreciation_per_day)
        if daily_dep <= 0:
            continue
        depreciation_total += daily_dep * days_eq
        # ลงเป็นรายวันให้กราฟด้วย
        for i in range(days_eq):
            d = eq_start + timedelta(days=i)
            expense_by_day[d] += daily_dep

    total_expense = round(grn_total + repairs_total + depreciation_total, 2)

    # ---------- 4) สถานะใบจัดส่ง / ใบวางบิล ----------
    # ใบจัดส่งทั้งหมดตอนนี้ (ไม่จำกัดช่วง)
    waiting_dn = DeliveryDoc.query.filter(DeliveryDoc.status == DeliveryStatus.PENDING).count()
    done_dn = DeliveryDoc.query.filter(DeliveryDoc.status == DeliveryStatus.DONE).count()

    # ใบวางบิล (BL) ภายในช่วง + บิลเกินกำหนดชำระ (ดูวันที่วันนี้)
    bills = SalesDoc.query.filter(SalesDoc.doc_type == "BL").all()
    billed_in_range = [b for b in bills if b.date and start <= b.date <= end]

    overdue_count = 0
    for b in bills:
        status = (b.status or "").upper()
        credit_days = b.credit_days or 0
        if not b.date:
            continue
        due_date = b.date + timedelta(days=credit_days)
        if status != "PAID" and due_date < today:
            overdue_count += 1

    # ---------- 5) อุปกรณ์ที่กำลังถูกเช่า + ลูกค้า ----------
    rented_equips = Equipment.query.filter(Equipment.status == "RENTED").all()
    rented_ids = [e.id for e in rented_equips] or [-1]

    logs = (
        EquipmentLog.query
        .filter(EquipmentLog.equipment_id.in_(rented_ids))
        .order_by(EquipmentLog.equipment_id, EquipmentLog.created_at.desc())
        .all()
    )

    last_rent_log = {}
    for lg in logs:
        if lg.action == "RENT_OUT" and lg.equipment_id not in last_rent_log:
            last_rent_log[lg.equipment_id] = lg

    renting_items = []
    for eq in rented_equips:
        lg = last_rent_log.get(eq.id)
        cust_name = lg.customer_name if lg and lg.customer_name else "-"
        renting_items.append({
            "sku": eq.sku,
            "name": eq.name,
            "customer": cust_name,
        })

    # ---------- 6) Top 5 อุปกรณ์ทำเงินสูงสุด ----------
    item_income = defaultdict(float)
    for d in rc_docs:
        for it in (d.items or []):
            item_income[it.name] += _safe_num(getattr(it, "line_total", 0.0))

    top_items = sorted(
        [{"name": name, "amount": round(val, 2)} for name, val in item_income.items()],
        key=lambda x: x["amount"],
        reverse=True
    )[:5]

    # ---------- 7) อุปกรณ์ที่ส่งซ่อม (ยังไม่ DONE) ----------
    open_repairs_q = (
        db.session.query(RepairJob, Equipment, Customer)
        .join(Equipment, RepairJob.equipment_id == Equipment.id)
        .outerjoin(Customer, RepairJob.customer_id == Customer.id)
        .filter(RepairJob.status != "DONE")
        .all()
    )

    repairs_list = []
    status_th = {
        "OPEN": "รอเริ่มงาน",
        "IN_PROGRESS": "กำลังซ่อม",
        "DONE": "ซ่อมเสร็จ",
    }
    for job, eq, cust in open_repairs_q:
        repairs_list.append({
            "job_no": job.number,
            "equipment": f"{eq.sku} · {eq.name}",
            "customer": cust.name if cust else "-",
            "status": status_th.get(job.status or "", job.status or "-"),
        })

    # ---------- 8) Top 5 ลูกค้าที่เช่าเรามากที่สุด ----------
    customer_income = defaultdict(float)
    for d in rc_docs:
        cust = d.customer
        cust_name = cust.name if cust else "(ไม่ระบุ)"
        customer_income[cust_name] += _doc_amount(d)

    top_customers = sorted(
        [{"name": name, "amount": round(val, 2)} for name, val in customer_income.items()],
        key=lambda x: x["amount"],
        reverse=True
    )[:5]

    # ---------- 9) ข้อมูลกราฟ (รายวัน + เปรียบเทียบเดือน) ----------
    labels = [d.strftime("%Y-%m-%d") for d in days]
    income_series = [round(income_by_day.get(d, 0.0), 2) for d in days]
    expense_series = [round(expense_by_day.get(d, 0.0), 2) for d in days]

    # helper หา first/last day ของเดือน
    def _month_range(y: int, m: int):
        first = date(y, m, 1)
        if m == 12:
            last = date(y + 1, 1, 1) - timedelta(days=1)
        else:
            last = date(y, m + 1, 1) - timedelta(days=1)
        return first, last

    cy, cm = today.year, today.month
    if cm == 1:
        py, pm = cy - 1, 12
    else:
        py, pm = cy, cm - 1
    ly, lm = cy - 1, cm

    def _sum_rc_in_month(y: int, m: int) -> float:
        first, last = _month_range(y, m)
        docs = (
            SalesDoc.query
            .filter(
                SalesDoc.doc_type == "RC",
                SalesDoc.status.in_(rc_statuses),
                SalesDoc.date.between(first, last),
            )
            .all()
        )
        return round(sum(_doc_amount(d) for d in docs), 2)

    cur_month_val = _sum_rc_in_month(cy, cm)
    prev_month_val = _sum_rc_in_month(py, pm)
    last_year_val = _sum_rc_in_month(ly, lm)

    def _pct_change(cur: float, base: float):
        if not base:
            return None
        return round((cur - base) * 100.0 / base, 1)

    month_compare = {
        "current": {
            "label": f"{cm:02d}/{cy}",
            "value": cur_month_val,
        },
        "prev": {
            "label": f"{pm:02d}/{py}",
            "value": prev_month_val,
        },
        "last_year": {
            "label": f"{lm:02d}/{ly}",
            "value": last_year_val,
        },
        "delta_prev_pct": _pct_change(cur_month_val, prev_month_val),
        "delta_ly_pct": _pct_change(cur_month_val, last_year_val),
    }

    stats = {
        "income_total": round(total_income, 2),
        "expense_total": round(total_expense, 2),
        "expense_breakdown": {
            "depr": round(depreciation_total, 2),
            "repair": round(repairs_total, 2),
            "grn": round(grn_total, 2),
        },
        "delivery": {
            "waiting": waiting_dn,
            "done": done_dn,
        },
        "billing": {
            "billed": len(billed_in_range),
            "overdue": overdue_count,
        },
        "renting": {
            "count": len(renting_items),
            "items": renting_items,
        },
        "top_items": top_items,
        "repairs": repairs_list,
        "top_customers": top_customers,
        "chart": {
            "labels": labels,
            "income": income_series,
            "expense": expense_series,
            "month_compare": month_compare,
        },
    }

    return render_template(
        "dashboard.html",
        start=start,
        end=end,
        rng=rng,
        stats=stats,
        today=today,
    )

# ---- Auth ----
from werkzeug.security import check_password_hash  # ถ้ายังไม่ได้ import ไว้ด้านบน ให้ใส่บรรทัดนี้เพิ่ม

@app.route("/auth/login", methods=["GET", "POST"])
def login():
    # ถ้าล็อกอินอยู่แล้ว ให้เด้งไปหน้า dashboard เลย
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""

        if not username or not password:
            flash("กรุณากรอกชื่อผู้ใช้และรหัสผ่าน", "danger")
            return render_template("auth/login.html")

        # หา user ตาม username
        user = User.query.filter_by(username=username).first()

        # ---------- เช็ครหัสผ่าน ----------
        password_ok = False

        if user:
            if hasattr(user, "check_password"):
                # กรณี model มี method check_password()
                try:
                    password_ok = user.check_password(password)
                except Exception as e:
                    print(f"[login] user.check_password error: {e}")
                    password_ok = False
            elif hasattr(user, "password_hash"):
                # กรณีเก่า: เก็บ hash ไว้ใน field password_hash
                try:
                    if user.password_hash:
                        password_ok = check_password_hash(user.password_hash, password)
                except Exception as e:
                    print(f"[login] check_password_hash error: {e}")
                    password_ok = False
            elif hasattr(user, "password"):
                # fallback สุดท้าย: เก็บ plain text ไว้ใน field password
                password_ok = (user.password == password)

        if password_ok:
            # ถ้ามีฟิลด์ is_active และเป็น False ก็ไม่ให้เข้า
            if hasattr(user, "is_active") and not user.is_active:
                flash("บัญชีนี้ถูกปิดการใช้งาน", "danger")
            else:
                login_user(user)
                next_url = request.args.get("next") or url_for("dashboard")
                print(f"[login] user '{user.username}' logged in")
                return redirect(next_url)
        else:
            flash("ชื่อผู้ใช้หรือรหัสผ่านผิด", "danger")
            print(f"[login] invalid login for username='{username}'")

    # GET หรือกรณีเช็คไม่ผ่าน กลับมาแสดงหน้า login
    return render_template("auth/login.html")



@app.route("/auth/logout", methods=["POST"])
def logout():
    logout_user()
    return redirect(url_for("login"))

# ---- User Management ----
@app.route("/admin/users")
@permission_required("users.manage")
def users_list():
    users = User.query.order_by(User.username.asc()).all()
    roles = Role.query.order_by(Role.name.asc()).all()
    return render_template("admin/users_list.html", users=users, roles=roles)

@app.route("/admin/users/new", methods=["GET", "POST"])
@permission_required("users.manage")
def users_new():
    roles = Role.query.order_by(Role.name.asc()).all()
    perms = Permission.query.order_by(Permission.code.asc()).all()
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        full_name = (request.form.get("full_name") or "").strip()
        password = request.form.get("password") or ""
        confirm  = request.form.get("confirm") or ""
        is_active = bool(request.form.get("is_active"))
        if not username or not password:
            flash("กรอกชื่อผู้ใช้และรหัสผ่าน", "danger"); return redirect(url_for("users_new"))
        if password != confirm:
            flash("รหัสผ่านไม่ตรงกัน", "danger"); return redirect(url_for("users_new"))
        if User.query.filter_by(username=username).first():
            flash("มีชื่อผู้ใช้นี้แล้ว", "danger"); return redirect(url_for("users_new"))
        u = User(
            username=username,
            full_name=full_name,
            password_hash=generate_password_hash(password),
            is_active=is_active,
        )
        db.session.add(u); db.session.flush()
        for rid in request.form.getlist("roles"):
            db.session.add(UserRole(user_id=u.id, role_id=int(rid)))
        for pid in request.form.getlist("perms"):
            db.session.add(UserPermission(user_id=u.id, perm_id=int(pid)))
        db.session.commit()
        flash("เพิ่มผู้ใช้เรียบร้อย", "success")
        return redirect(url_for("users_list"))
    return render_template("admin/users_form.html", roles=roles, perms=perms)

@app.route("/admin/users/<int:uid>/edit", methods=["GET", "POST"])
@permission_required("users.manage")
def users_edit(uid):
    u = db.session.get(User, uid) or abort(404)
    roles = Role.query.order_by(Role.name.asc()).all()
    perms = Permission.query.order_by(Permission.code.asc()).all()
    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        password  = request.form.get("password") or ""
        confirm   = request.form.get("confirm") or ""
        is_active = bool(request.form.get("is_active"))
        u.full_name = full_name
        u.is_active = is_active
        if password:
            if password != confirm:
                flash("รหัสผ่านใหม่ไม่ตรงกัน", "danger")
                return redirect(url_for("users_edit", uid=u.id))
            u.password_hash = generate_password_hash(password)
        db.session.expire(u, ["roles", "perms"])
        db.session.query(UserRole).filter_by(user_id=u.id).delete(synchronize_session=False)
        db.session.query(UserPermission).filter_by(user_id=u.id).delete(synchronize_session=False)
        db.session.flush()
        for rid in request.form.getlist("roles"):
            db.session.add(UserRole(user_id=u.id, role_id=int(rid)))
        for pid in request.form.getlist("perms"):
            db.session.add(UserPermission(user_id=u.id, perm_id=int(pid)))
        db.session.commit()
        flash("บันทึกการแก้ไขแล้ว", "success")
        return redirect(url_for("users_list"))
    role_ids_have = {r.id for r in u.roles}
    perm_ids_have = {p.id for p in u.perms}
    return render_template(
        "admin/users_form.html",
        u=u, roles=roles, perms=perms,
        role_ids_have=role_ids_have, perm_ids_have=perm_ids_have,
        is_edit=True,
    )

@app.route("/admin/users/<int:uid>/delete", methods=["POST"])
@permission_required("users.manage")
def users_delete(uid):
    u = db.session.get(User, uid) or abort(404)
    if u.id == current_user.id:
        flash("ไม่สามารถลบผู้ใช้ที่กำลังใช้งานอยู่ได้", "warning")
        return redirect(url_for("users_list"))
    if u.username == "admin":
        flash("ห้ามลบผู้ใช้ admin", "warning")
        return redirect(url_for("users_list"))
    if any(r.code == "admin" for r in u.roles) and not _has_other_admin(u.id):
        flash("ต้องมีผู้ดูแลระบบอย่างน้อย 1 คน ไม่สามารถลบได้", "warning")
        return redirect(url_for("users_list"))
    db.session.expire(u, ["roles", "perms"])
    db.session.query(UserRole).filter_by(user_id=u.id).delete(synchronize_session=False)
    db.session.query(UserPermission).filter_by(user_id=u.id).delete(synchronize_session=False)
    db.session.delete(u)
    db.session.commit()
    flash("ลบผู้ใช้แล้ว", "success")
    return redirect(url_for("users_list"))

# ---------- Inject Company to all templates ----------
@app.context_processor
def inject_company_profile():
    try:
        return {"company": get_company()}
    except Exception:
        return {"company": None}

# ---------- Company Settings ----------
UPLOAD_DIR = os.path.join(app.static_folder, "uploads", "company")
ALLOWED_LOGO = {".png", ".jpg", ".jpeg", ".webp", ".svg"}
MAX_LOGO_MB = 2

@app.route("/admin/company", methods=["GET", "POST"])
@permission_required("company.manage")
def company_edit():
    prof = get_company()
    if request.method == "POST":
        prof.name = (request.form.get("name") or "").strip()
        prof.address = (request.form.get("address") or "").strip()
        prof.district = (request.form.get("district") or "").strip()
        prof.amphoe = (request.form.get("amphoe") or "").strip()
        prof.province = (request.form.get("province") or "").strip()
        prof.postcode = (request.form.get("postcode") or "").strip()
        prof.phone = (request.form.get("phone") or "").strip()
        prof.tax_id = (request.form.get("tax_id") or "").strip()
        f = request.files.get("logo")
        if f and f.filename:
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            ext = os.path.splitext(f.filename.lower())[1]
            if ext not in ALLOWED_LOGO:
                flash("ไฟล์โลโก้ต้องเป็น PNG/JPG/JPEG/WEBP/SVG", "warning")
                return redirect(url_for("company_edit"))
            f.seek(0, os.SEEK_END)
            size_mb = f.tell() / (1024 * 1024)
            f.seek(0)
            if size_mb > MAX_LOGO_MB:
                flash(f"ไฟล์ใหญ่เกิน {MAX_LOGO_MB}MB", "warning")
                return redirect(url_for("company_edit"))
            filename = "logo" + ext
            save_path = os.path.join(UPLOAD_DIR, secure_filename(filename))
            f.save(save_path)
            prof.logo_path = os.path.join("uploads", "company", filename).replace("\\", "/")
        db.session.commit()
        flash("บันทึกข้อมูลบริษัทเรียบร้อย", "success")
        return redirect(url_for("company_edit"))
    return render_template("admin/company_form.html", prof=prof, has_logo=bool(prof.logo_path))

# ---------- Purchases (PO/GRN) ----------
@app.route("/purchases/po")
@permission_required("purchases.view")
def po_list():
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    qry = PurchaseOrder.query.options(joinedload(PurchaseOrder.supplier))

    if start_d:
        qry = qry.filter(PurchaseOrder.po_date >= start_d)
    if end_d:
        qry = qry.filter(PurchaseOrder.po_date <= end_d)

    if q:
        like = f"%{q}%"
        qry = qry.outerjoin(PurchaseOrder.supplier).filter(
            or_(
                PurchaseOrder.number.ilike(like),
                Supplier.name.ilike(like),
            )
        )

    pos = qry.order_by(PurchaseOrder.id.desc()).all()
    return render_template("purchases/po_list.html", pos=pos, q=q, start=start, end=end)

@app.route("/purchases/po/new", methods=["GET", "POST"])
@permission_required("purchases.create")
def po_new():
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()

    def _parse_items_any(form):
        """
        รองรับ 2 แบบ:
        A) แบบใหม่: items-0-name, items-0-sku, items-0-brand, items-0-warehouse, items-0-qty, items-0-unit,
                    items-0-unit_price, items-0-discount_pct
        B) แบบเดิม: item_name[], item_sku[], item_brand[], item_qty[], item_unit[], item_cost[], item_disc[]
        คืน list ของ dict ที่ "มีรายการจริง"
        """
        # default warehouse (ทั้งเอกสาร) ถ้า form ไม่มี warehouse ต่อบรรทัด
        default_wh = (form.get("warehouse") or "").strip() or "MAIN"

        # --- แบบใหม่: items-{i}-{field} ---
        has_new = any(k.startswith("items-") for k in form.keys())
        if has_new:
            import re
            pat = re.compile(r"^items-(\d+)-([a-zA-Z_]+)$")
            rows_map = {}

            for k, v in form.items():
                m = pat.match(k)
                if not m:
                    continue
                idx = int(m.group(1))
                field = m.group(2)
                rows_map.setdefault(idx, {})[field] = (v or "").strip()

            out = []
            for idx in sorted(rows_map.keys()):
                r = rows_map[idx]
                name = (r.get("name") or "").strip()
                sku = (r.get("sku") or "").strip()
                brand = (r.get("brand") or "").strip()
                unit = (r.get("unit") or "").strip() or "ชิ้น"

                # ✅ warehouse ต่อบรรทัด (ถ้ามี) ไม่งั้นใช้ default_wh
                warehouse = (r.get("warehouse") or "").strip() or default_wh

                try:
                    qty = float(r.get("qty") or 0)
                except Exception:
                    qty = 0.0

                # template ใหม่ใช้ unit_price / discount_pct
                try:
                    unit_cost = float(r.get("unit_price") or 0)
                except Exception:
                    unit_cost = 0.0

                try:
                    discount_pct = float(r.get("discount_pct") or 0)
                except Exception:
                    discount_pct = 0.0

                # ถือว่าเป็นรายการจริง: ต้องมี name หรือ sku และ qty > 0
                if (name or sku) and qty > 0:
                    out.append({
                        "name": name,
                        "sku": sku,
                        "brand": brand,
                        "warehouse": warehouse,
                        "qty": qty,
                        "unit": unit,
                        "unit_cost": unit_cost,
                        "discount_pct": discount_pct,
                    })
            return out

        # --- แบบเดิม: item_name[] ---
        names = form.getlist("item_name[]")
        skus  = form.getlist("item_sku[]")
        brands = form.getlist("item_brand[]")
        qtys  = form.getlist("item_qty[]")
        units = form.getlist("item_unit[]")
        costs = form.getlist("item_cost[]")
        discs = form.getlist("item_disc[]")

        # (ถ้าแบบเดิมยังไม่มี warehouse[] ก็ใช้ default_wh ให้ทั้งหมด)
        warehouses = form.getlist("item_warehouse[]")  # optional

        out = []
        for i, name in enumerate(names):
            name = (name or "").strip()
            sku = (skus[i] or "").strip() if i < len(skus) else ""
            brand = (brands[i] or "").strip() if i < len(brands) else ""
            unit = (units[i] or "ชิ้น") if i < len(units) else "ชิ้น"

            warehouse = (
                (warehouses[i] or "").strip() if i < len(warehouses) and warehouses else ""
            ) or default_wh

            if not name and not sku:
                continue

            try:
                qty = float(qtys[i] or 0) if i < len(qtys) else 0.0
            except Exception:
                qty = 0.0
            if qty <= 0:
                continue

            try:
                unit_cost = float(costs[i] or 0) if i < len(costs) else 0.0
            except Exception:
                unit_cost = 0.0

            try:
                discount_pct = float(discs[i] or 0) if i < len(discs) else 0.0
            except Exception:
                discount_pct = 0.0

            out.append({
                "name": name,
                "sku": sku,
                "brand": brand,
                "warehouse": warehouse,
                "qty": qty,
                "unit": unit,
                "unit_cost": unit_cost,
                "discount_pct": discount_pct,
            })

        return out

    if request.method == "POST":
        supplier_id = request.form.get("supplier_id", type=int) or 0
        rows = _parse_items_any(request.form)

        if not supplier_id:
            flash("กรุณาเลือกผู้ขาย (Supplier)", "danger")
            return render_template(
                "purchases/po_form.html",
                suppliers=suppliers,
                doc=None,
                mode="create",
                selected_id=None,
            ), 400

        if not rows:
            flash("กรุณาใส่รายการอย่างน้อย 1 รายการ", "danger")
            return render_template(
                "purchases/po_form.html",
                suppliers=suppliers,
                doc=None,
                mode="create",
                selected_id=supplier_id,
            ), 400

        po = PurchaseOrder(
            number=_gen_running("PO", PurchaseOrder),
            supplier_id=supplier_id,
            po_date=date.today(),
            status="DRAFT",
        )
        db.session.add(po)
        db.session.flush()

        for r in rows:
            db.session.add(POItem(
                po_id=po.id,
                name=r["name"],
                sku=r.get("sku") or "",
                brand=r.get("brand") or "",
                qty=r["qty"],
                unit=r["unit"],
                unit_cost=r["unit_cost"],
                discount_pct=r["discount_pct"],
            ))
            # หมายเหตุ: ตอนนี้ POItem model ของคุณอาจยังไม่มี field warehouse
            # ถ้ามีแล้วค่อยเปิดใช้:
            # warehouse=r.get("warehouse") or "MAIN",

        db.session.commit()
        flash("สร้างใบสั่งซื้อเรียบร้อย", "success")
        return redirect(url_for("po_view", pid=po.id))

    selected_id = request.args.get("selected_id", type=int)
    return render_template(
        "purchases/po_form.html",
        suppliers=suppliers,
        doc=None,
        mode="create",
        selected_id=selected_id,
    )


@app.route("/purchases/po/<int:pid>")
@permission_required("purchases.view")
def po_view(pid):
    po = PurchaseOrder.query.get_or_404(pid)
    return render_template("purchases/po_view.html", po=po)

@app.route("/purchases/po/<int:pid>/set_status", methods=["POST"])
@permission_required("purchases.create")
def po_set_status(pid):
    po = PurchaseOrder.query.get_or_404(pid)
    new_status = request.form.get("status") or "DRAFT"
    if new_status not in ("DRAFT", "APPROVED", "ORDERED"):
        flash("สถานะไม่ถูกต้อง", "danger")
        return redirect(url_for("po_view", pid=pid))
    po.status = new_status
    db.session.commit()
    flash("อัปเดตสถานะแล้ว", "success")
    return redirect(url_for("po_view", pid=pid))

@app.route("/purchases/grn")
@permission_required("goods.receive")
def grn_list():
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    qry = GoodsReceipt.query.options(
        joinedload(GoodsReceipt.po).joinedload(PurchaseOrder.supplier),
        joinedload(GoodsReceipt.po),
    )

    if start_d:
        qry = qry.filter(GoodsReceipt.grn_date >= start_d)
    if end_d:
        qry = qry.filter(GoodsReceipt.grn_date <= end_d)

    if q:
        like = f"%{q}%"
        qry = qry.outerjoin(GoodsReceipt.po).outerjoin(PurchaseOrder.supplier).outerjoin(GoodsReceipt.po).filter(
            or_(
                GoodsReceipt.number.ilike(like),
                Supplier.name.ilike(like),
                PurchaseOrder.number.ilike(like),
            )
        )

    grns = qry.order_by(GoodsReceipt.id.desc()).all()
    return render_template("purchases/grn_list.html", grns=grns, q=q, start=start, end=end)

@app.post("/purchases/po/<int:pid>/create_grn")
@permission_required("purchases.manage")
def po_create_grn(pid):
    # --- โหลด PO ---
    po = (
        PurchaseOrder.query
        .options(
            joinedload(PurchaseOrder.items),
            joinedload(PurchaseOrder.supplier),
        )
        .get_or_404(pid)
    )

    # --- กันกดซ้ำ: ถ้ามี GRN ของ PO นี้แล้ว ให้พาไปใบเดิม ---
    existing = GoodsReceipt.query.filter_by(po_id=po.id).order_by(GoodsReceipt.id.desc()).first()
    if existing:
        flash("มีใบรับสินค้า (RC) ของใบสั่งซื้อนี้แล้ว", "info")
        return redirect(url_for("grn_view", gid=existing.id))

    # --- สร้างหัวเอกสาร GRN ---
    grn = GoodsReceipt()

    # ฟังก์ชันช่วย set attribute แบบปลอดภัย
    def _set_if_has(obj, field, value):
        try:
            if hasattr(obj, field):
                setattr(obj, field, value)
                return True
        except Exception:
            pass
        return False

    # number
    if hasattr(grn, "number"):
        grn.number = _gen_running("RC", GoodsReceipt)

    # link PO
    if hasattr(grn, "po_id"):
        grn.po_id = po.id
    elif hasattr(grn, "po"):
        grn.po = po

    # supplier
    if hasattr(grn, "supplier_id"):
        grn.supplier_id = po.supplier_id
    elif hasattr(grn, "supplier"):
        grn.supplier = po.supplier

    # date (เก็บค่าวันที่จริงไว้ใช้ต่อ)
    grn_date_val = date.today()
    if hasattr(grn, "grn_date"):
        grn.grn_date = grn_date_val
    elif hasattr(grn, "receive_date"):
        grn.receive_date = grn_date_val
    elif hasattr(grn, "doc_date"):
        grn.doc_date = grn_date_val

    # status
    if hasattr(grn, "status"):
        grn.status = "RECEIVED"

    try:
        db.session.add(grn)
        db.session.flush()  # ให้ได้ grn.id

        # --- สร้างรายการ GRN จาก POItem ---
        created = 0

        for it in (po.items or []):
            name = (getattr(it, "name", "") or "").strip()
            sku = (getattr(it, "sku", "") or "").strip()
            if not (name or sku):
                continue

            qty = float(getattr(it, "qty", 0) or 0)
            if qty <= 0:
                continue

            unit_cost = float(getattr(it, "unit_cost", None) or getattr(it, "unit_price", 0) or 0)
            discount_pct = float(getattr(it, "discount_pct", 0) or 0)

            before = qty * unit_cost
            disc_amt = before * (discount_pct / 100.0)
            line_total = max(0.0, before - disc_amt)

            gi = GRNItem()

            # link grn
            if hasattr(gi, "grn_id"):
                gi.grn_id = grn.id
            elif hasattr(gi, "grn"):
                gi.grn = grn

            # ข้อมูลสินค้า
            _set_if_has(gi, "sku", sku)
            _set_if_has(gi, "name", name)
            _set_if_has(gi, "unit", getattr(it, "unit", "") or "ชิ้น")
            _set_if_has(gi, "qty", getattr(it, "qty", qty))
            _set_if_has(gi, "brand", getattr(it, "brand", "") or "")

            # ราคา/ส่วนลด
            if not _set_if_has(gi, "unit_cost", unit_cost):
                _set_if_has(gi, "unit_price", unit_cost)

            _set_if_has(gi, "discount_pct", discount_pct)
            _set_if_has(gi, "line_total", line_total)

            db.session.add(gi)
            db.session.flush()  # ✅ ให้ได้ gi.id เพื่อเอาไปใส่ incoming.grn_item_id

            # --- สร้างรายการ "รอเพิ่มเข้าระบบ" ตามจำนวนที่รับเข้า ---
            try:
                incoming_n = int(qty)
            except Exception:
                incoming_n = 0
            if incoming_n < 0:
                incoming_n = 0

            for _ in range(incoming_n):
                inc = IncomingEquipment(
                    grn_id=grn.id,
                    grn_item_id=gi.id,           # ✅ จำเป็นมาก
                    name=(name or "").strip(),   # ✅ ใช้ name (ไม่ใช่ item_name)
                    brand=(getattr(it, "brand", "") or "").strip(),
                    unit_cost=float(unit_cost or 0),
                    received_date=grn_date_val,  # ✅ ไม่อ้าง grn.date / grn.grn_date แบบมั่ว
                    status="PENDING",
                )
                db.session.add(inc)

            created += 1

        if created <= 0:
            db.session.rollback()
            flash("ไม่พบรายการในใบสั่งซื้อ จึงไม่สามารถสร้าง GRN ได้", "danger")
            return redirect(url_for("po_view", pid=po.id))

        db.session.commit()
        flash("สร้างใบรับสินค้า (RC) เรียบร้อย", "success")
        return redirect(url_for("grn_view", gid=grn.id))

    except Exception as e:
        db.session.rollback()
        raise



@app.get("/purchases/grn/<int:gid>")
@permission_required("purchases.view")
def grn_view(gid):
    grn = (
        GoodsReceipt.query
        .options(
            joinedload(GoodsReceipt.po).joinedload(PurchaseOrder.supplier),
            joinedload(GoodsReceipt.items),
        )
        .get_or_404(gid)
    )
    # ✅ ส่งทั้ง grn และ doc (เพื่อให้ template เดิมที่ใช้ doc ไม่พัง)
    return render_template("purchases/grn_view.html", grn=grn, doc=grn)


@app.get("/purchases/grn/<int:gid>/print")
@permission_required("purchases.view")
def grn_print(gid):
    """พิมพ์ใบรับสินค้า (RC)"""
    grn = (
        GoodsReceipt.query
        .options(
            joinedload(GoodsReceipt.po).joinedload(PurchaseOrder.supplier),
            joinedload(GoodsReceipt.items),
        )
        .get_or_404(gid)
    )

    company = get_company()
    return render_template(
        "purchases/grn_print.html",
        grn=grn,
        doc=grn,
        company=company,
        today=date.today(),
    )


# ---------- Purchases: Supplier APIs ----------
@app.get("/api/suppliers")
@permission_required("purchases.view")
def api_suppliers_list():
    q = (request.args.get("q") or "").strip()
    qry = Supplier.query
    if q:
        qry = qry.filter(Supplier.name.ilike(f"%{q}%"))
    rows = qry.order_by(Supplier.name.asc()).all()
    return jsonify([{"id": s.id, "name": s.name} for s in rows])

@app.route("/api/suppliers/create", methods=["POST"])
@permission_required("purchases.create")
def api_suppliers_create():
    name = (request.form.get("name") or "").strip()
    if not name:
        flash("กรุณากรอกชื่อบริษัท/ร้าน", "danger")
        return redirect(request.referrer or url_for("po_new"))
    s = Supplier(
        name=name,
        tax_id=(request.form.get("tax_id") or "").strip(),
        address=(request.form.get("address") or "").strip(),
        district=(request.form.get("district") or "").strip(),
        amphoe=(request.form.get("amphoe") or "").strip(),
        province=(request.form.get("province") or "").strip(),
        postcode=(request.form.get("postcode") or "").strip(),
        phone=(request.form.get("phone") or "").strip(),
    )
    db.session.add(s)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash("มีชื่อผู้ขายนี้อยู่แล้ว", "warning")
        return redirect(request.referrer or url_for("po_new"))
    flash("เพิ่มผู้ขายแล้ว", "success")
    return redirect(url_for("po_new", selected_id=s.id))

# ---------- Purchases: PO (PRINT A4) ----------
@app.route("/purchases/po/<int:pid>/print")
@permission_required("purchases.view")
def po_print(pid):
    po = PurchaseOrder.query.get_or_404(pid)
    return render_template("purchases/po_print.html", po=po, today=date.today())

# ---------- Customers ----------
@app.route("/customers")
@permission_required("customers.view")
def customers_list():
    q = (request.args.get("q") or "").strip()
    qry = Customer.query
    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(
            Customer.name.ilike(like),
            Customer.phone.ilike(like),
            Customer.contact_name.ilike(like),
            Customer.tax_id.ilike(like),
        ))
    customers = qry.order_by(Customer.name.asc()).all()
    return render_template("customers/customers_list.html", customers=customers, q=q)

# ---------------------------
# Customers: Rent History
# ---------------------------
@app.get("/customers/rent-history")
@permission_required("customers.view")
def rent_history_list():
    q = (request.args.get("q") or "").strip()

    cq = Customer.query
    if q:
        like = f"%{q}%"
        cq = cq.filter(
            or_(
                Customer.name.ilike(like),
                Customer.phone.ilike(like),
                Customer.tax_id.ilike(like),
            )
        )
    customers = cq.order_by(Customer.name.asc()).all()

    # สถิติจากเอกสาร IV (ไม่นับ CANCELLED)
    stats_rows = (
        db.session.query(
            SalesDoc.customer_id.label("cid"),
            func.count(SalesDoc.id).label("count"),
            func.coalesce(func.sum(SalesDoc.amount_grand), 0).label("total"),
        )
        .filter(func.upper(SalesDoc.doc_type) == "IV")
        .filter(func.upper(func.coalesce(SalesDoc.status, "DRAFT")) != "CANCELLED")
        .group_by(SalesDoc.customer_id)
        .all()
    )
    stats = {r.cid: SimpleNamespace(count=int(r.count or 0), total=float(r.total or 0)) for r in stats_rows}

    return render_template("customers/rent_history_list.html", customers=customers, stats=stats, q=q)


@app.get("/customers/<int:cid>/rent-history")
@permission_required("customers.view")
def rent_history_view(cid):
    customer = Customer.query.get_or_404(cid)

    docs = (
        SalesDoc.query
        .filter(SalesDoc.customer_id == cid)
        .filter(func.upper(SalesDoc.doc_type) == "IV")
        .filter(func.upper(func.coalesce(SalesDoc.status, "DRAFT")) != "CANCELLED")
        .order_by(SalesDoc.date.desc().nullslast(), SalesDoc.id.desc())
        .all()
    )

    kpi_count = len(docs)
    kpi_total = float(sum([(d.amount_grand or 0) for d in docs]) or 0)
    kpi_avg = (kpi_total / kpi_count) if kpi_count else 0

    return render_template(
        "customers/rent_history_view.html",
        customer=customer,
        docs=docs,
        kpi_count=kpi_count,
        kpi_total=kpi_total,
        kpi_avg=kpi_avg,
    )

@app.route("/customers/new", methods=["GET", "POST"])
@permission_required("customers.manage")
def customers_new():
    if request.method == "POST":
        c = Customer(
            name=(request.form.get("name") or "").strip(),
            address=(request.form.get("address") or "").strip(),
            district=(request.form.get("district") or "").strip(),
            amphoe=(request.form.get("amphoe") or "").strip(),
            province=(request.form.get("province") or "").strip(),
            postcode=(request.form.get("postcode") or "").strip(),
            phone=(request.form.get("phone") or "").strip(),
            tax_id=(request.form.get("tax_id") or "").strip(),
            contact_name=(request.form.get("contact_name") or "").strip(),
            contact_phone=(request.form.get("contact_phone") or "").strip(),
            credit_term_days=int((request.form.get("credit_term_days") or 0) or 0),
            payment_terms=(request.form.get("payment_terms") or "").strip(),
        )
        if not c.name:
            flash("กรุณากรอกชื่อลูกค้า", "danger")
            return redirect(url_for("customers_new"))
        db.session.add(c)
        db.session.commit()
        flash("เพิ่มลูกค้าแล้ว", "success")
        return redirect(url_for("customers_list"))
    return render_template("customers/customers_form.html", is_edit=False, c=None)

@app.route("/customers/<int:cid>/edit", methods=["GET", "POST"])
@permission_required("customers.manage")
def customers_edit(cid):
    c = Customer.query.get_or_404(cid)
    if request.method == "POST":
        c.name = (request.form.get("name") or "").strip()
        c.address = (request.form.get("address") or "").strip()
        c.district = (request.form.get("district") or "").strip()
        c.amphoe = (request.form.get("amphoe") or "").strip()
        c.province = (request.form.get("province") or "").strip()
        c.postcode = (request.form.get("postcode") or "").strip()
        c.phone = (request.form.get("phone") or "").strip()
        c.tax_id = (request.form.get("tax_id") or "").strip()
        c.contact_name = (request.form.get("contact_name") or "").strip()
        c.contact_phone = (request.form.get("contact_phone") or "").strip()
        c.credit_term_days = int((request.form.get("credit_term_days") or 0) or 0)
        c.payment_terms = (request.form.get("payment_terms") or "").strip()
        if not c.name:
            flash("กรุณากรอกชื่อลูกค้า", "danger")
            return redirect(url_for("customers_edit", cid=c.id))
        db.session.commit()
        flash("บันทึกการแก้ไขแล้ว", "success")
        return redirect(url_for("customers_list"))
    return render_template("customers/customers_form.html", is_edit=True, c=c)

@app.post("/customers/<int:cid>/delete")
@permission_required("customers.manage")
def customers_delete(cid):
    c = Customer.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    flash("ลบลูกค้าแล้ว", "success")
    return redirect(url_for("customers_list"))

# ---------- Categories ----------
@app.route("/equipment/categories")
@permission_required("equipment.manage")
def cat_list():
    cats = Category.query.order_by(Category.prefix_sku.asc()).all()
    return render_template("equipment/cat_list.html", cats=cats)

@app.route("/equipment/categories/new", methods=["GET","POST"])
@permission_required("equipment.manage")
def cat_new():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        prefix = (request.form.get("prefix_sku") or "").strip()
        if not name or not prefix:
            flash("กรอกชื่อหมวดหมู่และ Prefix SKU", "danger")
            return redirect(url_for("cat_new"))
        if Category.query.filter((Category.name==name)|(Category.prefix_sku==prefix)).first():
            flash("ชื่อหมวดหมู่หรือ Prefix ซ้ำ", "warning")
            return redirect(url_for("cat_new"))
        db.session.add(Category(name=name, prefix_sku=prefix))
        db.session.commit()
        flash("เพิ่มหมวดหมู่แล้ว", "success")
        return redirect(url_for("cat_list"))
    return render_template("equipment/cat_form.html")

# ---------- Equipment ----------
@app.route("/equipment")
@permission_required("equipment.view")
def equip_list():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").upper()
    warehouse = (request.args.get("warehouse") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    qry = Equipment.query
    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(Equipment.sku.ilike(like), Equipment.name.ilike(like)))
    if status in EQUIP_STATUS:
        qry = qry.filter(Equipment.status == status)
    if warehouse:
        qry = qry.filter(Equipment.warehouse == warehouse)
    # date range filter (received_date)
    if start_d:
        qry = qry.filter(Equipment.received_date >= start_d)
    if end_d:
        qry = qry.filter(Equipment.received_date <= end_d)

    rows = qry.order_by(Equipment.received_date.desc(), Equipment.id.desc()).all()
    return render_template(
        "equipment/equip_list.html",
        rows=rows,
        q=q,
        status=status,
        warehouse=warehouse,
        start=start,
        end=end,
        status_th=EQUIP_STATUS_THAI,
    )





# ---------- Stock Reservation (Reserve at QU/BK line) ----------
def _norm_brand(b: str | None) -> str:
    b = (b or "").strip()
    return b

def _reserved_qty(category_id: int, brand: str | None, warehouse: str | None, exclude_doc: tuple[str,int] | None = None) -> float:
    q = StockReservation.query.filter_by(category_id=category_id, status="ACTIVE")
    wh = (warehouse or "").strip() or "MAIN"
    q = q.filter(StockReservation.warehouse == wh)

    nb = _norm_brand(brand)
    if nb:
        q = q.filter(StockReservation.brand == nb)
    else:
        q = q.filter((StockReservation.brand == None) | (StockReservation.brand == ""))

    if exclude_doc:
        dt, did = exclude_doc
        q = q.filter(~((StockReservation.doc_type == dt) & (StockReservation.doc_id == did)))

    return float(q.with_entities(func.coalesce(func.sum(StockReservation.qty), 0.0)).scalar() or 0.0)

def _ready_total(category_id: int, brand: str | None, warehouse: str | None) -> int:
    wh = (warehouse or "").strip() or "MAIN"
    q = Equipment.query.filter_by(category_id=category_id).filter(Equipment.status == "READY").filter(Equipment.warehouse == wh)
    nb = _norm_brand(brand)
    if nb:
        q = q.filter(Equipment.brand == nb)
    else:
        q = q.filter((Equipment.brand == None) | (Equipment.brand == ""))
    return int(q.count())


def _check_and_reserve(doc_type: str, doc_id: int, sales_item_id: int | None, category_id: int, brand: str | None, warehouse: str | None, qty: float):
    qty = float(qty or 0)
    if qty <= 0:
        return
    ready_total = _ready_total(category_id, brand, warehouse)
    reserved = _reserved_qty(category_id, brand, warehouse)
    available = ready_total - reserved
    if qty > available + 1e-9:
        cat = Category.query.get(category_id)
        raise ValueError(f"สต็อกไม่พอ: {cat.name if cat else 'หมวด'} / {brand or 'ไม่ระบุ'} / คลัง {((warehouse or '').strip() or 'MAIN')} (คงเหลือ {available:.0f} จาก READY {ready_total} - RESERVED {reserved})")
    r = StockReservation(
        doc_type=(doc_type or "").upper(),
        doc_id=doc_id,
        sales_item_id=sales_item_id,
        category_id=category_id,
        brand=_norm_brand(brand),
        warehouse=((warehouse or '').strip() or 'MAIN'),
        qty=qty,
        status="ACTIVE",
    )
    db.session.add(r)

def _release_reservations_for_doc(doc_type: str, doc_id: int):
    StockReservation.query.filter_by(doc_type=(doc_type or "").upper(), doc_id=doc_id, status="ACTIVE").delete(synchronize_session=False)

def _release_reservation_for_item(sales_item_id: int):
    StockReservation.query.filter_by(sales_item_id=sales_item_id, status="ACTIVE").delete(synchronize_session=False)


def _warehouse_choices() -> list[str]:
    whs = [r[0] for r in db.session.query(func.coalesce(Equipment.warehouse, "MAIN")).distinct().all()]
    whs = sorted({(w or "MAIN") for w in whs})
    return whs or ["MAIN"]

def _brand_choices_by_category(category_id: int, warehouse: str | None = None) -> list[str]:
    wh = (warehouse or "").strip() or "MAIN"
    rows = (
        db.session.query(func.coalesce(Equipment.brand, ""))
        .filter(Equipment.category_id == category_id)
        .filter(Equipment.warehouse == wh)
        .distinct()
        .all()
    )
    brands = sorted({(r[0] or "").strip() for r in rows})
    return brands
# ---------- Warehouse / Stock (Realtime) ----------

@app.get("/warehouse/stock")
@permission_required("equipment.view")
def warehouse_stock():
    """หน้าคลังสินค้า (สรุป Stock แบบ Realtime) + แยกตาม 'คลัง/สาขา' + RESERVED/AVAILABLE"""
    q = (request.args.get("q") or "").strip()
    wh_filter = (request.args.get("warehouse") or "").strip()

    # รายชื่อคลังทั้งหมด (ดึงจากทั้ง equipment และ reservation เพื่อให้ครบ)
    whs_e = [r[0] for r in db.session.query(func.coalesce(Equipment.warehouse, "MAIN")).distinct().all()]
    whs_r = [r[0] for r in db.session.query(func.coalesce(StockReservation.warehouse, "MAIN")).distinct().all()]
    warehouses = sorted({(w or "MAIN") for w in (whs_e + whs_r)})

    # --- aggregate equipment ---
    equip_rows = (
        db.session.query(
            Equipment.category_id.label("category_id"),
            func.coalesce(Equipment.warehouse, "MAIN").label("warehouse"),
            func.coalesce(Equipment.brand, "").label("brand"),
            func.sum(case((Equipment.status == "READY", 1), else_=0)).label("ready_qty"),
            func.sum(case((Equipment.status == "RENTED", 1), else_=0)).label("rented_qty"),
            func.sum(case((Equipment.status == "REPAIR", 1), else_=0)).label("repair_qty"),
            func.sum(case((Equipment.status == "LOST", 1), else_=0)).label("lost_qty"),
            func.count(Equipment.id).label("total_in"),
        )
        .group_by(Equipment.category_id, func.coalesce(Equipment.warehouse, "MAIN"), func.coalesce(Equipment.brand, ""))
        .all()
    )

    # --- aggregate reservations ---
    resv_rows = (
        db.session.query(
            StockReservation.category_id.label("category_id"),
            func.coalesce(StockReservation.warehouse, "MAIN").label("warehouse"),
            func.coalesce(StockReservation.brand, "").label("brand"),
            func.coalesce(func.sum(StockReservation.qty), 0.0).label("reserved_qty"),
        )
        .filter(StockReservation.status == "ACTIVE")
        .group_by(StockReservation.category_id, func.coalesce(StockReservation.warehouse, "MAIN"), func.coalesce(StockReservation.brand, ""))
        .all()
    )

    cat_map = {c.id: c for c in Category.query.all()}

    # merge
    data = {}
    def _k(cat_id, wh, br):
        return (int(cat_id), (wh or "MAIN"), (br or ""))

    for r in equip_rows:
        key=_k(r.category_id, r.warehouse, r.brand)
        data.setdefault(key, {
            "cat_id": r.category_id, "warehouse": r.warehouse, "brand": r.brand,
            "ready_qty": 0, "rented_qty": 0, "repair_qty": 0, "lost_qty": 0, "total_in": 0, "reserved_qty": 0.0,
        })
        d=data[key]
        d["ready_qty"]=int(r.ready_qty or 0)
        d["rented_qty"]=int(r.rented_qty or 0)
        d["repair_qty"]=int(r.repair_qty or 0)
        d["lost_qty"]=int(r.lost_qty or 0)
        d["total_in"]=int(r.total_in or 0)

    for r in resv_rows:
        key=_k(r.category_id, r.warehouse, r.brand)
        data.setdefault(key, {
            "cat_id": r.category_id, "warehouse": r.warehouse, "brand": r.brand,
            "ready_qty": 0, "rented_qty": 0, "repair_qty": 0, "lost_qty": 0, "total_in": 0, "reserved_qty": 0.0,
        })
        data[key]["reserved_qty"]=float(r.reserved_qty or 0.0)

    # build rows
    rows=[]
    for (cat_id, wh, br), d in data.items():
        cat=cat_map.get(cat_id)
        if not cat:
            continue
        if wh_filter and wh != wh_filter:
            continue
        d = dict(d)
        d["cat_name"]=cat.name
        d["cat_prefix"]=cat.prefix_sku
        rows.append(SimpleNamespace(
            cat_id=cat_id,
            cat_name=cat.name,
            cat_prefix=cat.prefix_sku,
            warehouse=wh,
            brand=(br or ""),
            ready_qty=int(d["ready_qty"]),
            reserved_qty=float(d["reserved_qty"]),
            rented_qty=int(d["rented_qty"]),
            repair_qty=int(d["repair_qty"]),
            lost_qty=int(d["lost_qty"]),
            total_in=int(d["total_in"]),
        ))

    # filter by q (ชื่อหมวด/Prefix/Brand/คลัง)
    if q:
        ql=q.lower()
        rows=[r for r in rows if (ql in (r.cat_name or "").lower()) or (ql in (r.cat_prefix or "").lower()) or (ql in (r.brand or "").lower()) or (ql in (r.warehouse or "").lower())]

    rows = sorted(rows, key=lambda r: (r.warehouse, r.cat_name or "", r.brand or ""))

    return render_template("warehouse/stock.html", ep="warehouse_stock", rows=rows, q=q, warehouse=wh_filter, warehouses=warehouses)


@app.get("/api/warehouse/stock")
@permission_required("equipment.view")
def api_warehouse_stock():
    """API สำหรับดึง stock แบบ realtime ตาม category/prefix/brand (READY/RESERVED/AVAILABLE)"""
    cat_id = request.args.get("category_id", type=int)
    prefix = (request.args.get("prefix") or "").strip()
    brand = (request.args.get("brand") or "").strip()

    # Resolve category by prefix if needed
    if (not cat_id) and prefix:
        cat = Category.query.filter(Category.prefix_sku == prefix).first()
        cat_id = cat.id if cat else None

    if not cat_id:
        return jsonify({"ok": False, "error": "missing category_id/prefix"}), 400

    warehouse = (request.args.get('warehouse') or '').strip() or 'MAIN'

    ready_total = _ready_total(cat_id, brand or None, warehouse)
    reserved = _reserved_qty(cat_id, brand or None, warehouse)
    available = max(0.0, float(ready_total) - float(reserved))

    return jsonify({
        "ok": True,
        "category_id": cat_id,
        "brand": brand,
        "ready": int(ready_total),
        "reserved": float(reserved),
        "available": float(available),
    })
    # ---------- Equipment (New) ----------
@app.route("/equipment/new", methods=["GET","POST"])
@permission_required("equipment.manage")
def equip_new():
    cats = Category.query.order_by(Category.name.asc()).all()

    # รายการอุปกรณ์ "รอเพิ่มเข้าระบบ" (มาจากการรับสินค้าเข้า / GRN)
    q = IncomingEquipment.query.filter_by(status="PENDING")
    # ถ้ามาจากหน้า GRN จะส่ง ?grn_id=... มา เพื่อกรองรายการเฉพาะ GRN นั้น
    grn_id = request.args.get("grn_id", type=int)
    if grn_id:
        q = q.filter(IncomingEquipment.grn_id == grn_id)
    pending = q.order_by(IncomingEquipment.id.desc()).all()

    # ถ้ามาจากปุ่ม "เพิ่ม" จะส่ง incoming_id มา
    # NOTE: ฟอร์ม POST ของบางเทมเพลตอาจไม่ได้ส่ง incoming_id กลับมา ทำให้ mark DONE ไม่ทำงาน
    # เราจึงอ่านจาก request.values (รวม args+form) และทำ fallback จับคู่จากข้อมูลที่กรอก
    incoming_id = request.values.get("incoming_id", type=int)
    inc = IncomingEquipment.query.get(incoming_id) if incoming_id else None
    if inc and (inc.status or "").upper() != "PENDING":
        inc = None

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        brand = (request.form.get("brand") or "").strip()
        cat_id = request.form.get("category_id", type=int)
        received = request.form.get("received_date")
        cost = request.form.get("cost", type=float) or 0.0
        ly = request.form.get("life_years", type=int) or 0
        lm = request.form.get("life_months", type=int) or 0
        ld = request.form.get("life_days", type=int) or 0
        if not name or not cat_id or not received:
            flash("กรอกข้อมูลให้ครบ (ชื่อ/หมวดหมู่/วันที่รับเข้า)", "danger")
            return redirect(url_for("equip_new"))

        cat = Category.query.get_or_404(cat_id)
        rdate = datetime.fromisoformat(received).date()
        sku = gen_sku(cat.prefix_sku, rdate)

        img_path = ""
        f = request.files.get("image")
        if f and f.filename:
            try:
                img_path = _save_image(f, f"equip_{sku}")
            except ValueError as e:
                flash(str(e), "warning")
                return redirect(url_for("equip_new"))

        eq = Equipment(
            sku=sku, name=name, brand=brand,
            warehouse=((request.form.get('warehouse') or '').strip() or 'MAIN'),
            category_id=cat.id,
            received_date=rdate, cost=cost,
            life_years=ly, life_months=lm, life_days=ld,
            image_path=img_path, status="READY",
        )
        db.session.add(eq)
        db.session.flush()

        # ถ้าสร้างจากรายการรอเพิ่ม -> ปิดรายการนั้น
        # NOTE: บางเทมเพลตอาจไม่ได้ส่ง incoming_id ตอน POST
        # เราจะพยายาม "เดา" รายการรอเพิ่มจากข้อมูลที่กรอก (name/brand/cost/date)
        if not inc:
            try:
                q = IncomingEquipment.query.filter_by(status="PENDING")
                if name:
                    q = q.filter(IncomingEquipment.name == name)
                if brand:
                    q = q.filter(IncomingEquipment.brand == brand)
                # match cost ใกล้เคียง
                q = q.order_by(IncomingEquipment.id.asc())
                cand = q.first()
                if cand is not None:
                    cand_cost = float(getattr(cand, "unit_cost", 0) or 0)
                    # date match ถ้ามี
                    ok_date = True
                    if rdate and getattr(cand, "received_date", None):
                        ok_date = (cand.received_date == rdate)
                    if abs(cand_cost - float(cost or 0)) < 0.01 and ok_date:
                        inc = cand
            except Exception:
                pass

        if inc:
            inc.status = "DONE"

        db.session.add(EquipmentLog(
            equipment_id=eq.id,
            action="ADD",
            note=("เพิ่มอุปกรณ์ (มาจากรับสินค้าเข้า)" if inc else "เพิ่มอุปกรณ์"),
            user_id=(current_user.id if current_user.is_authenticated else None),
        ))
        db.session.commit()
        flash("เพิ่มอุปกรณ์แล้ว", "success")
        return redirect(url_for("equip_view", eid=eq.id))

    # GET: เตรียมค่า preset อัตโนมัติจากรายการรอเพิ่ม
    preset = {}
    if inc:
        preset = {
            "name": inc.name,
            "brand": inc.brand,
            "received_date": (inc.received_date.isoformat() if inc.received_date else date.today().isoformat()),
            "cost": float(inc.unit_cost or 0),
        }

    return render_template(
        "equipment/equip_form.html",
        cats=cats,
        pending=pending,
        preset=preset,
        incoming=inc,
        warehouses=_warehouse_choices(),
    )

@app.route("/equipment/<int:eid>")
@permission_required("equipment.view")
def equip_view(eid):
    e = Equipment.query.get_or_404(eid)
    logs = EquipmentLog.query.filter_by(equipment_id=e.id).order_by(EquipmentLog.created_at.desc()).all()
    return render_template("equipment/equip_view.html", e=e, logs=logs, status_th=EQUIP_STATUS_THAI)

@app.route("/equipment/<int:eid>/edit", methods=["GET","POST"])
@permission_required("equipment.manage")
def equip_edit(eid):
    e = Equipment.query.get_or_404(eid)
    cats = Category.query.order_by(Category.name.asc()).all()
    if request.method == "POST":
        e.name = (request.form.get("name") or "").strip()
        e.brand = (request.form.get("brand") or "").strip()
        e.warehouse = ((request.form.get('warehouse') or '').strip() or 'MAIN')
        e.category_id = request.form.get("category_id", type=int) or e.category_id
        received = request.form.get("received_date")
        e.cost = request.form.get("cost", type=float) or 0.0
        e.life_years  = request.form.get("life_years", type=int) or 0
        e.life_months = request.form.get("life_months", type=int) or 0
        e.life_days   = request.form.get("life_days", type=int) or 0
        new_status = (request.form.get("status") or e.status).upper()
        if new_status in EQUIP_STATUS and new_status != e.status:
            e.status = new_status
            db.session.add(EquipmentLog(
                equipment_id=e.id,
                action="STATUS",
                note=f"สถานะเป็น {EQUIP_STATUS_THAI[new_status]}",
                user_id=(current_user.id if current_user.is_authenticated else None),
            ))
        if received:
            e.received_date = datetime.fromisoformat(received).date()
            cat = Category.query.get(e.category_id)
            if cat:
                e.sku = gen_sku(cat.prefix_sku, e.received_date)
        f = request.files.get("image")
        if f and f.filename:
            try:
                e.image_path = _save_image(f, f"equip_{e.sku}")
            except ValueError as ex:
                flash(str(ex), "warning")
                return redirect(url_for("equip_edit", eid=e.id))
        db.session.add(EquipmentLog(
            equipment_id=e.id,
            action="EDIT",
            note="แก้ไขรายละเอียด",
            user_id=(current_user.id if current_user.is_authenticated else None),
        ))
        db.session.commit()
        flash("บันทึกแล้ว", "success")
        return redirect(url_for("equip_view", eid=e.id))
    return render_template("equipment/equip_form.html", cats=cats, e=e, status_th=EQUIP_STATUS_THAI, warehouses=_warehouse_choices())

@app.post("/equipment/<int:eid>/delete")
@permission_required("equipment.manage")
def equip_delete(eid):
    e = Equipment.query.get_or_404(eid)
    db.session.query(EquipmentLog).filter_by(equipment_id=e.id).delete(synchronize_session=False)
    db.session.delete(e); db.session.commit()
    flash("ลบอุปกรณ์แล้ว", "success")
    return redirect(url_for("equip_list"))

# ---------- Promotions ----------
@app.route("/promos")
@permission_required("promos.view")
def promos_list():
    q = (request.args.get("q") or "").strip()
    qry = Promotion.query
    if q:
        qry = qry.filter(Promotion.name.ilike(f"%{q}%"))
    promos = qry.order_by(Promotion.id.desc()).all()
    return render_template("promos/promo_list.html", promos=promos, q=q)

@app.route("/promos/new", methods=["GET","POST"])
@permission_required("promos.manage")
def promo_new():
    if request.method == "POST":
        p = Promotion(
            name=(request.form.get("name") or "").strip(),
            active=bool(request.form.get("active")),
            start_date=(datetime.fromisoformat(request.form["start_date"]).date()
                        if request.form.get("start_date") else None),
            end_date=(datetime.fromisoformat(request.form["end_date"]).date()
                        if request.form.get("end_date") else None),
            min_items=request.form.get("min_items", type=int) or 0,
            rental_unit=(request.form.get("rental_unit") or "DAY").upper(),
            min_duration=request.form.get("min_duration", type=int) or 0,
            discount_type=(request.form.get("discount_type") or "PCT").upper(),
            discount_value=request.form.get("discount_value", type=float) or 0.0,
            cheapest_units_to_discount=request.form.get("cheapest_units_to_discount", type=int) or 1,
            note=(request.form.get("note") or "").strip()
        )
        if not p.name:
            flash("กรุณากรอกชื่อโปร", "danger"); return redirect(url_for("promo_new"))
        db.session.add(p); db.session.commit()
        flash("เพิ่มโปรโมชั่นแล้ว", "success")
        return redirect(url_for("promos_list"))
    return render_template("promos/promo_form.html", p=None)

@app.route("/promos/<int:pid>/edit", methods=["GET","POST"])
@permission_required("promos.manage")
def promo_edit(pid):
    p = Promotion.query.get_or_404(pid)
    if request.method == "POST":
        p.name = (request.form.get("name") or "").strip()
        p.active = bool(request.form.get("active"))
        p.start_date = (datetime.fromisoformat(request.form["start_date"]).date()
                        if request.form.get("start_date") else None)
        p.end_date = (datetime.fromisoformat(request.form["end_date"]).date()
                        if request.form.get("end_date") else None)
        p.min_items = request.form.get("min_items", type=int) or 0
        p.rental_unit = (request.form.get("rental_unit") or "DAY").upper()
        p.min_duration = request.form.get("min_duration", type=int) or 0
        p.discount_type = (request.form.get("discount_type") or "PCT").upper()
        p.discount_value = request.form.get("discount_value", type=float) or 0.0
        p.cheapest_units_to_discount = request.form.get("cheapest_units_to_discount", type=int) or 1
        p.note = (request.form.get("note") or "").strip()
        if not p.name:
            flash("กรุณากรอกชื่อโปร", "danger"); return redirect(url_for("promo_edit", pid=p.id))
        db.session.commit()
        flash("บันทึกโปรโมชั่นแล้ว", "success")
        return redirect(url_for("promos_list"))
    return render_template("promos/promo_form.html", p=p)

@app.post("/promos/<int:pid>/delete")
@permission_required("promos.manage")
def promo_delete(pid):
    p = Promotion.query.get_or_404(pid)
    db.session.delete(p); db.session.commit()
    flash("ลบโปรโมชั่นแล้ว", "success")
    return redirect(url_for("promos_list"))

def _best_promo_today() -> list[Promotion]:
    today = date.today()
    return (Promotion.query
            .filter_by(active=True)
            .filter((Promotion.start_date==None) | (Promotion.start_date<=today))
            .filter((Promotion.end_date==None)   | (Promotion.end_date>=today))
            .order_by(Promotion.id.desc())
            .all())

def _choose_best_promo_for_items(items: list[dict], rental_days: int | None):
    """
    ประเมินทุกโปรที่ active วันนี้ แล้วเลือกตัวที่เหมาะสุด
    เกณฑ์เรียง:
      1) ส่วนลดบาทมากสุด
      2) ถ้าส่วนลดเท่ากัน → min_items มากกว่า
      3) ถ้าเท่ากันอีก → min_duration (แปลงเป็นวัน) มากกว่า
      4) ถ้ายังเท่ากัน → id มากกว่า (อันใหม่กว่า)
    """
    promos = _best_promo_today()
    best = None
    best_disc = 0.0
    best_key = None  # key เอาไว้เทียบ

    for p in promos:
        disc = float(compute_promo_discount(items, rental_days=rental_days, promo=p) or 0.0)
        if disc <= 0:
            continue  # โปรนี้ไม่เข้าเงื่อนไข

        # แปลงเงื่อนไขวันขั้นต่ำของโปรเป็นจำนวนวันจริง
        min_days_required = _unit_to_days(p.rental_unit or "DAY", p.min_duration or 0)

        # key สำหรับจัดลำดับ
        key = (
            disc,
            p.min_items or 0,
            min_days_required,
            p.id or 0,
        )

        if best is None or key > best_key:
            best = p
            best_disc = disc
            best_key = key

    return best, best_disc


@app.post("/api/promos/evaluate")
@permission_required("promos.view")
def api_promos_evaluate():
    """รับ items ปัจจุบันจากหน้าแบบฟอร์ม → ประเมินทุกโปรที่เปิดวันนี้ → เลือกโปรที่ดีที่สุด"""
    data = request.get_json(force=True) or {}
    items = data.get("items") or []
    rental_days = data.get("rental_days")  # ส่งมาก็ได้ ไม่ส่งมาก็ให้ None

    promos = _best_promo_today()
    best = None
    best_disc = 0.0
    best_key = None  # ใช้เก็บ key สำหรับเทียบว่าโปรไหน "ดีกว่า"

    for p in promos:
        disc = float(compute_promo_discount(items, rental_days=rental_days, promo=p) or 0.0)

        # ถ้าส่วนลดไม่เข้าเงื่อนไข หรือได้ 0 ก็ข้ามโปรนี้ไปเลย
        if disc <= 0:
            continue

        # แปลงเงื่อนไขวันขั้นต่ำของโปรเป็นจำนวนวันจริง ๆ เพื่อเอาไว้เทียบ
        min_days_required = _unit_to_days(p.rental_unit or "DAY", p.min_duration or 0)

        # key สำหรับจัดลำดับโปร:
        #   1) ส่วนลดบาทมากสุด
        #   2) ถ้าส่วนลดเท่ากัน → min_items มากกว่า
        #   3) ถ้ายังเท่ากัน → min_days_required มากกว่า
        #   4) ถ้ายังเท่ากัน → id ใหม่กว่า
        key = (
            disc,
            p.min_items or 0,
            min_days_required,
            p.id or 0,
        )

        if best is None or key > best_key:
            best = p
            best_disc = disc
            best_key = key

    if not best or best_disc <= 0:
        return jsonify({
            "ok": True,
            "hasPromo": False,
            "message": "ไม่มีโปรโมชันที่เข้าเงื่อนไข"
        })

    return jsonify({
        "ok": True,
        "hasPromo": True,
        "promo": {"id": best.id, "name": best.name},
        "discount": round(best_disc, 2),
    })



def _apply_discount_as_negative_line(d: SalesDoc, amount: float, label: str):
    if amount <= 0:
        return
    line = SalesItem(
        doc=d, name=f"[PROMO] {label}",
        qty=1, rent_unit="DAY", rent_duration=1,
        unit_price=-amount, discount_pct=0,
        line_subtotal=-amount, line_total=-amount
    )
    db.session.add(line)
    _calc_sales_totals(d)


@app.post("/sales/quotes/<int:qid>/check_promo")
@permission_required("sales.manage")
def qu_check_promo(qid):
    """เช็คและนำโปรโมชันที่ดีที่สุดมาใช้กับใบเสนอราคา qid"""
    d = (
        SalesDoc.query
        .options(joinedload(SalesDoc.items), joinedload(SalesDoc.customer))
        .get_or_404(qid)
    )

    items = _items_from_doc(d)

    promos = _best_promo_today()
    if not promos:
        flash("ไม่พบโปรโมชันที่เปิดใช้งานวันนี้", "warning")
        return redirect(url_for("qu_view", qid=d.id))

    # ประเมินทุกโปร เลือกโปร "ที่ดีที่สุด"
    # เกณฑ์:
    #   1) ส่วนลดบาทมากสุด
    #   2) ถ้าส่วนลดเท่ากัน → min_items มากกว่า
    #   3) ถ้ายังเท่ากัน → min_duration (แปลงเป็นวัน) มากกว่า
    #   4) ถ้ายังเท่ากัน → id ใหม่กว่า (โปรใหม่กว่า)
    best = None
    best_disc = 0.0
    best_key = None

    for p in promos:
        disc = float(
            compute_promo_discount(items, rental_days=None, promo=p) or 0.0
        )
        if disc <= 0:
            # โปรนี้ไม่เข้าเงื่อนไข / ไม่มีส่วนลดจริง ข้าม
            continue

        # แปลงเงื่อนไขวันขั้นต่ำของโปรเป็นจำนวนวัน เพื่อใช้เทียบกัน
        min_days_required = _unit_to_days(
            p.rental_unit or "DAY",
            p.min_duration or 0,
        )

        key = (
            disc,
            p.min_items or 0,
            min_days_required,
            p.id or 0,
        )

        if best is None or key > best_key:
            best = p
            best_disc = disc
            best_key = key

    if best and best_disc > 0:
        # ใส่ส่วนลดเป็นแถวลบ และเขียนบันทึกในหมายเหตุ
        _apply_discount_as_negative_line(d, best_disc, best.name)
        note = f"[AUTO PROMO] ใช้โปร '{best.name}' ลด {best_disc:,.2f} บาท"
        d.remark = (d.remark + "\n" + note).strip() if d.remark else note
        db.session.commit()
        flash(
            f"นำส่วนลดจากโปร '{best.name}' มาใช้แล้ว ({best_disc:,.2f} บาท) "
            f"และบันทึกข้อความในหมายเหตุ",
            "success",
        )
    else:
        flash("ไม่มีโปรโมชันที่เข้าเงื่อนไขสำหรับเอกสารนี้", "info")

    return redirect(url_for("qu_view", qid=d.id))



# ---------- Sales: Quotes ----------

# ---------------------------
# Compatibility: old /quotes URL -> /sales/quotes
# ---------------------------
@app.get("/quotes")
@app.get("/quotes/")
def quotes_legacy_redirect():
    return redirect(url_for("qu_list"))


@app.route("/sales/quotes")
@permission_required("sales.view")
def qu_list():
    q = (request.args.get("q") or "").strip()

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    # เลือกเฉพาะเอกสาร QU + filter ชื่อลูกค้าตาม q เหมือนเดิม
    qry = SalesDoc.query.filter(SalesDoc.doc_type == "QU")
    if start_d:
        qry = qry.filter(SalesDoc.date >= start_d)
    if end_d:
        qry = qry.filter(SalesDoc.date <= end_d)
    if q:
        qry = qry.join(Customer).filter(Customer.name.ilike(f"%{q}%"))

    rows = qry.order_by(SalesDoc.id.desc()).all()

    # --- เพิ่มส่วนดึงข้อมูลใบส่งสินค้าที่ถูกสร้างจากใบเสนอราคาแต่ละใบ ---
    ids = [d.id for d in rows]
    deliveries_map = {}
    if ids:
        dls = (
            DeliveryDoc.query
            .filter(
                DeliveryDoc.source_type == "QUOTATION",
                DeliveryDoc.source_id.in_(ids),
            )
            .all()
        )
        # map: key = id ของใบเสนอราคา, value = DeliveryDoc ที่สร้างจากใบนั้น
        deliveries_map = {d.source_id: d for d in dls}

    return render_template(
        "sales/qu_list.html",
        rows=rows,
        q=q,
        start=start,
        end=end,
        deliveries_map=deliveries_map,
        doc_type="QU",   # ✅ ส่งไปให้ template ใช้เช็คว่ามีใบส่งแล้วหรือยัง
    )



# ---------- Accounts Receivable (AR) / Aging ----------
def _doc_credit_days_for_customer(doc: "SalesDoc", cust: "Customer") -> int:
    """เครดิตที่ใช้คำนวณกำหนดชำระ: ใช้ของเอกสารก่อน ถ้าไม่กำหนดใช้ของลูกค้า"""
    try:
        d = int((doc.credit_days or 0) or 0)
    except Exception:
        d = 0
    if d <= 0:
        try:
            d = int((cust.credit_term_days or 0) or 0)
        except Exception:
            d = 0
    return max(0, d)

def _customer_ar_documents(cust: "Customer", warehouse: str | None = None):
    """
    เอกสารลูกหนี้: ใช้ BL/IV ที่ยังไม่ชำระ (ไม่รวม CANCELLED)
    คืน list[dict] สำหรับ render/report
    """
    q = SalesDoc.query.filter(SalesDoc.customer_id == cust.id)
    q = q.filter(SalesDoc.doc_type.in_(["BL", "IV"]))
    q = q.filter(SalesDoc.status.notin_(["PAID", "CANCELLED"]))
    if warehouse:
        q = q.filter(SalesDoc.warehouse == warehouse)
    docs = q.order_by(SalesDoc.date.asc()).all()

    today = date.today()
    rows = []
    for d in docs:
        credit = _doc_credit_days_for_customer(d, cust)
        due = (d.date or today) + timedelta(days=credit)
        overdue_days = (today - due).days
        bucket = "CURRENT" if overdue_days <= 0 else (
            "1-30" if overdue_days <= 30 else
            "31-60" if overdue_days <= 60 else
            "61-90" if overdue_days <= 90 else
            "90+"
        )
        rows.append({
            "id": d.id,
            "number": d.number,
            "doc_type": d.doc_type,
            "status": d.status,
            "warehouse": d.warehouse,
            "date": d.date,
            "credit_days": credit,
            "due_date": due,
            "overdue_days": max(0, overdue_days),
            "bucket": bucket,
            "amount": float(d.amount_grand or d.amount_total or 0.0),
            "url": url_for("sales_doc_view", sid=d.id) if "sales_doc_view" in app.view_functions else None,
        })
    return rows

def _aging_summary(rows):
    s = {"CURRENT": 0.0, "1-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0, "TOTAL": 0.0}
    for r in rows:
        amt = float(r.get("amount") or 0.0)
        s["TOTAL"] += amt
        s[r["bucket"]] += amt
    return s

@app.get("/customers/<int:cid>/aging")
@permission_required("customers.view")
def customers_aging(cid):
    c = Customer.query.get_or_404(cid)
    wh = (request.args.get("warehouse") or "").strip() or None
    rows = _customer_ar_documents(c, warehouse=wh)
    summary = _aging_summary(rows)
    return render_template(
        "customers/customers_aging.html",
        c=c,
        rows=rows,
        summary=summary,
        warehouse=wh,
    )

@app.get("/reports/ar-aging")
@permission_required("customers.view")
def report_ar_aging():
    wh = (request.args.get("warehouse") or "").strip() or None
    q = Customer.query.order_by(Customer.name.asc()).all()
    data = []
    for c in q:
        rows = _customer_ar_documents(c, warehouse=wh)
        if not rows:
            continue
        data.append({
            "customer": c,
            "summary": _aging_summary(rows),
            "rows_count": len(rows),
        })
    # เรียงลูกหนี้มาก -> น้อย
    data.sort(key=lambda x: x["summary"]["TOTAL"], reverse=True)
    return render_template("reports/ar_aging.html", data=data, warehouse=wh)

@app.route("/sales/quotes/new", methods=["GET", "POST"])
@permission_required("sales.manage")
def qu_new():
    customers = Customer.query.order_by(Customer.name.asc()).all()
    import re
    def _extract_sku(title: str) -> str | None:
        m = re.search(r"\[([^\[\]]+?)\]", title or "")
        return m.group(1).strip() if m else None
    def _to_float(x, default=0.0) -> float:
        try: return float(x)
        except Exception: return float(default)
    def _to_int(x, default=1):  # เดิม default=0
        try: return int(x)
        except Exception: return int(default)

    def _try_parse_date(s: str | None):
        try:
            s = (s or "").strip()
            if not s:
                return None
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None


    if request.method == "POST":
        cid = request.form.get("customer_id", type=int) or 0
        if not cid:
            flash("กรุณาเลือกลูกค้า", "danger")
            return redirect(url_for("qu_new"))
        doc = SalesDoc(
            number=_gen_sales_running("QU"),
            doc_type="QU",
            status="DRAFT",
            customer_id=cid,
            warehouse=((request.form.get('warehouse') or '').strip() or 'MAIN'),
            project_name=(request.form.get("project_name") or "").strip(),
            po_customer=(request.form.get("po_customer") or "").strip(),
            credit_days=(request.form.get("credit_days", type=int) if (request.form.get("credit_days") or "").strip() != "" else (Customer.query.get(cid).credit_term_days or 0)),
            tax_mode=(request.form.get("tax_mode") or "EXC").upper(),
            wht_pct=request.form.get("wht_pct", type=int) or 0,
            date=date.today(),
            remark=(request.form.get("remark") or "").strip(),
            billing_mode=((request.form.get("billing_mode") or "ONCE").upper()),
            installment_count=(request.form.get("installment_count", type=int) or 0),
            contract_start=_try_parse_date(request.form.get("contract_start")),
            contract_end=_try_parse_date(request.form.get("contract_end")),
        )
        db.session.add(doc)
        db.session.flush()
        names   = request.form.getlist("name[]")
        brands_ = request.form.getlist("brand[]")
        qtys    = request.form.getlist("qty[]")
        units   = request.form.getlist("unit[]")
        durs    = request.form.getlist("duration[]")
        prices  = request.form.getlist("price[]")
        dps     = request.form.getlist("disc[]")
        added_count = 0
        for i, n in enumerate(names):
            n = (n or "").strip()
            if not n:
                continue
            image_path = ""
            cat_id = None
            cat_prefix = None
            sku = _extract_sku(n)
            # หา prefix/หมวดจากชื่อรายการ (เช่น "สว่าน [SPTE6901]" หรือ "สว่าน [SPTE6901-001]")
            cat_prefix = None
            cat_id = None
            pref = _extract_sku(n)
            if pref:
                # ถ้าเป็น SKU เต็มให้ตัดเป็น prefix ก่อน "-" เช่น SPTE6901-001 -> SPTE6901
                p = pref.split("-")[0].strip()
                cat_prefix = p
                cat = Category.query.filter_by(prefix_sku=p).first()
                if cat:
                    cat_id = cat.id

            if sku:
                # ตอนทำ QU ให้เลือก "หมวดหมู่" จาก prefix_sku เท่านั้น (ยังไม่ล็อคตัวอุปกรณ์จริง)
                base_sku = (sku.split("-")[0].strip() if sku else "")
                cat = Category.query.filter_by(prefix_sku=base_sku).first()
                if cat:
                    cat_id = cat.id
                    cat_prefix = (cat.prefix_sku or "").strip()
                    # ปรับชื่อรายการให้เป็นชื่อหมวด (ไว้ให้ลูกค้าอ่านง่าย)
                    n = f"{cat.name} [{cat_prefix}]" if cat_prefix else (cat.name or n)

            db.session.add(SalesItem(
                doc_id=doc.id,
                name=n,
                image_path=image_path,
                category_id=cat_id,
                category_prefix=cat_prefix,
                brand=(_norm_brand(brands_[i] if i < len(brands_) else "") or None),
                qty=_to_float(qtys[i] if i < len(qtys) else 0),
                rent_unit=((units[i] if i < len(units) else "DAY") or "DAY").upper(),
                rent_duration=_to_int(durs[i] if i < len(durs) else 1),
                unit_price=_to_float(prices[i] if i < len(prices) else 0),
                discount_pct=_to_float(dps[i] if i < len(dps) else 0),
            ))
            added_count += 1
        if added_count == 0:
            db.session.rollback()
            flash("กรุณาใส่อย่างน้อย 1 รายการ", "danger")
            return redirect(url_for("qu_new"))
        db.session.flush()
        # Reserve stock at line-level (B)
        try:
            for it in SalesItem.query.filter_by(doc_id=doc.id).all():
                if it.category_id:
                    _check_and_reserve("QU", doc.id, it.id, it.category_id, it.brand, doc.warehouse, it.qty or 0)
        except ValueError as e:
            db.session.rollback()
            flash(str(e), "danger")
            return redirect(url_for("qu_new"))

        _calc_sales_totals(doc)
        db.session.commit()
        flash("บันทึกใบเสนอราคาแล้ว (DRAFT)", "success")
        return redirect(url_for("qu_view", qid=doc.id))
    brands = [b[0] for b in db.session.query(Equipment.brand).filter(Equipment.brand != "").distinct().order_by(Equipment.brand.asc()).all()]
    return render_template("sales/qu_form.html", customers=customers, brands=brands, warehouses=_warehouse_choices(), warehouse_default="MAIN")


@app.route("/sales/quotes/<int:qid>/edit", methods=["GET","POST"])
@permission_required("sales.manage")
def qu_edit(qid):
    doc = (
        SalesDoc.query
        .options(joinedload(SalesDoc.items), joinedload(SalesDoc.customer))
        .get_or_404(qid)
    )
    if doc.doc_type != "QU":
        abort(404)
    if doc.status == "APPROVED":
        flash("ใบเสนอราคาอนุมัติแล้ว ไม่อนุญาตให้แก้ไข", "warning")
        return redirect(url_for("qu_view", qid=doc.id))

    customers = Customer.query.order_by(Customer.name.asc()).all()
    brands = [b[0] for b in db.session.query(Equipment.brand).filter(Equipment.brand != "").distinct().order_by(Equipment.brand.asc()).all()]

    import re
    def _extract_token_in_brackets(title: str) -> str | None:
        m = re.search(r"\[([^\[\]]+?)\]", title or "")
        return m.group(1).strip() if m else None

    def _to_float(x, default=0.0) -> float:
        try: return float(x)
        except Exception: return float(default)

    def _to_int(x, default=1):
        try: return int(float(x))
        except Exception: return int(default)

    def _try_parse_date(s: str | None):
        try:
            s = (s or "").strip()
            if not s:
                return None
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    if request.method == "POST":
        cid = request.form.get("customer_id", type=int) or 0
        if not cid:
            flash("กรุณาเลือกลูกค้า", "danger")
            return redirect(url_for("qu_edit", qid=doc.id))

        # update header
        doc.customer_id = cid
        doc.project_name = (request.form.get("project_name") or "").strip()
        doc.warehouse = ((request.form.get('warehouse') or '').strip() or 'MAIN')
        doc.po_customer = (request.form.get("po_customer") or "").strip()
        doc.credit_days = request.form.get("credit_days", type=int) or 0
        doc.tax_mode = (request.form.get("tax_mode") or "EXC").upper()
        doc.wht_pct = request.form.get("wht_pct", type=int) or 0
        doc.remark = (request.form.get("remark") or "").strip()
        doc.billing_mode = ((request.form.get("billing_mode") or "ONCE").upper())
        doc.installment_count = (request.form.get("installment_count", type=int) or 0)
        doc.contract_start = _try_parse_date(request.form.get("contract_start"))
        doc.contract_end = _try_parse_date(request.form.get("contract_end"))

        # release old reservations + replace lines
        _release_reservations_for_doc("QU", doc.id)
        SalesItem.query.filter_by(doc_id=doc.id).delete(synchronize_session=False)

        names   = request.form.getlist("name[]")
        brands_ = request.form.getlist("brand[]")
        qtys    = request.form.getlist("qty[]")
        units   = request.form.getlist("unit[]")
        durs    = request.form.getlist("duration[]")
        prices  = request.form.getlist("price[]")
        dps     = request.form.getlist("disc[]")

        added_count = 0
        for i, n in enumerate(names):
            n = (n or "").strip()
            if not n:
                continue

            # หา category จาก token ใน [] (ถือเป็น prefix เช่น SPTE6901 หรือ SPTE6901-001)
            cat_id = None
            cat_prefix = None
            token = _extract_token_in_brackets(n)
            if token:
                p = token.split("-")[0].strip()
                cat_prefix = p
                cat = Category.query.filter_by(prefix_sku=p).first()
                if cat:
                    cat_id = cat.id

            # image_path: ถ้า token เป็น SKU เต็มและพบใน equipment -> ใช้รูป
            image_path = ""
            if token:
                eq = Equipment.query.filter_by(sku=token).first()
                if eq:
                    image_path = eq.image_path or ""

            line = SalesItem(
                doc_id=doc.id,
                name=n,
                category_id=cat_id,
                category_prefix=cat_prefix,
                brand=(_norm_brand(brands_[i] if i < len(brands_) else "") or None),
                qty=_to_int(qtys[i] if i < len(qtys) else 1, 1),
                rent_unit=((units[i] if i < len(units) else "DAY") or "DAY").upper(),
                rent_duration=_to_int(durs[i] if i < len(durs) else 1, 1),
                unit_price=_to_float(prices[i] if i < len(prices) else 0.0, 0.0),
                discount_pct=_to_float(dps[i] if i < len(dps) else 0.0, 0.0),
                image_path=image_path,
            )
            db.session.add(line)
            added_count += 1

        if added_count == 0:
            db.session.rollback()
            flash("กรุณาใส่อย่างน้อย 1 รายการ", "danger")
            return redirect(url_for("qu_edit", qid=doc.id))

        db.session.flush()

        # reserve ใหม่ตามบรรทัด (B)
        try:
            for it in SalesItem.query.filter_by(doc_id=doc.id).all():
                if it.category_id:
                    _check_and_reserve("QU", doc.id, it.id, it.category_id, it.brand, doc.warehouse, it.qty or 0)
        except ValueError as e:
            db.session.rollback()
            flash(str(e), "danger")
            return redirect(url_for("qu_edit", qid=doc.id))

        _calc_sales_totals(doc)
        db.session.commit()
        flash(f"บันทึกการแก้ไขใบเสนอราคาแล้ว ({added_count} รายการ)", "success")
        return redirect(url_for("qu_view", qid=doc.id))

    return render_template(
        "sales/qu_form.html",
        customers=customers,
        brands=brands,
        warehouses=_warehouse_choices(),
        warehouse_default=(doc.warehouse or 'MAIN'),
        doc=doc,
        items=doc.items,
        action_url=url_for("qu_edit", qid=doc.id),
        form_title="แก้ไขใบเสนอราคา",
        submit_label="บันทึกการแก้ไข",
    )

@app.post("/sales/quotes/<int:qid>/delete")
@permission_required("sales.manage")
def qu_delete(qid):
    doc = SalesDoc.query.get_or_404(qid)

    # กันลบเอกสารผิดประเภท
    if (doc.doc_type or "").upper() != "QU":
        flash("เอกสารนี้ไม่ใช่ใบเสนอราคา (QU)", "danger")
        return redirect(url_for("qu_list"))

    # แนะนำความปลอดภัย: อนุญาตให้ลบเฉพาะ DRAFT
    # ถ้าอยากลบได้ทุกสถานะ คุณค่อยแจ้งผม แล้วจะทำแบบ cascade/soft delete ให้
    if (doc.status or "").upper() not in ("DRAFT", "CANCELLED"):
        flash("ลบได้เฉพาะใบเสนอราคา DRAFT หรือ CANCELLED เท่านั้น (ถ้าต้องการลบสถานะอื่นบอกได้)", "warning")
        return redirect(url_for("qu_view", qid=doc.id))

    # ถ้ามีระบบจองสต๊อก ให้พยายามปล่อยจองก่อน
        # try:
        #     _release_reservations("QU", doc.id)
        # except Exception:
    #     pass

    # ลบรายการลูกก่อน กัน FK error
    SalesItem.query.filter_by(doc_id=doc.id).delete()

    db.session.delete(doc)
    db.session.commit()

    flash("ลบใบเสนอราคาเรียบร้อยแล้ว", "success")
    return redirect(url_for("qu_list"))


@permission_required("sales.manage")
def qu_delete(qid):
    doc = SalesDoc.query.get_or_404(qid)
    if doc.doc_type != "QU":
        abort(404)
    if doc.status == "APPROVED":
        flash("ใบเสนอราคาอนุมัติแล้ว ไม่อนุญาตให้ลบ", "warning")
        return redirect(url_for("qu_view", qid=doc.id))

    # ถ้ามีใบส่งสินค้าถูกสร้างจากใบนี้ ให้กันลบทิ้งเพื่อไม่ให้ข้อมูลขาด
    has_dl = DeliveryDoc.query.filter_by(source_type="QUOTATION", source_id=doc.id).first()
    if has_dl:
        flash("มีใบส่งสินค้าถูกสร้างจากใบเสนอราคานี้แล้ว ไม่อนุญาตให้ลบ", "warning")
        return redirect(url_for("qu_view", qid=doc.id))

    _release_reservations_for_doc("QU", doc.id)
    SalesItem.query.filter_by(doc_id=doc.id).delete(synchronize_session=False)
    db.session.delete(doc)
    db.session.commit()
    flash("ลบใบเสนอราคาแล้ว", "success")
    return redirect(url_for("qu_list"))


@app.route("/sales/quotes/<int:qid>")
@permission_required("sales.view")
def qu_view(qid):
    d = SalesDoc.query.get_or_404(qid)
    bk = SalesDoc.query.filter_by(parent_id=d.id, doc_type="BK").first()
    return render_template("sales/qu_view.html", d=d, bk=bk)


@app.post("/sales/quotes/<int:qid>/items/<int:item_id>/status")
@login_required
@permission_required("sales.manage")
def qu_item_set_status(qid, item_id):
    d = SalesDoc.query.options(joinedload(SalesDoc.items)).get_or_404(qid)
    if (d.doc_type or "").upper() != "QU":
        abort(400)

    it = SalesItem.query.filter_by(id=item_id, doc_id=d.id).first_or_404()

    status = (request.form.get("line_status") or "APPROVED").strip().upper()
    if status not in ("APPROVED", "REJECTED"):
        status = "APPROVED"

    it.line_status = status

    # ✅ ถ้ารายการถูก "ไม่อนุมัติ" ให้ยกเลิกการจองสต็อกของรายการนี้ (ถ้ามี)
    if status == "REJECTED":
        (StockReservation.query
         .filter_by(doc_type="QU", doc_id=d.id, sales_item_id=it.id)
         .update({"status": "CANCELLED", "qty": 0.0}, synchronize_session=False))
    else:
        # กลับมาอนุมัติ → ทำให้ reservation ACTIVE (ถ้ามี) และใส่ qty ให้ตรงกับ item.qty
        (StockReservation.query
         .filter_by(doc_type="QU", doc_id=d.id, sales_item_id=it.id)
         .update({"status": "ACTIVE", "qty": float(it.qty or 0)}, synchronize_session=False))

    db.session.commit()
    flash("อัปเดตสถานะรายการแล้ว", "success")
    return redirect(url_for("qu_view", qid=d.id))


@app.template_filter("unit_th")
def unit_th(v: str) -> str:
    m = {"HOUR": "ชั่วโมง", "DAY": "วัน", "MONTH": "เดือน", "YEAR": "ปี"}
    return m.get((v or "").upper(), v or "")

@app.template_filter("sale_status_th")
def sale_status_th(v: str) -> str:
    m = {
        "DRAFT": "ร่าง",
        "APPROVED": "อนุมัติแล้ว",
        "UNPAID": "ยังไม่ชำระเงิน",
        "PAID": "ชำระเงินแล้ว",
        "UNISSUED": "ยังไม่ออกเอกสาร",
        "ISSUED": "ออกเอกสารแล้ว",
        "PENDING": "รอดำเนินการ",
    }
    return m.get((v or "").upper(), v or "")

@app.template_filter("tax_mode_th")
def tax_mode_th(v: str) -> str:
    m = {"EXC": "ค่าของไม่รวมภาษี (+VAT 7%)", "INC": "รวมภาษีแล้ว", "NONE": "ไม่คิดภาษี"}
    return m.get((v or "").upper(), v or "")

# ===== Helper: หาชื่อลูกค้าปัจจุบันของอุปกรณ์ที่กำลังเช่า =====
# ===== Helper: หาชื่อลูกค้าปัจจุบันของอุปกรณ์ที่กำลังเช่า =====
@app.template_global("renting_customer_for_sku")
def renting_customer_for_sku(sku: str) -> str:
    """
    คืนชื่อ 'ลูกค้าที่เช่าอยู่' ของอุปกรณ์ตาม SKU

    วิธีหา:
      - ดูเอกสารขายทุกประเภท (QU / BL / IV / RC / DN / RN ฯลฯ)
      - ที่มี text ของ SKU นี้อยู่ในชื่อรายการสินค้า (SalesItem.name)
      - เลือกเอกสารที่ "วันที่" ล่าสุด และดึงชื่อลูกค้าออกมา

    ถ้าไม่เจออะไรเลย จะคืน "-"
    """
    try:
        if not sku:
            return "-"

        sku = sku.strip()

        # เผื่อเคสที่ตัวแปรโมเดลยังไม่อยู่ใน globals
        if "Customer" not in globals() or "SalesDoc" not in globals() or "SalesItem" not in globals():
            return "-"

        # หาทั้งแบบมี [] ครอบ และแบบเป็นข้อความดิบ ๆ
        like1 = f"%[{sku}]%"
        like2 = f"%{sku}%"

        row = (
            db.session.query(Customer.name)
            .join(SalesDoc, SalesDoc.customer_id == Customer.id)
            .join(SalesItem, SalesItem.doc_id == SalesDoc.id)
            .filter(
                or_(
                    SalesItem.name.ilike(like1),
                    SalesItem.name.ilike(like2),
                )
            )
            .order_by(SalesDoc.date.desc(), SalesDoc.id.desc())
            .first()
        )

        if not row:
            return "-"

        name = row[0] or ""
        return name or "-"
    except Exception:
        # กันไม่ให้ dashboard พัง ถ้ามี error ใด ๆ
        return "-"


@app.template_global()
def renting_customer_for_sku(sku: str) -> str:
    """
    คืนชื่อ 'ลูกค้าที่เช่าอยู่' ของอุปกรณ์ตาม SKU
    ใช้จากใบเสนอราคา (QU) ที่อนุมัติล่าสุด
    """
    from sqlalchemy import func

    if not sku:
        return "-"

    sku = sku.strip()
    like = f"%[{sku}]%"

    row = (
        db.session.query(Customer.name)
        .join(SalesDoc, SalesDoc.customer_id == Customer.id)
        .join(SalesItem, SalesItem.doc_id == SalesDoc.id)
        .filter(
            SalesDoc.doc_type == "QU",
            SalesDoc.status == "APPROVED",
            SalesItem.name.ilike(like),
        )
        .order_by(SalesDoc.date.desc(), SalesDoc.id.desc())
        .first()
    )

    if not row:
        return "-"

    name = row[0] or ""
    return name or "-"


# ---- API: Equipment search (SKU + name) ----
@app.get("/api/equipment/search")
@permission_required("equipment.view")
def api_equipment_search():
    q = (request.args.get("q") or "").strip()
    include_rented = request.args.get("include_rented", type=int) == 1
    limit = request.args.get("limit", type=int) or 20
    qry = Equipment.query
    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(Equipment.sku.ilike(like), Equipment.name.ilike(like)))
    if not include_rented:
        qry = qry.filter(Equipment.status != "RENTED")
    rows = qry.order_by(Equipment.name.asc()).limit(limit).all()
    def _f(x):
        try: return float(x or 0)
        except Exception: return 0.0
    out = []
    for e in rows:
        out.append({
            "id": e.id,
            "sku": e.sku,
            "name": e.name,
            "image": e.image_path or "",
            "price_per_day":   _f(getattr(e, "price_per_day_break_even", 0)),
            "price_per_month": _f(getattr(e, "price_per_month_break_even", 0)),
            "price_per_year":  _f(getattr(e, "price_per_year_break_even", 0)),
            "label": f"{e.sku} · {e.name}",
            "value": f"[{e.sku}] {e.name}",
        })
    return jsonify(out)

# ---- API: Catalog search (Categories + Equipment) for autocomplete in PO/QU ----
@app.get("/api/catalog/search")
@permission_required("equipment.view")
def api_catalog_search():
    q = (request.args.get("q") or "").strip()
    limit = request.args.get("limit", type=int) or 20
    out = []

    if not q:
        return jsonify(out)

    like = f"%{q}%"

    # Categories: search by prefix_sku or name
    cats = (
        Category.query
        .filter(or_(Category.prefix_sku.ilike(like), Category.name.ilike(like)))
        .order_by(Category.name.asc())
        .limit(limit)
        .all()
    )
    for c in cats:
        p = (c.prefix_sku or "").strip()
        nm = (c.name or "").strip() or p
        if not p and not nm:
            continue
        out.append({
            "type": "CATEGORY",
            "sku": p,  # for QU/PO input we use prefix_sku as code
            "name": nm,
            "label": f"[{p}] {nm}" if p else nm,
        })

    # Equipment: search by full sku or name (include rented)
    eqs = (
        Equipment.query
        .filter(or_(Equipment.sku.ilike(like), Equipment.name.ilike(like)))
        .order_by(Equipment.name.asc())
        .limit(limit)
        .all()
    )
    for e in eqs:
        sku = (e.sku or "").strip()
        nm = (e.name or "").strip() or sku
        if not sku and not nm:
            continue
        out.append({
            "type": "EQUIPMENT",
            "sku": sku,  # full sku
            "name": nm,
            "brand": getattr(e, "brand", None),
            "warehouse": getattr(e, "warehouse", None),
            "label": f"[{sku}] {nm}",
        })

    # Dedup by (type, sku)
    seen = set()
    uniq = []
    for it in out:
        key = (it.get("type"), it.get("sku") or it.get("name"))
        if key in seen:
            continue
        seen.add(key)
        uniq.append(it)

    return jsonify(uniq[:limit])

@app.get("/api/catalog/lookup")
@permission_required("equipment.view")
def api_catalog_lookup():
    """Lookup one item for PO/QU autofill.
    Priority:
      1) exact Equipment.sku
      2) exact Category.prefix_sku
      3) best match by name (equipment first, then category)
    """
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify({"ok": False})

    # 1) exact equipment sku
    e = Equipment.query.filter(func.upper(Equipment.sku) == q.upper()).first()
    if e:
        return jsonify({
            "ok": True,
            "type": "EQUIPMENT",
            "sku": (e.sku or "").strip(),
            "name": (e.name or "").strip(),
            "brand": getattr(e, "brand", None),
            "warehouse": getattr(e, "warehouse", None),
            "label": f"[{(e.sku or '').strip()}] {(e.name or '').strip()}",
        })

    # 2) exact category prefix
    c = Category.query.filter(func.upper(Category.prefix_sku) == q.upper()).first()
    if c:
        p = (c.prefix_sku or "").strip()
        nm = (c.name or "").strip() or p
        return jsonify({
            "ok": True,
            "type": "CATEGORY",
            "sku": p,
            "name": nm,
            "label": f"[{p}] {nm}" if p else nm,
        })

    like = f"%{q}%"
    e = Equipment.query.filter(Equipment.name.ilike(like)).order_by(Equipment.name.asc()).first()
    if e:
        return jsonify({
            "ok": True,
            "type": "EQUIPMENT",
            "sku": (e.sku or "").strip(),
            "name": (e.name or "").strip(),
            "brand": getattr(e, "brand", None),
            "warehouse": getattr(e, "warehouse", None),
            "label": f"[{(e.sku or '').strip()}] {(e.name or '').strip()}",
        })
    c = Category.query.filter(Category.name.ilike(like)).order_by(Category.name.asc()).first()
    if c:
        p = (c.prefix_sku or "").strip()
        nm = (c.name or "").strip() or p
        return jsonify({
            "ok": True,
            "type": "CATEGORY",
            "sku": p,
            "name": nm,
            "label": f"[{p}] {nm}" if p else nm,
        })

    return jsonify({"ok": False})


def _build_item_img_map(d: SalesDoc) -> dict[int, str]:
    """คืน mapping item_id -> static url ของรูปสินค้า (รองรับหลายรูปแบบ path)

    ลำดับความพยายาม:
      1) it.image_path (ถ้าเป็น path ใน static)
      2) it.allocated_skus (SKU จริงที่เลือกจาก BK) -> หาไฟล์ใน static/uploads/equipment/
         - ชื่อไฟล์รองรับ: <sku>.<ext>, equip_<sku>.<ext>, equipment_<sku>.<ext>, img_<sku>.<ext>
         - ถ้าไม่เจอ จะ fallback แบบ "รุ่นเดียวกัน" โดยตัดเลขท้าย (เช่น ...-002) แล้วหาไฟล์ที่ขึ้นต้นด้วย prefix นั้น
      3) SKU ในชื่อแบบ [SKU] ชื่อ (ของเดิม)
    """
    import os, re
    from flask import current_app, url_for

    def _url(relpath: str) -> str:
        return url_for("static", filename=relpath)

    def _abs_static(relpath: str) -> str:
        return os.path.join(current_app.root_path, "static", relpath.replace("/", os.sep))

    def _normalize_static_rel(p: str) -> str:
        p = (p or "").strip()
        if not p:
            return ""
        # ถ้าเก็บมาเป็น URL เต็ม / data: ให้ข้าม (template handle เอง)
        if "://" in p or p.startswith("data:"):
            return ""
        p = p.lstrip("/")
        if p.startswith("static/"):
            p = p[7:]
        return p

    def _exists(relpath: str) -> bool:
        return bool(relpath) and os.path.exists(_abs_static(relpath))

    def _extract_sku_from_name(title: str) -> str | None:
        m = re.search(r"\[([^\[\]]+?)\]", title or "")
        return m.group(1).strip() if m else None

    # เตรียม index ชื่อไฟล์ใน uploads/equipment สำหรับ fallback แบบ prefix
    equip_dir_rel = "uploads/equipment"
    equip_dir_abs = _abs_static(equip_dir_rel)
    equip_files = []
    try:
        if os.path.isdir(equip_dir_abs):
            equip_files = [f for f in os.listdir(equip_dir_abs) if os.path.isfile(os.path.join(equip_dir_abs, f))]
    except Exception:
        equip_files = []

    exts = (".jpg", ".jpeg", ".png", ".webp")
    prefixes = ("", "equip_", "equipment_", "img_")

    def _find_by_sku(sku: str) -> str | None:
        sku = (sku or "").strip()
        if not sku:
            return None

        # 1) ตรงตัว
        for px in prefixes:
            for ext in exts:
                cand = f"{equip_dir_rel}/{px}{sku}{ext}"
                if _exists(cand):
                    return _url(cand)

        # 2) fallback: ตัดเลขท้ายหลัง '-' แล้วหาไฟล์ที่ขึ้นต้นด้วย prefix นั้น
        #    เช่น SP-001-070126-002 -> SP-001-070126-
        base_prefix = sku
        if "-" in sku:
            base_prefix = sku.rsplit("-", 1)[0] + "-"

        # หาไฟล์ในโฟลเดอร์ที่เริ่มด้วย base_prefix
        # ตัวอย่างชื่อไฟล์: equip_SP-001-070126-001.jpg
        for px in prefixes:
            start = f"{px}{base_prefix}"
            found = None
            for fn in equip_files:
                low = fn.lower()
                if not low.endswith(exts):
                    continue
                if fn.startswith(start):
                    found = fn
                    break
            if found:
                cand = f"{equip_dir_rel}/{found}"
                if _exists(cand):
                    return _url(cand)

        return None

    img_map: dict[int, str] = {}

    for it in getattr(d, "items", []) or []:
        # (A) image_path ตรง ๆ
        rel = _normalize_static_rel(getattr(it, "image_path", "") or "")
        if rel and _exists(rel):
            img_map[it.id] = _url(rel)
            continue

        # (B) allocated_skus -> ใช้ SKU ตัวแรก
        alloc = (getattr(it, "allocated_skus", "") or "").strip()
        if alloc:
            sku_first = None
            for s in alloc.split(","):
                s = (s or "").strip()
                if s:
                    sku_first = s
                    break
            if sku_first:
                u = _find_by_sku(sku_first)
                if u:
                    img_map[it.id] = u
                    continue

        # (C) [SKU] ชื่อ (ของเดิม)
        sku = _extract_sku_from_name(getattr(it, "name", "") or "")
        if sku:
            u = _find_by_sku(sku)
            if u:
                img_map[it.id] = u
                continue

    return img_map

@app.get("/api/categories/search")
@permission_required("sales.manage")
def api_categories_search():
    """ค้นหาหมวดหมู่อุปกรณ์ (Category) ด้วย prefix_sku หรือชื่อหมวด
    ใช้สำหรับหน้าสร้างใบเสนอราคา (QU) เพื่อให้เลือก "หมวดหมู่" เท่านั้น ไม่ล็อคตัวอุปกรณ์จริง
    """
    q = (request.args.get("q") or "").strip()
    out = []
    qs = Category.query
    if q:
        like = f"%{q}%"
        qs = qs.filter(
            db.or_(
                Category.prefix_sku.ilike(like),
                Category.name.ilike(like),
            )
        )
    cats = qs.order_by(Category.prefix_sku.asc()).limit(50).all()
    for c in cats:
        out.append({
            "sku": (c.prefix_sku or "").strip(),  # ให้ฟิลด์ชื่อ sku เพื่อ reuse UI เดิม
            "prefix_sku": (c.prefix_sku or "").strip(),
            "name": (c.name or "").strip(),
            "category_name": (c.name or "").strip(),
        })
    return jsonify(out)


@app.route("/sales/quotes/<int:qid>/preview")
@permission_required("sales.view")
def qu_preview(qid: int):
    d = SalesDoc.query.options(
        joinedload(SalesDoc.items),
        joinedload(SalesDoc.customer),
    ).get_or_404(qid)
    img_map = _build_item_img_map(d)
    return render_template("sales/qu_print.html", d=d, today=date.today(), mode="preview", img_map=img_map)

@app.route("/sales/quotes/<int:qid>/print")
@permission_required("sales.view")
def qu_print(qid):
    d = SalesDoc.query.options(
        joinedload(SalesDoc.items),
        joinedload(SalesDoc.customer),
    ).get_or_404(qid)
    auto = request.args.get("auto")
    img_map = _build_item_img_map(d)
    return render_template(
        "sales/qu_print.html",
        d=d,
        today=date.today(),
        auto=bool(auto),
        mode="print",
        img_map=img_map,
    )

def _extract_item_skus(items):
    import re
    out = []
    for it in (items or []):
        name = (it.name or "").strip()
        m = re.search(r"\[([^\[\]]+?)\]", name)
        if m:
            out.append((it.id, m.group(1).strip()))
    return out

def _update_equipment_from_quote(d, target_status: str):
    id_sku = _extract_item_skus(d.items)
    if not id_sku:
        return 0, []
    skus = [sku for _, sku in id_sku]
    eqs = Equipment.query.filter(Equipment.sku.in_(skus)).all()
    sku2eq = {e.sku: e for e in eqs}
    changed = 0
    missing = []
    for _, sku in id_sku:
        e = sku2eq.get(sku)
        if not e:
            missing.append(sku); continue
        if e.status != target_status:
            prev = e.status
            e.status = target_status
            db.session.add(EquipmentLog(
                equipment_id=e.id,
                action="STATUS",
                note=f"จาก {EQUIP_STATUS_THAI.get(prev, prev)} → {EQUIP_STATUS_THAI.get(target_status, target_status)} จากใบเสนอราคา {d.number}",
                user_id=(current_user.id if current_user.is_authenticated else None),
            ))
            changed += 1
    return changed, missing

def _clone_items(from_doc: SalesDoc, to_doc: SalesDoc):
    """คัดลอกรายการจากเอกสารหนึ่งไปอีกเอกสารหนึ่ง
    - รองรับ field เพิ่มเติมแบบปลอดภัย (จะใส่เฉพาะคอลัมน์ที่มีจริงใน SalesItem)
    - ต้องคัดลอก field ที่ใช้ล็อคของ/แสดง SKU ด้วย (category_id, category_prefix, allocated_skus)
    """
    # รายชื่อคอลัมน์จริงของ SalesItem (กัน TypeError: invalid keyword argument)
    item_cols = set(getattr(SalesItem, "__table__").columns.keys())

    def _val(obj, key, default=None):
        return getattr(obj, key, default)

    for it in (from_doc.items or []):
        data = {}

        # --- core fields ---
        for k, default in [
            ("image_path", ""),
            ("name", ""),
            ("qty", 1),
            ("rent_unit", None),
            ("rent_duration", 1),
            ("unit_price", 0),
            ("discount_pct", 0),
            ("line_total", 0),
            ("line_subtotal", None),
        ]:
            if k in item_cols:
                v = _val(it, k, default)
                if v is None and default is not None:
                    v = default
                data[k] = v

        # --- lock/SKU fields ---
        for k in ["category_id", "category_prefix", "allocated_skus", "source_qu_item_id"]:
            if k in item_cols:
                data[k] = _val(it, k, None)

        # --- equipment link (ถ้ามีในรุ่นนี้) ---
        for k in ["equipment_id", "equipment_sku", "equipment_code", "equipment_name"]:
            if k in item_cols:
                data[k] = _val(it, k, None)

        db.session.add(SalesItem(doc=to_doc, **data))



def _create_child_doc(parent: SalesDoc, doc_type: str, init_status: str) -> SalesDoc:
    prefix = {"BL": "BL", "IV": "IV", "RC": "RC"}[doc_type]
    child = SalesDoc(
        number=_gen_running(prefix, SalesDoc),
        doc_type=doc_type,
        status=init_status,
        customer_id=parent.customer_id,
        po_customer=parent.po_customer,
        credit_days=parent.credit_days or 0,
        tax_mode=parent.tax_mode,
        wht_pct=parent.wht_pct or 0,
        date=date.today(),
        remark="",
        parent=parent,
        amount_subtotal=parent.amount_subtotal or 0.0,
        amount_vat=parent.amount_vat or 0.0,
        amount_total=parent.amount_total or 0.0,
        amount_wht=parent.amount_wht or 0.0,
        amount_grand=parent.amount_grand or 0.0,
    )
    db.session.add(child)
    db.session.flush()
    _clone_items(parent, child)
    return child

def _ensure_children_for_quote(qu: SalesDoc):
    children = {c.doc_type: c for c in SalesDoc.query.filter_by(parent_id=qu.id).all()}
    if "BL" not in children:
        _create_child_doc(qu, "BL", "UNPAID")
    if "IV" not in children:
        _create_child_doc(qu, "IV", "UNISSUED")
    if "RC" not in children:
        _create_child_doc(qu, "RC", "UNISSUED")

@app.post("/sales/quotes/<int:qid>/approve")
@login_required
@permission_required("sales.manage")
def qu_approve(qid):
    # โหลดใบเสนอราคาที่จะอนุมัติ
    d = (
        SalesDoc.query.options(
            joinedload(SalesDoc.items),
            joinedload(SalesDoc.customer),
        )
        .get_or_404(qid)
    )

    # กันกดอนุมัติซ้ำ
    if (d.status or "").upper() == "APPROVED":
        flash("เอกสารนี้อนุมัติแล้ว", "info")
        # ถ้าเป็นสัญญา ให้พาไปหน้า CT (ถ้ามี)
        ct = SalesDoc.query.filter_by(parent_id=d.id, doc_type="CT").first()
        if ct:
            return redirect(url_for("contract_view", cid=ct.id))
        return redirect(url_for("qu_view", qid=d.id))

    # เปลี่ยนสถานะใบเสนอราคา
    d.status = "APPROVED"

    # โหมดสัญญา/แบ่งงวดรายเดือน
    # โหมดสัญญา/แบ่งงวดรายเดือน
    if (d.billing_mode or '').upper() == 'INSTALLMENT':
        # 1) ต้องมีใบจองเพื่อใช้ล็อคของ (BK) เหมือนโหมดปกติ
        bk = SalesDoc.query.filter_by(parent_id=d.id, doc_type='BK').first()
        if not bk:
            bk = _create_booking_from_quote(d)

        # 2) ต้องมีสัญญา/PO ใหญ่ (CT) เพื่อสร้างตารางงวด
        ct = SalesDoc.query.filter_by(parent_id=d.id, doc_type='CT').first()
        if not ct:
            ct = _create_contract_from_quote(d)

        db.session.commit()
        flash('อนุมัติใบเสนอราคาแล้ว: สร้างใบจองเพื่อไปล็อคของ และสร้างสัญญา/งวดรายเดือนแล้ว', 'success')
        # พาไปหน้าใบจองก่อน เพื่อให้กดจัดสรร/อนุมัติ (ล็อคของ -> อุปกรณ์เปลี่ยนเป็นถูกเช่า)
        return redirect(url_for('bk_view', doc_id=bk.id))

    # โหมดปกติ (เดิม): สร้างใบจอง (BK) ถ้ายังไม่มี
    existing_bk = SalesDoc.query.filter_by(parent_id=d.id, doc_type="BK").first()
    if not existing_bk:
        bk = _create_booking_from_quote(d)
    else:
        bk = existing_bk

    db.session.commit()
    flash("อนุมัติใบเสนอราคา และสร้างใบจองเรียบร้อยแล้ว", "success")

    # ไปหน้าใบจอง (param ชื่อ doc_id)
    return redirect(url_for("bk_view", doc_id=bk.id))


@app.post("/sales/quotes/<int:qid>/cancel")
@login_required
@permission_required("sales.manage")
def qu_cancel(qid):
    """ยกเลิกใบเสนอราคา (QU) และคืนสต็อกที่จองไว้
    - เดิม: ไม่ให้ยกเลิกเมื่อ QU = APPROVED
    - ใหม่: อนุญาตให้ยกเลิกแม้ APPROVED ได้ "ถ้า" downstream ยังไม่ล็อคจริง
      เงื่อนไขสำคัญ:
        - BK (ถ้ามี) ต้องยังไม่ APPROVED
        - BK ต้องยังไม่มี allocated_skus
    """
    qu: SalesDoc = (
        SalesDoc.query.options(joinedload(SalesDoc.items)).get_or_404(qid)
    )
    if (qu.doc_type or "").upper() != "QU":
        abort(404)

    st = (qu.status or "").upper()
    if st == "CANCELLED":
        flash("ใบเสนอราคานี้ถูกยกเลิกแล้ว", "info")
        return redirect(url_for("qu_view", qid=qu.id))

    # หา BK ลูก (ถ้ามี)
    bk = (
        SalesDoc.query.options(joinedload(SalesDoc.items))
        .filter_by(parent_id=qu.id, doc_type="BK")
        .first()
    )

    # ถ้า QU อนุมัติแล้ว -> ให้ยกเลิกได้เฉพาะกรณี BK ยังไม่ล็อคของจริง
    if st == "APPROVED":
        if not bk:
            # อนุมัติแล้วแต่ไม่มี BK (ผิด flow) — กันไว้
            flash("ใบเสนอราคาอนุมัติแล้ว แต่ไม่พบใบจอง (BK) — ไม่อนุญาตให้ยกเลิก", "warning")
            return redirect(url_for("qu_view", qid=qu.id))

        bk_st = (bk.status or "").upper()
        if bk_st == "APPROVED":
            flash("มีใบจอง (BK) ที่อนุมัติแล้ว — ไม่อนุญาตให้ยกเลิกใบเสนอราคา", "warning")
            return redirect(url_for("qu_view", qid=qu.id))

        has_alloc = any((it.allocated_skus or "").strip() for it in (bk.items or []))
        if has_alloc:
            flash("มีการจัดสรร SKU แล้วในใบจอง (BK) — ไม่อนุญาตให้ยกเลิกใบเสนอราคา", "warning")
            return redirect(url_for("qu_view", qid=qu.id))

    # ===== ผ่านเงื่อนไข -> ยกเลิก QU + คืนจอง =====
    qu.status = "CANCELLED"
    _release_reservations_for_doc("QU", qu.id)

    # ถ้ามี BK และ BK ยังไม่ APPROVED -> ยกเลิก BK และคืนจองด้วย
    if bk and (bk.status or "").upper() != "APPROVED":
        has_alloc = any((it.allocated_skus or "").strip() for it in (bk.items or []))
        if has_alloc:
            flash("มีการจัดสรร SKU แล้วในใบจอง (BK) — ไม่ยกเลิก BK อัตโนมัติ", "warning")
        else:
            bk.status = "CANCELLED"
            _release_reservations_for_doc("BK", bk.id)

    db.session.commit()
    flash("ยกเลิกใบเสนอราคาและคืนสต็อกที่จองไว้แล้ว", "success")
    return redirect(url_for("qu_view", qid=qu.id))



@app.post("/sales/bookings/<int:bid>/cancel")
@login_required
@permission_required("sales.manage")
def bk_cancel(bid):
    """ยกเลิกใบจอง (BK) และคืนสต็อกที่จองไว้ (RESERVED)
    - อนุญาตเฉพาะที่ยังไม่ APPROVED
    - ถ้ามีการ Allocate SKU แล้ว จะคืนสถานะอุปกรณ์กลับเป็น READY และล้าง allocated_skus
    """
    bk: SalesDoc = (
        SalesDoc.query.options(joinedload(SalesDoc.items), joinedload(SalesDoc.customer)).get_or_404(bid)
    )
    if (bk.doc_type or "").upper() != "BK":
        abort(404)

    st = (bk.status or "").upper()
    if st == "APPROVED":
        flash("ใบจองนี้อนุมัติแล้ว ไม่อนุญาตให้ยกเลิก", "warning")
        return redirect(url_for("bk_view", doc_id=bk.id))
    if st == "CANCELLED":
        flash("ใบจองนี้ถูกยกเลิกแล้ว", "info")
        return redirect(url_for("bk_view", doc_id=bk.id))

    cust_name = bk.customer.name if getattr(bk, "customer", None) else ""
    # คืนอุปกรณ์ที่ถูก Allocate แล้ว (ถ้ามี)
    for it in (bk.items or []):
        skus = [s.strip() for s in (it.allocated_skus or "").split(",") if s.strip()]
        if not skus:
            continue
        eqs = Equipment.query.filter(Equipment.sku.in_(skus)).all()
        by_sku = {e.sku: e for e in eqs}
        for sku in skus:
            eq = by_sku.get(sku)
            if not eq:
                continue
            # คืนเป็น READY เฉพาะกรณีที่ยัง RENTED อยู่
            if (eq.status or "").upper() == "RENTED":
                eq.status = "READY"
                _equip_log(eq, "CANCEL_BOOKING", f"ยกเลิก {bk.number} | ลูกค้า: {cust_name}")
        it.allocated_skus = ""  # ล้าง
        # sync QU on bk_cancel
        try:
            if bk.parent_id:
                qu = SalesDoc.query.get(bk.parent_id)
                if qu and (qu.doc_type or '').upper() == 'QU' and getattr(it, 'source_qu_item_id', None):
                    qu_it = SalesItem.query.get(it.source_qu_item_id)
                    if qu_it and qu_it.doc_id == qu.id:
                        qu_it.allocated_skus = ""
        except Exception:
            pass
        # เผื่อมี reservation ผูก item นี้ค้างอยู่
        _release_reservation_for_item(it.id)

    # คืนจองทั้งหมดของ BK
    _release_reservations_for_doc("BK", bk.id)
    bk.status = "CANCELLED"
    db.session.commit()

    flash("ยกเลิกใบจองและคืนสต็อกเรียบร้อยแล้ว", "success")
    return redirect(url_for("bk_view", doc_id=bk.id))

# ---------- Sales: Lists/View/Toggle/Print ----------
def _doc_list(doc_type: str, title_th: str):
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)
    qry = SalesDoc.query.filter(SalesDoc.doc_type == doc_type)
    if start_d:
        qry = qry.filter(SalesDoc.date >= start_d)
    if end_d:
        qry = qry.filter(SalesDoc.date <= end_d)
    if q:
        qry = qry.join(Customer).filter(Customer.name.ilike(f"%{q}%"))
    rows = qry.order_by(SalesDoc.id.desc()).all()
    return render_template(
        "sales/qu_list.html",
        rows=rows,
        q=q,
        start=start,
        end=end,
        doc_type=doc_type,
        page_title=title_th,
        show_new=False,
    )

@app.route("/sales/bills")
@permission_required("sales.view")
def bl_list():
    return _doc_list("BL", "ใบวางบิล")

@app.route("/sales/invoices")
@permission_required("sales.view")
def iv_list():
    return _doc_list("IV", "ใบกำกับภาษี")

@app.route("/sales/receipts")
@permission_required("sales.view")
def rc_list():
    return _doc_list("RC", "ใบเสร็จรับเงิน")



# ================== Sales: Contracts / Installments ==================

@app.route("/sales/contracts")
@permission_required("sales.view")
def contract_list():
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    qry = SalesDoc.query.options(joinedload(SalesDoc.customer)).filter(SalesDoc.doc_type == "CT")

    if start_d:
        qry = qry.filter(SalesDoc.date >= start_d)
    if end_d:
        qry = qry.filter(SalesDoc.date <= end_d)

    if q:
        qry = qry.join(Customer).filter(Customer.name.ilike(f"%{q}%"))

    rows = qry.order_by(SalesDoc.id.desc()).all()

    # summary per contract
    stats = {}
    for ct in rows:
        total = SalesInstallment.query.filter_by(contract_id=ct.id).count()
        receipted = SalesInstallment.query.filter_by(contract_id=ct.id, status="RECEIPTED").count()
        invoiced = SalesInstallment.query.filter(
            SalesInstallment.contract_id == ct.id,
            SalesInstallment.invoice_id.isnot(None)
        ).count()
        stats[ct.id] = {"total": total, "receipted": receipted, "invoiced": invoiced}

    return render_template(
        "sales/contracts_list.html",
        rows=rows,
        q=q,
        start=start,
        end=end,
        stats=stats,
    )

@app.route("/sales/contracts/<int:cid>")
@permission_required("sales.view")
def contract_view(cid):
    ct = SalesDoc.query.options(joinedload(SalesDoc.customer)).get_or_404(cid)
    if ct.doc_type != "CT":
        abort(404)

    # ensure installments exist (กรณีย้าย DB/เพิ่มฟีเจอร์ทีหลัง)
    if SalesInstallment.query.filter_by(contract_id=ct.id).count() == 0 and ct.parent_id:
        qu = SalesDoc.query.options(joinedload(SalesDoc.items)).get(ct.parent_id)
        _ensure_installments_for_contract(ct, qu)
        db.session.commit()

    insts = SalesInstallment.query.filter_by(contract_id=ct.id).order_by(SalesInstallment.installment_no.asc()).all()

    # summary
    total = len(insts)
    receipted = len([x for x in insts if (x.status or "").upper() == "RECEIPTED"])
    invoiced = len([x for x in insts if x.invoice_id])

    return render_template(
        "sales/contract_view.html",
        ct=ct,
        insts=insts,
        summary={"total": total, "receipted": receipted, "invoiced": invoiced},
    )

@app.post("/sales/contracts/<int:cid>/installments/<int:iid>/set_po")
@permission_required("sales.manage")
def contract_installment_set_po(cid, iid):
    ct = SalesDoc.query.get_or_404(cid)
    if ct.doc_type != "CT":
        abort(404)
    inst = SalesInstallment.query.filter_by(contract_id=cid, id=iid).first_or_404()
    inst.po_customer_sub = (request.form.get("po_customer_sub") or "").strip()
    db.session.commit()
    flash("บันทึก PO ย่อยของลูกค้าแล้ว", "success")
    return redirect(url_for("contract_view", cid=cid))

@app.post("/sales/contracts/<int:cid>/installments/<int:iid>/create_docs")
@permission_required("sales.manage")
def contract_installment_create_docs(cid, iid):
    ct = SalesDoc.query.get_or_404(cid)
    if ct.doc_type != "CT":
        abort(404)
    inst = SalesInstallment.query.filter_by(contract_id=cid, id=iid).first_or_404()

    # กันสร้างซ้ำ
    if inst.bill_id or inst.invoice_id or inst.receipt_id:
        flash("งวดนี้มีเอกสารถูกสร้างแล้ว", "info")
        return redirect(url_for("contract_view", cid=cid))

    _create_docs_for_installment(ct, inst)
    db.session.commit()
    flash("สร้างเอกสารงวดนี้เรียบร้อยแล้ว (BL/IV/RC)", "success")
    return redirect(url_for("contract_view", cid=cid))

@app.post("/sales/contracts/<int:cid>/installments/<int:iid>/mark_paid")
@permission_required("sales.manage")
def contract_installment_mark_paid(cid, iid):
    ct = SalesDoc.query.get_or_404(cid)
    if ct.doc_type != "CT":
        abort(404)
    inst = SalesInstallment.query.filter_by(contract_id=cid, id=iid).first_or_404()
    if not inst.receipt_id:
        flash("ยังไม่มีใบเสร็จของงวดนี้", "warning")
        return redirect(url_for("contract_view", cid=cid))
    rc = SalesDoc.query.get_or_404(inst.receipt_id)
    rc.status = "PAID"
    inst.status = "RECEIPTED"
    db.session.commit()
    flash("บันทึกว่าชำระแล้ว (อัปเดตสถานะงวดเป็น RECEIPTED)", "success")
    return redirect(url_for("contract_view", cid=cid))

@app.route("/sales/contracts/<int:cid>/export.xlsx")
@permission_required("sales.view")
def contract_export_xlsx(cid):
    ct = SalesDoc.query.options(joinedload(SalesDoc.customer)).get_or_404(cid)
    if ct.doc_type != "CT":
        abort(404)
    insts = SalesInstallment.query.filter_by(contract_id=ct.id).order_by(SalesInstallment.installment_no.asc()).all()

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    headers = [
        "Contract No", "Customer", "Project", "Start", "End",
        "Installment No", "Period Start", "Period End", "Bill Date", "Due Date",
        "Status", "PO Sub",
        "Subtotal", "VAT", "Total", "WHT", "Grand",
        "BL No", "IV No", "RC No"
    ]
    ws.append(headers)

    for inst in insts:
        bl_no = inst.bill.number if inst.bill else ""
        iv_no = inst.invoice.number if inst.invoice else ""
        rc_no = inst.receipt.number if inst.receipt else ""
        ws.append([
            ct.number,
            (ct.customer.name if ct.customer else ""),
            ct.project_name,
            (ct.contract_start.strftime("%Y-%m-%d") if ct.contract_start else ""),
            (ct.contract_end.strftime("%Y-%m-%d") if ct.contract_end else ""),
            inst.installment_no,
            inst.period_start.strftime("%Y-%m-%d"),
            inst.period_end.strftime("%Y-%m-%d"),
            inst.bill_date.strftime("%Y-%m-%d"),
            inst.due_date.strftime("%Y-%m-%d"),
            inst.status,
            inst.po_customer_sub,
            inst.amount_subtotal,
            inst.amount_vat,
            inst.amount_total,
            inst.amount_wht,
            inst.amount_grand,
            bl_no, iv_no, rc_no
        ])

    # autosize
    for col in range(1, len(headers) + 1):
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"{ct.number}_schedule.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _doc_view(doc_id: int, doc_type: str, title_th: str):
    d = SalesDoc.query.options(
        joinedload(SalesDoc.items),
        joinedload(SalesDoc.customer),
    ).get_or_404(doc_id)
    if d.doc_type != doc_type:
        abort(404)
    return render_template("sales/qu_view.html", d=d, page_title=title_th, hide_approve=True, is_child_doc=True)

# -------------------------
# Excel Exports (Reports)
# -------------------------
def _xlsx_send(wb, filename: str):
    """ส่งไฟล์ Excel (openpyxl Workbook) ออกเป็น attachment"""
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@app.get("/sales/docs/<int:doc_id>/export.xlsx")
@permission_required("sales.view")
def sales_doc_export_xlsx(doc_id):
    """Export เอกสารขาย 1 ใบ (QU/BK/BL/IV/RC/...)"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    d: SalesDoc = (
        SalesDoc.query
        .options(joinedload(SalesDoc.items), joinedload(SalesDoc.customer))
        .get_or_404(doc_id)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Document"

    # Header
    ws.append(["Doc Type", d.doc_type])
    ws.append(["Number", d.number])
    ws.append(["Status", d.status])
    ws.append(["Date", d.date.strftime("%Y-%m-%d") if d.date else ""])
    ws.append(["Customer", d.customer.name if d.customer else ""])
    ws.append(["Project", getattr(d, "project_name", "") or ""])
    ws.append(["PO(Customer)", getattr(d, "po_customer", "") or ""])
    ws.append(["Tax Mode", getattr(d, "tax_mode", "") or ""])
    ws.append(["WHT %", getattr(d, "wht_pct", "") or ""])
    ws.append(["Billing Mode", getattr(d, "billing_mode", "") or ""])
    ws.append(["Contract Start", getattr(d, "contract_start", None).strftime("%Y-%m-%d") if getattr(d, "contract_start", None) else ""])
    ws.append(["Contract End", getattr(d, "contract_end", None).strftime("%Y-%m-%d") if getattr(d, "contract_end", None) else ""])
    ws.append(["Installments", getattr(d, "installment_count", "") or ""])
    ws.append(["Remark", getattr(d, "remark", "") or ""])
    ws.append([])

    # Items
    ws.append(["#", "Item", "Category Prefix", "Allocated SKUs", "Qty", "Unit", "Duration", "Unit Price", "Disc %", "Line Subtotal", "Line Total"])
    for i, it in enumerate(d.items or [], start=1):
        ws.append([
            i,
            it.name,
            getattr(it, "category_prefix", "") or "",
            getattr(it, "allocated_skus", "") or "",
            float(it.qty or 0),
            (it.rent_unit or ""),
            float(getattr(it, "rent_duration", 0) or 0),
            float(it.unit_price or 0),
            float(it.discount_pct or 0),
            float(it.line_subtotal or 0),
            float(it.line_total or 0),
        ])

    ws.append([])
    ws.append(["Subtotal", float(getattr(d, "amount_subtotal", 0) or 0)])
    ws.append(["VAT", float(getattr(d, "amount_vat", 0) or 0)])
    ws.append(["Total", float(getattr(d, "amount_total", 0) or 0)])
    ws.append(["WHT", float(getattr(d, "amount_wht", 0) or 0)])
    ws.append(["Grand", float(getattr(d, "amount_grand", 0) or 0)])

    # column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18

    return _xlsx_send(wb, f"{(d.doc_type or 'DOC')}_{d.number or d.id}.xlsx")

@app.get("/purchases/po/<int:pid>/export.xlsx")
@permission_required("purchases.view")
def po_export_xlsx(pid):
    """Export ใบสั่งซื้อ (PO) 1 ใบ"""
    from openpyxl import Workbook
    po: PurchaseOrder = (
        PurchaseOrder.query
        .options(joinedload(PurchaseOrder.items), joinedload(PurchaseOrder.supplier))
        .get_or_404(pid)
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "PO"

    ws.append(["Number", po.number])
    ws.append(["Status", po.status])
    ws.append(["Date", po.po_date.strftime("%Y-%m-%d") if po.po_date else ""])
    ws.append(["Supplier", po.supplier.name if po.supplier else ""])
    ws.append(["Remark", getattr(po, "remark", "") or ""])
    ws.append([])

    ws.append(["#", "SKU", "Item", "Brand", "Qty", "Unit", "Cost", "Disc %", "Line Total"])
    for i, it in enumerate(po.items or [], start=1):
        ws.append([
            i,
            getattr(it, "sku", "") or "",
            getattr(it, "name", "") or "",
            getattr(it, "brand", "") or "",
            float(getattr(it, "qty", 0) or 0),
            getattr(it, "unit", "") or "",
            float(getattr(it, "cost", 0) or 0),
            float(getattr(it, "disc_pct", 0) or 0),
            float(getattr(it, "line_total", 0) or 0),
        ])

    ws.append([])
    ws.append(["Subtotal", float(getattr(po, "amount_subtotal", 0) or 0)])
    ws.append(["VAT", float(getattr(po, "amount_vat", 0) or 0)])
    ws.append(["Grand", float(getattr(po, "amount_grand", 0) or 0)])

    return _xlsx_send(wb, f"PO_{po.number or po.id}.xlsx")

@app.get("/purchases/grn/<int:gid>/export.xlsx")
@permission_required("purchases.view")
def grn_export_xlsx(gid):
    """Export ใบรับสินค้า (RC) 1 ใบ"""
    from openpyxl import Workbook
    g: GoodsReceipt = (
        GoodsReceipt.query
        .options(joinedload(GoodsReceipt.items), joinedload(GoodsReceipt.po), joinedload(GoodsReceipt.po).joinedload(PurchaseOrder.supplier))
        .get_or_404(gid)
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "GRN"

    ws.append(["Number", g.number])
    ws.append(["Status", g.status])
    ws.append(["Date", g.grn_date.strftime("%Y-%m-%d") if g.grn_date else ""])
    ws.append(["Supplier", g.supplier.name if g.supplier else ""])
    ws.append(["PO", g.po.number if getattr(g, "po", None) else ""])
    ws.append([])

    ws.append(["#", "SKU", "Item", "Brand", "Qty", "Unit", "Cost", "Disc %", "Line Total"])
    for i, it in enumerate(g.items or [], start=1):
        ws.append([
            i,
            getattr(it, "sku", "") or "",
            getattr(it, "name", "") or "",
            getattr(it, "brand", "") or "",
            float(getattr(it, "qty", 0) or 0),
            getattr(it, "unit", "") or "",
            float(getattr(it, "cost", 0) or 0),
            float(getattr(it, "disc_pct", 0) or 0),
            float(getattr(it, "line_total", 0) or 0),
        ])

    ws.append([])
    ws.append(["Subtotal", float(getattr(g, "amount_subtotal", 0) or 0)])
    ws.append(["VAT", float(getattr(g, "amount_vat", 0) or 0)])
    ws.append(["Grand", float(getattr(g, "amount_grand", 0) or 0)])

    return _xlsx_send(wb, f"GRN_{g.number or g.id}.xlsx")

@app.get("/deliveries/<int:did>/export.xlsx")
@permission_required("transport.view")
def delivery_export_xlsx(did):
    """Export ใบส่งสินค้า (Delivery) 1 ใบ"""
    from openpyxl import Workbook
    d: DeliveryDoc = (
        DeliveryDoc.query
        .options(joinedload(DeliveryDoc.items), joinedload(DeliveryDoc.customer))
        .get_or_404(did)
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery"

    ws.append(["Number", d.number])
    ws.append(["Status", d.status])
    ws.append(["Date", d.date.strftime("%Y-%m-%d") if d.date else ""])
    ws.append(["Customer", d.customer.name if d.customer else ""])
    ws.append(["Type", getattr(d, "delivery_type", "") or ""])
    ws.append([])

    ws.append(["#", "Item", "Qty", "Unit", "Remark"])
    for i, it in enumerate(d.items or [], start=1):
        ws.append([
            i,
            getattr(it, "name", "") or "",
            float(getattr(it, "qty", 0) or 0),
            getattr(it, "unit", "") or "",
            getattr(it, "remark", "") or "",
        ])

    return _xlsx_send(wb, f"DN_{d.number or d.id}.xlsx")

@app.get("/claims/<int:cid>/export.xlsx")
@permission_required("claims.view")
def claim_export_xlsx(cid):
    """Export งานเคลม 1 รายการ"""
    from openpyxl import Workbook
    c: Claim = (
        Claim.query
        .options(joinedload(Claim.items), joinedload(Claim.customer))
        .get_or_404(cid)
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Claim"

    ws.append(["Number", c.number])
    ws.append(["Status", c.status])
    ws.append(["Date", c.date.strftime("%Y-%m-%d") if c.date else ""])
    ws.append(["Customer", c.customer.name if c.customer else ""])
    ws.append(["Remark", getattr(c, "remark", "") or ""])
    ws.append([])

    ws.append(["#", "Item", "Qty", "Unit", "Reason", "Status"])
    for i, it in enumerate(c.items or [], start=1):
        ws.append([
            i,
            getattr(it, "name", "") or "",
            float(getattr(it, "qty", 0) or 0),
            getattr(it, "unit", "") or "",
            getattr(it, "reason", "") or "",
            getattr(it, "status", "") or "",
        ])

    return _xlsx_send(wb, f"CLAIM_{c.number or c.id}.xlsx")

@app.get("/repairs/<int:jid>/export.xlsx")
@permission_required("repairs.view")
def repair_export_xlsx(jid):
    """Export ใบงานซ่อม 1 ใบ"""
    from openpyxl import Workbook
    job: RepairJob = (
        RepairJob.query
        .options(joinedload(RepairJob.items), joinedload(RepairJob.equipment), joinedload(RepairJob.customer))
        .get_or_404(jid)
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Repair"

    ws.append(["Number", job.number])
    ws.append(["Status", job.status])
    ws.append(["Opened", job.opened_at.strftime("%Y-%m-%d %H:%M") if job.opened_at else ""])
    ws.append(["Closed", job.closed_at.strftime("%Y-%m-%d %H:%M") if job.closed_at else ""])
    ws.append(["Equipment", job.equipment.sku if job.equipment else ""])
    ws.append(["Customer", job.customer.name if job.customer else ""])
    ws.append(["Total Cost", float(getattr(job, "total_cost", 0) or 0)])
    ws.append(["Note", getattr(job, "note", "") or ""])
    ws.append([])

    ws.append(["#", "Part/Item", "Qty", "Unit", "Cost", "Line Total"])
    for i, it in enumerate(job.items or [], start=1):
        ws.append([
            i,
            getattr(it, "name", "") or "",
            float(getattr(it, "qty", 0) or 0),
            getattr(it, "unit", "") or "",
            float(getattr(it, "cost", 0) or 0),
            float(getattr(it, "line_total", 0) or 0),
        ])

    return _xlsx_send(wb, f"REPAIR_{job.number or job.id}.xlsx")


@app.route("/sales/bills/<int:did>")
@permission_required("sales.view")
def bl_view(did): return _doc_view(did, "BL", "ใบวางบิล")

@app.route("/sales/invoices/<int:did>")
@permission_required("sales.view")
def iv_view(did): return _doc_view(did, "IV", "ใบกำกับภาษี")

@app.route("/sales/receipts/<int:did>")
@permission_required("sales.view")
def rc_view(did): return _doc_view(did, "RC", "ใบเสร็จรับเงิน")

# ---------- Flow Helpers (QU → BK → BL → IV → RC) ----------

def _booking_flow_children_statuses(bk: 'SalesDoc') -> dict:
    """Return map {doc_type: STATUS_UPPER} for BK children (BL/IV/RC)."""
    if not bk or (bk.doc_type or '').upper() != 'BK':
        return {}
    return {c.doc_type: (c.status or '').upper()
            for c in SalesDoc.query.filter_by(parent_id=bk.id).all()}


def booking_flow_ready(bk: 'SalesDoc') -> bool:
    """True if BL=PAID and IV=ISSUED and RC=ISSUED for this BK.

    Note: Manual close only. We do NOT auto-close BK/QU here.
    """
    ch = _booking_flow_children_statuses(bk)
    return ch.get('BL') == 'PAID' and ch.get('IV') == 'ISSUED' and ch.get('RC') == 'ISSUED'


@app.post("/sales/bookings/<int:bid>/close")
@permission_required("sales.manage")
def bk_close(bid):
    """Manual close BK work (and close QU parent if exists).

    Allowed only when BK is APPROVED and flow is ready.
    """
    bk = SalesDoc.query.get_or_404(bid)
    if (bk.doc_type or '').upper() != 'BK':
        abort(404)

    st = (bk.status or '').upper()
    if st in ['CANCELLED']:
        flash('ไม่สามารถปิดงาน: เอกสารถูกยกเลิกแล้ว', 'warning')
        return redirect(url_for('bk_view', doc_id=bk.id))

    if st != 'APPROVED':
        flash('ต้องอนุมัติใบจองก่อนจึงจะปิดงานได้', 'warning')
        return redirect(url_for('bk_view', doc_id=bk.id))

    if not booking_flow_ready(bk):
        flash('ยังปิดงานไม่ได้: ต้องให้ BL=ชำระแล้ว และ IV/RC=ออกแล้ว ก่อน', 'warning')
        return redirect(url_for('bk_view', doc_id=bk.id))

    bk.status = 'CLOSED'

    close_parent = (request.form.get('close_parent') or '').strip() in ['1','true','True','on','yes']
    if close_parent and bk.parent and (bk.parent.doc_type or '').upper() == 'QU':
        bk.parent.status = 'CLOSED'

    db.session.commit()
    flash('ปิดงานเรียบร้อย', 'success')
    return redirect(url_for('bk_view', doc_id=bk.id))


@app.post("/sales/bookings/<int:bid>/reopen")
@permission_required("sales.manage")
def bk_reopen(bid):
    """Reopen BK work (set status back to APPROVED).

    Does not change equipment statuses.
    """
    bk = SalesDoc.query.get_or_404(bid)
    if (bk.doc_type or '').upper() != 'BK':
        abort(404)

    st = (bk.status or '').upper()
    if st != 'CLOSED':
        flash('เอกสารยังไม่ได้ปิดงาน', 'info')
        return redirect(url_for('bk_view', doc_id=bk.id))

    bk.status = 'APPROVED'

    reopen_parent = (request.form.get('reopen_parent') or '').strip() in ['1','true','True','on','yes']
    if reopen_parent and bk.parent and (bk.parent.doc_type or '').upper() == 'QU' and (bk.parent.status or '').upper() == 'CLOSED':
        bk.parent.status = 'APPROVED'

    db.session.commit()
    flash('เปิดงานอีกครั้งแล้ว', 'success')
    return redirect(url_for('bk_view', doc_id=bk.id))




@app.post("/sales/bills/<int:did>/toggle")
@permission_required("sales.manage")
def bl_toggle(did):
    d = SalesDoc.query.get_or_404(did)
    if (d.doc_type or "").upper() != "BL":
        abort(404)

    d.status = "PAID" if (d.status or "").upper() != "PAID" else "UNPAID"

    # update booking flow if this doc belongs to BK
    try:
        if getattr(d, "parent", None) and (d.parent.doc_type or "").upper() == "BK":
            _update_booking_flow_state(d.parent)
    except Exception:
        pass

    db.session.commit()
    return redirect(url_for("bl_view", did=did))


@app.post("/sales/invoices/<int:did>/toggle")
@permission_required("sales.manage")
def iv_toggle(did):
    d = SalesDoc.query.get_or_404(did)
    if (d.doc_type or "").upper() != "IV":
        abort(404)

    d.status = "ISSUED" if (d.status or "").upper() != "ISSUED" else "UNISSUED"

    # update booking flow if this doc belongs to BK
    try:
        if getattr(d, "parent", None) and (d.parent.doc_type or "").upper() == "BK":
            _update_booking_flow_state(d.parent)
    except Exception:
        pass

    db.session.commit()
    return redirect(url_for("iv_view", did=did))


@app.post("/sales/receipts/<int:did>/toggle")
@permission_required("sales.manage")
def rc_toggle(did):
    d = SalesDoc.query.get_or_404(did)
    if d.doc_type != "RC":
        abort(404)
    d.status = "ISSUED" if (d.status or "").upper() != "ISSUED" else "UNISSUED"
    # update booking flow if this doc belongs to BK
    if d.parent and d.parent.doc_type == "BK":
        _update_booking_flow_state(d.parent)
    db.session.commit()
    return redirect(url_for("rc_view", did=did))

@app.route("/sales/bills/<int:did>/print")
@permission_required("sales.view")
def bl_print(did):
    d = (SalesDoc.query
         .options(
             joinedload(SalesDoc.items),
             joinedload(SalesDoc.customer),
             selectinload(SalesDoc.parent),
         )
         .get_or_404(did))
    img_map = _build_item_img_map(d)
    return render_template("sales/sd_print.html",
                           d=d,
                           today=date.today(),
                           mode="print",
                           img_map=img_map)

@app.route("/sales/invoices/<int:did>/print")
@permission_required("sales.view")
def iv_print(did):
    d = (SalesDoc.query
         .options(
             joinedload(SalesDoc.items),
             joinedload(SalesDoc.customer),
             selectinload(SalesDoc.parent),
         )
         .get_or_404(did))
    bl_ref = (SalesDoc.query
              .filter_by(parent_id=d.parent_id, doc_type="BL")
              .order_by(SalesDoc.id.desc())
              .first())
    img_map = _build_item_img_map(d)
    return render_template("sales/sd_print.html",
                           d=d,
                           bl_ref=bl_ref,
                           today=date.today(),
                           mode="print",
                           img_map=img_map)

@app.route("/sales/receipts/<int:did>/print")
@permission_required("sales.view")
def rc_print(did):
    d = (SalesDoc.query
         .options(
             joinedload(SalesDoc.items),
             joinedload(SalesDoc.customer),
             selectinload(SalesDoc.parent),
         )
         .get_or_404(did))
    bl_ref = (SalesDoc.query
              .filter_by(parent_id=d.parent_id, doc_type="BL")
              .order_by(SalesDoc.id.desc())
              .first())
    iv_ref = (SalesDoc.query
              .filter_by(parent_id=d.parent_id, doc_type="IV")
              .order_by(SalesDoc.id.desc())
              .first())
    img_map = _build_item_img_map(d)
    return render_template("sales/sd_print.html",
                           d=d,
                           bl_ref=bl_ref,
                           iv_ref=iv_ref,
                           today=date.today(),
                           mode="print",
                           img_map=img_map)

# ---- API: Active promotions (วันนี้) ----
@app.get("/api/promos/active")
@permission_required("promos.view")
def api_promos_active():
    today = date.today()
    promos = (Promotion.query
              .filter(Promotion.active==True)
              .filter((Promotion.start_date==None) | (Promotion.start_date <= today))
              .filter((Promotion.end_date==None)   | (Promotion.end_date >= today))
              .order_by(Promotion.id.desc())
              .all())
    def _row(p: Promotion):
        return {
            "id": p.id,
            "name": p.name,
            "active": bool(p.active),
            "start_date": p.start_date.isoformat() if p.start_date else None,
            "end_date": p.end_date.isoformat() if p.end_date else None,
            "min_items": p.min_items or 0,
            "rental_unit": (p.rental_unit or "DAY").upper(),
            "min_duration": p.min_duration or 0,
            "discount_type": (p.discount_type or "PCT").upper(),
            "discount_value": float(p.discount_value or 0),
            "cheapest_units_to_discount": p.cheapest_units_to_discount or 1,
        }
    return jsonify([_row(p) for p in promos])

@app.template_filter("unit_th_condensed")
def unit_th_condensed(u: str) -> str:
    m = {"DAY": "วัน", "MONTH": "เดือน", "YEAR": "ปี"}
    return m.get((u or "").upper(), u or "")


# ========== SPARE PARTS ROUTES ==========
@app.route("/spares")
@permission_required("spares.view")
def spares_list():
    ep = "spares_list"
    q = request.args.get("q", "").strip()
    query = SparePart.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(SparePart.code.ilike(like), SparePart.name.ilike(like)))
    rows = query.order_by(SparePart.code.asc()).all()
    return render_template("spares/list.html", ep=ep, rows=rows, q=q)

@app.route("/spares/new", methods=["GET", "POST"])
@permission_required("spares.create")
def spare_new():
    ep = "spare_new"
    if request.method == "POST":
        code = request.form.get("code","").strip()
        name = request.form.get("name","").strip()
        unit = request.form.get("unit","ชิ้น").strip() or "ชิ้น"
        unit_cost = Decimal(request.form.get("unit_cost","0") or "0")
        stock_qty = Decimal(request.form.get("stock_qty","0") or "0")
        notes = request.form.get("notes","").strip()
        sp = SparePart(code=code, name=name, unit=unit, unit_cost=unit_cost, stock_qty=stock_qty, notes=notes)
        db.session.add(sp)
        db.session.commit()
        flash("เพิ่มอะไหล่เรียบร้อย", "success")
        return redirect(url_for("spares_list"))
    return render_template("spares/form.html", ep=ep, mode="new")

@app.route("/spares/<int:sid>/edit", methods=["GET", "POST"])
@permission_required("spares.edit")
def spare_edit(sid):
    ep = "spare_edit"
    sp = SparePart.query.get_or_404(sid)
    if request.method == "POST":
        sp.code = request.form.get("code","").strip() or sp.code
        sp.name = request.form.get("name","").strip() or sp.name
        sp.unit = request.form.get("unit","ชิ้น").strip() or "ชิ้น"
        sp.unit_cost = Decimal(request.form.get("unit_cost","0") or "0")
        sp.stock_qty = Decimal(request.form.get("stock_qty","0") or "0")
        sp.notes = request.form.get("notes","").strip()
        db.session.commit()
        flash("บันทึกแล้ว", "success")
        return redirect(url_for("spares_list"))
    return render_template("spares/form.html", ep=ep, mode="edit", sp=sp)

# ========== CLAIMS ROUTES ==========

def _extract_sku_tokens(text: str) -> list[str]:
    """
    ดึง token ที่อยู่ใน [] เช่น "... [SP-001-081124-001]" -> ["SP-001-081124-001"]
    และเผื่อกรณีมีหลายอัน
    """
    if not text:
        return []
    return re.findall(r"\[([A-Za-z0-9_\-\.]+)\]", text)  # ดึงทุกตัวที่อยู่ใน []

def _resolve_equipment_from_claim_item(claim_item):
    """พยายามหา Equipment จาก ClaimItem ให้ได้มากที่สุด"""
    # 1) อุปกรณ์ทดแทนในใบเคลม (ตรงตัวที่สุด)
    eq = getattr(claim_item, "replacement_equipment", None)
    if eq:
        return eq

    # 2) กรณีโปรเจคในอนาคตมี field นี้ (ปัจจุบันไม่มี) — กันไว้
    if getattr(claim_item, "equipment_id", None):
        try:
            e = Equipment.query.get(int(claim_item.equipment_id))
            if e:
                return e
        except Exception:
            pass

    # 3) ใช้ sales_item เพื่อไล่หา
    si = getattr(claim_item, "sales_item", None)
    if si:
        # 3.1 ถ้ามีฟิลด์ sku (เผื่อเพิ่มในอนาคต)
        sku = _norm_sku(getattr(si, "sku", None))
        if sku:
            eq = Equipment.query.filter_by(sku=sku).first()
            if not eq:
                eq = Equipment.query.filter(Equipment.sku.ilike(f"%{sku}%")).first()
            if eq:
                return eq

        # 3.2 หา [SKU] จากชื่อ
        for token in _extract_tokens_from_text(getattr(si, "name", "")):
            eq = Equipment.query.filter_by(sku=token).first()
            if eq:
                return eq
            eq = Equipment.query.filter(Equipment.sku.ilike(f"%{token}%")).first()
            if eq:
                return eq

        # 3.3 สุดท้ายลองจับคู่ด้วยชื่อ (เผื่อไม่ได้ใส่วงเล็บ)
        name = str(getattr(si, "name", "")).strip()
        if name:
            eq = Equipment.query.filter(Equipment.name.ilike(f"%{name}%")).first()
            if eq:
                return eq

    return None

def _resolve_equipment_from_sales_item(si):
    """
    พยายามหา Equipment จากรายการในใบเสนอราคา (SalesDoc item)
    ใช้ตอนสร้างใบคืนสินค้า
    """
    if not si:
        return None

    # 1) ถ้ามีฟิลด์ equipment_id (กรณีผูกตรงอยู่แล้ว)
    eq_id = getattr(si, "equipment_id", None)
    if eq_id:
        try:
            e = Equipment.query.get(int(eq_id))
            if e:
                return e
        except Exception:
            pass

    # 2) ลองหา [SKU] จากชื่อ เช่น "สว่านไร้สาย [DR-001-241117-001]"
    name = str(getattr(si, "name", "") or "").strip()
    if name:
        m = re.search(r"\[([^\[\]]+?)\]", name)
        if m:
            sku = m.group(1).strip()
            eq = Equipment.query.filter_by(sku=sku).first()
            if not eq:
                eq = Equipment.query.filter(Equipment.sku.ilike(f"%{sku}%")).first()
            if eq:
                return eq

        # 3) สุดท้ายลองจับคู่ด้วยชื่อเต็ม
        eq = Equipment.query.filter(Equipment.name.ilike(f"%{name}%")).first()
        if eq:
            return eq

    return None




@app.route("/claims")
@permission_required("claims.view")
def claims_list():
    ep = "claims_list"
    q = request.args.get("q", "").strip()

    # ตอนนี้ยังไม่ได้ใช้ q filter อะไร เพิ่มทีหลังก็ได้
    rows = Claim.query.order_by(Claim.date.desc(), Claim.number.desc()).all()

    # --- เพิ่มส่วน map หาใบส่งสินค้าที่สร้างจากเคลมแต่ละใบ ---
    ids = [c.id for c in rows]
    deliveries_map = {}
    if ids:
        dls = (
            DeliveryDoc.query
            .filter(
                DeliveryDoc.source_type == "CLAIM",
                DeliveryDoc.source_id.in_(ids),
            )
            .all()
        )
        deliveries_map = {d.source_id: d for d in dls}

    return render_template(
        "claims/list.html",
        ep=ep,
        rows=rows,
        q=q,
        deliveries_map=deliveries_map,   # ✅ ส่งไปให้ template ใช้
    )


@app.route("/claims/new", methods=["GET","POST"])
@permission_required("claims.manage")
def claim_new():
    ep = "claim_new"
    customers = Customer.query.order_by(Customer.name.asc()).all()
    selected_customer_id = request.args.get("customer_id", type=int)
    quotes = []
    if selected_customer_id:
        quotes = (SalesDoc.query
                  .filter_by(doc_type="QU", customer_id=selected_customer_id, status="APPROVED")
                  .order_by(SalesDoc.date.desc())
                  .all())
    if request.method == "POST":
        customer_id = request.form.get("customer_id", type=int)
        quote_id = request.form.get("quote_id", type=int)
        if not (customer_id and quote_id):
            flash("กรุณาเลือกลูกค้าและใบเสนอราคา", "warning")
            return redirect(url_for("claim_new"))
        qu = SalesDoc.query.get_or_404(quote_id)
        if qu.doc_type != "QU" or (qu.status or "").upper() != "APPROVED":
            flash("ต้องเลือกใบเสนอราคาที่อนุมัติแล้วเท่านั้น", "warning")
            return redirect(url_for("claim_new", customer_id=customer_id))
        return redirect(url_for("claim_build", quote_id=quote_id))
    return render_template("claims/new.html",
                           ep=ep,
                           customers=customers,
                           selected_customer_id=selected_customer_id,
                           quotes=quotes)

@app.route("/claims/build/<int:quote_id>", methods=["GET","POST"])
@permission_required("claims.manage")
def claim_build(quote_id):
    ep = "claim_new"
    qu = SalesDoc.query.get_or_404(quote_id)
    if qu.doc_type != "QU" or (qu.status or "").upper() != "APPROVED":
        flash("ต้องใช้ใบเสนอราคาที่อนุมัติแล้วเท่านั้น", "warning")
        return redirect(url_for("claim_new", customer_id=qu.customer_id))
    ready_equips = Equipment.query.filter_by(status="READY").order_by(Equipment.name.asc()).all()
    if request.method == "POST":
        clm = Claim(
            number=_next_claim_number_by_date_with_prefix("CL", date.today()),
            date=date.today(),
            status="SUBMITTED",
            customer_id=qu.customer_id,
            quote_id=qu.id,
            remark=request.form.get("remark","").strip()
        )
        db.session.add(clm)
        db.session.flush()
        for it in qu.items:
            if request.form.get(f"claim_item_{it.id}"):
                qty = float(request.form.get(f"qty_{it.id}", "1") or "1")
                repl_id_val = request.form.get(f"repl_{it.id}")
                repl_id = int(repl_id_val) if (repl_id_val and repl_id_val.isdigit()) else None
                ci = ClaimItem(
                    claim_id=clm.id,
                    sales_item_id=it.id,
                    qty_claim=qty,
                    replacement_equipment_id=repl_id
                )
                db.session.add(ci)
                if repl_id:
                    equip_repl = Equipment.query.get(repl_id)
                    if equip_repl:
                        if equip_repl.status != "RENTED":
                            equip_repl.status = "RENTED"
                        _equip_log(
                            equip_repl,
                            action="ส่งทดแทน",
                            note=f"ทดแทนในใบเคลม {clm.number} อ้างอิง QU {qu.number}",
                            ref_model="Claim",
                            ref_id=clm.id
                        )
                orig_equip = None
                orig_equip_id = getattr(it, "equipment_id", None)
                if orig_equip_id:
                    orig_equip = Equipment.query.get(orig_equip_id)
                else:
                    import re
                    m = re.search(r"\[([^\[\]]+?)\]", it.name or "")
                    if m:
                        sku = m.group(1).strip()
                        orig_equip = Equipment.query.filter_by(sku=sku).first()
                if orig_equip:
                    if orig_equip.status != "REPAIR":
                        orig_equip.status = "REPAIR"
                    _equip_log(
                        orig_equip,
                        action="เข้ารอซ่อม (จากเคลม)",
                        note=f"เข้ารอซ่อมจากใบเคลม {clm.number} อ้างอิง QU {qu.number}",
                        ref_model="Claim",
                        ref_id=clm.id
                    )
        db.session.commit()
        flash(f"สร้างใบเคลม {clm.number} แล้ว", "success")
        return redirect(url_for("repairs.list_", show="pending"))
    return render_template("claims/build.html", ep=ep, qu=qu, ready_equips=ready_equips)

@app.route("/claims/<int:claim_id>")
@login_required
@permission_required("claims.view")
def claim_view(claim_id):
    ep = "claim_view"
    c = Claim.query.get_or_404(claim_id)
    return render_template("claims/view.html", ep=ep, c=c)

@app.route("/claims/<int:claim_id>/print")
def claim_print(claim_id):
    from flask import url_for
    import os, re, glob
    c = Claim.query.get_or_404(claim_id)
    def to_url_from_path(p: str | None):
        if not p: return None
        p = str(p)
        if p.startswith("http"): return p
        if p.startswith("static/"): return "/" + p
        return url_for("static", filename=p)
    def file_exists_rel(relpath: str) -> bool:
        return os.path.exists(os.path.join(app.static_folder, relpath))
    def normalize_sku(s):
        if not s: return None
        return str(s).strip().replace("\u200b","").replace("\ufeff","")
    def extract_sku_from_any(it):
        for obj in (getattr(it, "sales_item", None), getattr(it, "equipment", None), it):
            if not obj: continue
            for attr in ("sku", "item_sku", "code", "item_code"):
                v = getattr(obj, attr, None)
                if v: return normalize_sku(v)
        for obj in (getattr(it, "sales_item", None), getattr(it, "equipment", None), it):
            if not obj: continue
            for attr in ("name", "item_name", "title", "desc", "description"):
                s = getattr(obj, attr, None)
                if not s: continue
                m = re.search(r"\[([^\]]+)\]", str(s))
                if m: return normalize_sku(m.group(1))
        return None
    EXT = (".jpg", ".jpeg", ".png", ".webp", ".gif")
    SEARCH_DIRS = ["uploads/equipment","uploads/equipments","uploads/equipment_img","uploads/images","uploads"]
    def find_by_sku(sku: str | None):
        if not sku: return None
        sku = normalize_sku(sku)
        for base in (f"uploads/equipment/equip_{sku}", f"uploads/equipment/{sku}"):
            for ext in EXT:
                rel = base + ext
                if file_exists_rel(rel):
                    return url_for("static", filename=rel)
        for d in SEARCH_DIRS:
            abs_dir = os.path.join(app.static_folder, d)
            if not os.path.isdir(abs_dir): continue
            for path in glob.glob(os.path.join(abs_dir, "*.*")):
                fname = os.path.basename(path).lower()
                if any(fname.endswith(e) for e in EXT) and (sku.lower() in fname):
                    rel = os.path.relpath(path, app.static_folder).replace("\\", "/")
                    return url_for("static", filename=rel)
        return None
    def left_img_url(it):
        for obj in filter(None, [
            getattr(it, "sales_item", None),
            getattr(getattr(it, "sales_item", None), "equipment", None),
            getattr(it, "equipment", None),
        ]):
            for attr in ("image_path","photo_path","image","photo","image_url","photo_url"):
                v = getattr(obj, attr, None)
                if v:
                    u = to_url_from_path(v)
                    if u: return u
        sku = extract_sku_from_any(it)
        return find_by_sku(sku)
    def right_img_url(it):
        eq = getattr(it, "replacement_equipment", None)
        if not eq: return None
        for attr in ("image_path","photo_path","image","photo","image_url","photo_url"):
            v = getattr(eq, attr, None)
            if v:
                u = to_url_from_path(v)
                if u: return u
        return find_by_sku(getattr(eq, "sku", None))
    img_left, img_right, dbg = {}, {}, {}
    for it in c.items:
        uL = left_img_url(it)
        uR = right_img_url(it)
        img_left[it.id]  = uL
        img_right[it.id] = uR
        dbg[it.id] = {"sku_left": extract_sku_from_any(it), "left": uL, "right": uR}
        print("[CLAIM_PRINT] item", it.id, "sku_left=", dbg[it.id]["sku_left"], "-> left_url=", uL)
    return render_template("claims/print.html",
        c=c, img_left=img_left, img_right=img_right, dbg=dbg)

# ==== Thai status display helpers ====
THAI_STATUS = {
    "READY":  "พร้อมใช้งาน",
    "RENTED": "ถูกเช่า",
    "REPAIR": "รอซ่อม",
    "CLAIMED":"อยู่ระหว่างเคลม",
    "LOST":   "สูญหาย",
    "SCRAP":  "ตัดจำหน่าย",
}
def status_th(code: str) -> str:
    return THAI_STATUS.get((code or "").upper(), code or "")

import re
def _next_claim_number_by_date_with_prefix(prefix: str = "CL", dt: date | None = None) -> str:
    dt = dt or date.today()
    yyyymmdd = dt.strftime("%Y%m%d")
    prefix_today = f"{prefix}{yyyymmdd}"
    like_prefix = f"{prefix_today}%"
    last = (db.session.query(Claim)
            .filter(Claim.number.like(like_prefix))
            .order_by(Claim.number.desc())
            .first())
    if not last:
        return f"{prefix_today}001"
    m = re.match(rf"^{prefix_today}(\d{{3}})$", last.number or "")
    if not m:
        return f"{prefix_today}001"
    seq = int(m.group(1)) + 1
    return f"{prefix_today}{seq:03d}"

@app.context_processor
def _inject_helpers():
    return dict(status_th=status_th)

# ---- PRINT CLAIM (unique endpoint) ----
@app.get("/claims/<int:cid>/print", endpoint="claims_print")
@login_required
@permission_required("claims.view")
def claims_print(cid):
    from flask import url_for
    import os
    c = Claim.query.get_or_404(cid)
    def eq_img_url(e):
        if not e: return None
        for attr in ("image_path", "photo_path", "image", "photo", "image_url", "photo_url"):
            p = getattr(e, attr, None)
            if p:
                if p.startswith("http"): return p
                if p.startswith("uploads/"): return url_for("static", filename=p)
                if p.startswith("static/"): return "/" + p
                return url_for("static", filename=p)
        candidates = [
            f"uploads/equipment/equip_{e.sku}.jpg",
            f"uploads/equipment/equip_{e.sku}.png",
            f"uploads/equipment/{e.sku}.jpg",
            f"uploads/equipment/{e.sku}.png",
        ]
        for rel in candidates:
            abs_path = os.path.join(app.static_folder, rel)
            if os.path.exists(abs_path):
                return url_for("static", filename=rel)
        return None
    img_left, img_right = {}, {}
    for it in c.items:
        eq_left = getattr(getattr(it, "sales_item", None), "equipment", None) or getattr(it, "equipment", None)
        img_left[it.id] = eq_img_url(eq_left)
        img_right[it.id] = eq_img_url(getattr(it, "replacement_equipment", None))
    return render_template(
        "claims/print.html",
        c=c,
        img_left=img_left,
        img_right=img_right,
        no_container=True
    )



# ========== RETURNS ROUTES (ใบคืนสินค้า) ==========

@app.route("/returns")
@login_required
@permission_required("sales.manage")
def returns_list():
    ep = "returns_list"
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    # ดึงใบคืน + ลูกค้า + ใบเสนอราคา + รายการ
    query = ReturnDoc.query.options(
        joinedload(ReturnDoc.customer),
        joinedload(ReturnDoc.quote),
        joinedload(ReturnDoc.booking),
        joinedload(ReturnDoc.items),
    )

    # ถ้ามี field is_deleted ให้กรองออก (กันกรณี soft delete)
    is_del_col = getattr(ReturnDoc, "is_deleted", None)
    if is_del_col is not None:
        query = query.filter(is_del_col.is_(False))

    # เลือกช่วงวันที่ (ตาม ReturnDoc.date)
    if start_d:
        query = query.filter(ReturnDoc.date >= start_d)
    if end_d:
        query = query.filter(ReturnDoc.date <= end_d)

    # ค้นหาจาก เลขที่ใบคืน / เลขที่ใบเสนอราคา / ชื่อลูกค้า
    if q:
        like = f"%{q}%"
        QuoteDoc = aliased(SalesDoc)
        BookingDoc = aliased(SalesDoc)
        query = (
            query
            .outerjoin(ReturnDoc.customer)
            .outerjoin(QuoteDoc, ReturnDoc.quote_id == QuoteDoc.id)
            .outerjoin(BookingDoc, ReturnDoc.booking_id == BookingDoc.id)
            .filter(
                or_(
                    ReturnDoc.number.ilike(like),
                    QuoteDoc.number.ilike(like),
                    BookingDoc.number.ilike(like),
                    Customer.name.ilike(like),
                )
            )
        )

    docs = (
        query
        .order_by(ReturnDoc.date.desc(), ReturnDoc.id.desc())
        .all()
    )

    return render_template(
        "returns/list.html",
        ep=ep,
        docs=docs,
        q=q,
        start=start,
        end=end,
    )


@app.route("/returns/new", methods=["GET", "POST"])
@login_required
@permission_required("sales.manage")
def returns_new():
    """
    สร้างใบคืนสินค้า / คืนอุปกรณ์เช่า

    mode=QU (default): เลือก QU ที่อนุมัติแล้วและยังไม่เคยสร้างใบคืน
    mode=BK: เลือก BK ที่อนุมัติแล้ว (คืนจากรายการ SKU ที่ allocate)
    """
    ep = "returns_new"

    mode = (request.args.get("mode") or "QU").strip().upper()
    if mode not in ("QU", "BK"):
        mode = "QU"

    customers = Customer.query.order_by(Customer.name.asc()).all()

    selected_customer_id = request.args.get("customer_id", type=int)
    quotes = []
    bookings = []

    if selected_customer_id:
        if mode == "QU":
            # QU ที่ถูกออกใบคืนสินค้าไปแล้ว (อ้าง QU)
            subq = db.session.query(ReturnDoc.quote_id).filter(ReturnDoc.quote_id.isnot(None)).subquery()

            quotes = (
                SalesDoc.query
                .filter(
                    SalesDoc.doc_type == "QU",
                    SalesDoc.customer_id == selected_customer_id,
                    func.upper(SalesDoc.status) == "APPROVED",
                    ~SalesDoc.id.in_(subq),
                )
                .order_by(SalesDoc.date.desc(), SalesDoc.number.desc())
                .all()
            )
        else:
            # BK คืนได้หลายครั้ง (partial return) -> ไม่ต้องกันซ้ำ
            bookings = (
                SalesDoc.query
                .options(joinedload(SalesDoc.customer))
                .filter(
                    SalesDoc.doc_type == "BK",
                    SalesDoc.customer_id == selected_customer_id,
                    func.upper(SalesDoc.status).in_(["APPROVED", "CLOSED", "OPEN"]),
                )
                .order_by(SalesDoc.date.desc(), SalesDoc.number.desc())
                .all()
            )

    if request.method == "POST":
        customer_id = request.form.get("customer_id", type=int)
        doc_id = request.form.get("doc_id", type=int)
        mode_post = (request.form.get("mode") or mode).strip().upper()

        if not (customer_id and doc_id):
            flash("กรุณาเลือกลูกค้าและเอกสารอ้างอิง", "warning")
            return redirect(url_for("returns_new", mode=mode_post))

        if mode_post == "BK":
            bk = SalesDoc.query.get_or_404(doc_id)
            if bk.doc_type != "BK":
                flash("ต้องใช้ใบจอง (BK) เท่านั้น", "warning")
                return redirect(url_for("returns_new", mode="BK", customer_id=customer_id))
            if (bk.status or "").upper() not in ("APPROVED", "OPEN", "CLOSED"):
                flash("ใบจองต้องอยู่สถานะ อนุมัติ/เปิดงาน/ปิดงาน เท่านั้น", "warning")
                return redirect(url_for("returns_new", mode="BK", customer_id=customer_id))
            return redirect(url_for("returns_build_bk", bk_id=bk.id))

        # default QU
        qu = SalesDoc.query.get_or_404(doc_id)
        if qu.doc_type != "QU" or (qu.status or "").upper() != "APPROVED":
            flash("ต้องใช้ใบเสนอราคาที่อนุมัติแล้วเท่านั้น", "warning")
            return redirect(url_for("returns_new", mode="QU", customer_id=customer_id))

        return redirect(url_for("returns_build", quote_id=doc_id))

    return render_template(
        "returns/new.html",
        ep=ep,
        mode=mode,
        customers=customers,
        selected_customer_id=selected_customer_id,
        quotes=quotes,
        bookings=bookings,
    )


@app.route("/returns/build/<int:quote_id>", methods=["GET", "POST"])
@login_required
@permission_required("sales.manage")
def returns_build(quote_id):
    ep = "returns_build"

    qu: SalesDoc = (
        SalesDoc.query
        .options(
            joinedload(SalesDoc.items),
            joinedload(SalesDoc.customer),
        )
        .get_or_404(quote_id)
    )

    # ต้องเป็นใบ QU ที่อนุมัติแล้วเท่านั้น
    if qu.doc_type != "QU" or (qu.status or "").upper() != "APPROVED":
        flash("ต้องใช้ใบเสนอราคาที่อนุมัติแล้วเท่านั้น", "warning")
        return redirect(url_for("returns_new", customer_id=qu.customer_id))

    # เตรียม rows สำหรับแสดงในหน้า build
    rows = []
    for it in qu.items:
        eq = _resolve_equipment_from_sales_item(it)
        if not eq:
            continue
        rows.append(SimpleNamespace(item=it, equipment=eq))

    if request.method == "POST":
        # ดูว่า form ส่ง field อะไรมาบ้าง (ช่วย debug)
        form_keys = list(request.form.keys())
        print("DEBUG returns_build form keys:", form_keys)

        selected = []

        # -------------------------------
        # รูปแบบใหม่: row_enabled[] + item_id[] + return_qty[]
        # -------------------------------
        item_ids = request.form.getlist("item_id[]")
        qty_list = request.form.getlist("return_qty[]")
        enabled_idx_raw = request.form.getlist("row_enabled[]")
        note_list = request.form.getlist("item_note[]")

        if item_ids:
            enabled_idx = {int(x) for x in enabled_idx_raw if x.isdigit()}

            for idx, (sid, qty_str) in enumerate(zip(item_ids, qty_list)):
                if idx not in enabled_idx:
                    continue

                try:
                    qty = float(qty_str)
                except (TypeError, ValueError):
                    qty = 0

                if qty <= 0:
                    continue

                # หา row ที่ตรงกับ sales_item.id
                row = next((r for r in rows if str(r.item.id) == str(sid)), None)
                if not row:
                    continue

                it = row.item
                eq = row.equipment
                selected.append((it, eq, qty))

        # -------------------------------
        # รูปแบบเก่า: return_item_<id> + qty_<id>
        # -------------------------------
        if not selected:
            for r in rows:
                it = r.item
                eq = r.equipment

                flag_name = f"return_item_{it.id}"
                if not request.form.get(flag_name):
                    continue

                qty = request.form.get(f"qty_{it.id}", type=float) or 0
                if qty <= 0:
                    continue

                selected.append((it, eq, qty))

        # ถ้าไม่มีอะไรถูกเลือกเลย
        if not selected:
            flash("กรุณาเลือกรายการที่จะคืนอย่างน้อย 1 รายการ", "warning")
            return redirect(url_for("returns_build", quote_id=quote_id))

        # สร้างเอกสารใบคืนสินค้า
        ret = ReturnDoc(
            number=_next_return_number_by_date_with_prefix("RT", date.today()),
            date=date.today(),
            customer_id=qu.customer_id,
            quote_id=qu.id,
            remark=(request.form.get("remark") or "").strip(),
            created_by=current_user.id if current_user.is_authenticated else None,
        )
        db.session.add(ret)
        db.session.flush()  # ให้ได้ ret.id

        # สร้างรายการคืน + อัปเดตสถานะอุปกรณ์ + log
        for it, eq, qty in selected:
            db.session.add(ReturnItem(
                doc_id=ret.id,
                equipment_id=eq.id,
                qty=qty,
            ))

            prev_status = eq.status or "READY"
            if prev_status != "READY":
                eq.status = "READY"

            _equip_log(
                eq,
                action="RETURN",
                note=f"คืนจากใบคืนสินค้า {ret.number} อ้างอิง QU {qu.number}",
                ref_model="ReturnDoc",
                ref_id=ret.id,
            )

        db.session.commit()
        flash(f"สร้างใบคืนสินค้า {ret.number} แล้ว", "success")
        return redirect(url_for("returns_view", rid=ret.id))

    # GET: แสดงหน้าเลือกอุปกรณ์
    return render_template(
        "returns/build.html",
        ep=ep,
        qu=qu,
        rows=rows,
    )



@app.route("/returns/build-bk/<int:bk_id>", methods=["GET", "POST"])
@login_required
@permission_required("sales.manage")
def returns_build_bk(bk_id):
    ep = "returns_build_bk"

    bk: SalesDoc = (
        SalesDoc.query
        .options(
            joinedload(SalesDoc.items),
            joinedload(SalesDoc.customer),
        )
        .get_or_404(bk_id)
    )

    if (bk.doc_type or "").upper() != "BK":
        flash("ต้องใช้ใบจอง (BK) เท่านั้น", "warning")
        return redirect(url_for("returns_new", mode="BK", customer_id=bk.customer_id))

    # รวบรวม SKU ที่ถูก allocate (ยังไม่คืน)
    rows = []
    for it in (bk.items or []):
        sku_csv = (it.allocated_skus or "").strip()
        if not sku_csv:
            continue
        skus = [s.strip() for s in sku_csv.split(",") if s.strip()]
        for sku in skus:
            eq = Equipment.query.filter_by(sku=sku).first()
            if not eq:
                continue
            rows.append(SimpleNamespace(item=it, equipment=eq))

    if request.method == "POST":
        selected = []
        # row_enabled[] + equipment_id[] + condition[] + damage_note[] + damage_cost[]
        eq_ids = request.form.getlist("equipment_id[]")
        conds = request.form.getlist("condition[]")
        notes = request.form.getlist("damage_note[]")
        costs = request.form.getlist("damage_cost[]")
        enabled_idx_raw = request.form.getlist("row_enabled[]")

        enabled_idx = {int(x) for x in enabled_idx_raw if x.isdigit()}

        for idx, eqid in enumerate(eq_ids):
            if idx not in enabled_idx:
                continue
            try:
                eqid_int = int(eqid)
            except Exception:
                continue

            cond = (conds[idx] if idx < len(conds) else "GOOD").strip().upper() or "GOOD"
            if cond not in ("GOOD", "REPAIR", "LOST"):
                cond = "GOOD"

            note = (notes[idx] if idx < len(notes) else "").strip()

            try:
                cost = float(costs[idx]) if idx < len(costs) and (costs[idx] or "").strip() else 0.0
            except Exception:
                cost = 0.0

            row = next((r for r in rows if r.equipment.id == eqid_int), None)
            if not row:
                continue
            selected.append((row.item, row.equipment, cond, note, cost))

        if not selected:
            flash("กรุณาเลือกรายการที่จะคืนอย่างน้อย 1 รายการ", "warning")
            return redirect(url_for("returns_build_bk", bk_id=bk_id))

        # หา QU แม่ (ถ้ามี) เพื่อเก็บ reference
        qu_id = None
        try:
            if bk.parent_id:
                parent = SalesDoc.query.get(bk.parent_id)
                if parent and (parent.doc_type or "").upper() == "QU":
                    qu_id = parent.id
        except Exception:
            qu_id = None

        ret = ReturnDoc(
            number=_next_return_number_by_date_with_prefix("RT", date.today()),
            date=date.today(),
            customer_id=bk.customer_id,
            quote_id=qu_id,
            booking_id=bk.id,
            ref_type="BK",
            remark=(request.form.get("remark") or "").strip(),
            created_by=current_user.id if current_user.is_authenticated else None,
        )
        db.session.add(ret)
        db.session.flush()

        # ทำคืนทีละ SKU: อัปเดตสถานะอุปกรณ์ + ตัด SKU ออกจาก allocated_skus ใน BK
        for it, eq, cond, note, cost in selected:
            db.session.add(
                ReturnItem(
                    doc_id=ret.id,
                    equipment_id=eq.id,
                    qty=1,
                    condition=cond,
                    damage_note=note,
                    damage_cost=cost,
                )
            )

            # update equipment status
            prev_status = (eq.status or "READY").upper()
            new_status = "READY"
            if cond == "REPAIR":
                new_status = "REPAIR"
            elif cond == "LOST":
                new_status = "LOST"

            if prev_status != new_status:
                eq.status = new_status

            # remove SKU from allocated_skus (BK item) + sync ไป QU แม่ (ถ้ามี mapping)
            try:
                # --- 1) remove จาก BK item ---
                sku_list = [s.strip() for s in (it.allocated_skus or "").split(",") if s.strip()]
                if eq.sku in sku_list:
                    sku_list = [s for s in sku_list if s != eq.sku]
                    it.allocated_skus = ",".join(sku_list) if sku_list else None

                # --- 2) sync QU allocated_skus on return (ถ้ามี mapping) ---
                try:
                    if bk and bk.parent_id:
                        qu = SalesDoc.query.get(bk.parent_id)
                        if (
                            qu
                            and (qu.doc_type or "").upper() == "QU"
                            and getattr(it, "source_qu_item_id", None)
                        ):
                            qu_it = SalesItem.query.get(it.source_qu_item_id)
                            if qu_it and qu_it.doc_id == qu.id:
                                qu_skus = [s.strip() for s in (qu_it.allocated_skus or "").split(",") if s.strip()]
                                if eq.sku in qu_skus:
                                    qu_skus = [s for s in qu_skus if s != eq.sku]
                                    qu_it.allocated_skus = ",".join(qu_skus) if qu_skus else None
                except Exception:
                    pass

            except Exception:
                pass

            # log
            th_cond = {"GOOD": "คืนปกติ", "REPAIR": "ส่งซ่อม", "LOST": "สูญหาย"}.get(cond, cond)
            extra = f" ({th_cond})" + (f" | {note}" if note else "")
            if cost and cost > 0:
                extra += f" | ค่าเสียหาย {cost:,.2f}"

            _equip_log(
                eq,
                action="RETURN",
                note=f"คืนจากใบคืนสินค้า {ret.number} อ้างอิง BK {bk.number}{extra}",
                ref_model="ReturnDoc",
                ref_id=ret.id,
            )

        db.session.commit()
        flash(f"สร้างใบคืนสินค้า {ret.number} แล้ว", "success")
        return redirect(url_for("returns_view", rid=ret.id))

    return render_template(
        "returns/build_bk.html",
        ep=ep,
        bk=bk,
        rows=rows,
    )


@app.route("/returns/<int:rid>")
@login_required
@permission_required("sales.view")
def returns_view(rid):
    ep = "returns_view"

    doc: ReturnDoc = (
        ReturnDoc.query
        .options(
            joinedload(ReturnDoc.customer),
            joinedload(ReturnDoc.quote).joinedload(SalesDoc.customer),
            joinedload(ReturnDoc.booking).joinedload(SalesDoc.customer),
            joinedload(ReturnDoc.items).joinedload(ReturnItem.equipment),
        )
        .get_or_404(rid)
    )

    return render_template(
        "returns/view.html",
        ep=ep,
        d=doc,   # ⬅ ใน template ให้ใช้ d.items, d.customer, d.quote ฯลฯ
    )

@app.route("/returns/<int:rid>/print")
@login_required
@permission_required("sales.manage")
def returns_print(rid):
    # โหลดใบคืน + customer + quote + รายการ + อุปกรณ์
    d: ReturnDoc = (
        ReturnDoc.query
        .options(
            joinedload(ReturnDoc.customer),
            joinedload(ReturnDoc.quote),
        joinedload(ReturnDoc.booking),
            joinedload(ReturnDoc.items).joinedload(ReturnItem.equipment),
        )
        .get_or_404(rid)
    )

    # ----- หาใบส่งสินค้าที่อ้างอิงใบเสนอราคาเดียวกัน (ถ้ามี) -----
    delivery = None
    try:
        if d.quote_id:
            delivery = (
                DeliveryDoc.query
                .filter(DeliveryDoc.quote_id == d.quote_id)
                .order_by(DeliveryDoc.date.desc(), DeliveryDoc.id.desc())
                .first()
            )
    except Exception:
        delivery = None
    # -----------------------------------------------------------

    # ----- ดึงข้อมูลบริษัท (ถ้ามีโมเดล Company ให้ใช้ ถ้าไม่มีไม่พัง) -----
    # ดึงจาก globals แล้วเก็บลงตัวแปร local ชื่อ Company ใหม่
    Company = globals().get("Company")
    company = None
    if Company:
        try:
            company = Company.query.first()
        except Exception:
            company = None
    # -----------------------------------------------------------

    today = date.today()

    return render_template(
        "returns/print.html",
        d=d,
        company=company,
        today=today,
        delivery=delivery,
    )

bp_repairs = Blueprint("repairs", __name__, url_prefix="/repairs")

# อ้างถึงตารางที่คุณมีอยู่แล้วใน app.py
# Equipment, EquipmentLog, Claim, ClaimItem, Customer, SparePart  (ชื่ออาจต่างเล็กน้อย – ปรับให้ตรงของคุณ)

# --- แทนที่ฟังก์ชันนี้ทั้งก้อน ---




@bp_repairs.route("/")
@login_required
@permission_required("repairs.view")
def list_():
    """หน้ารายการงานซ่อม + รอเปิดงานซ่อม (จากใบเคลม)"""
    q = (request.args.get("q") or "").strip()
    show = (request.args.get("show") or "").lower()

    # ---------- งานซ่อมที่เปิดแล้ว ----------
    qs = RepairJob.query
    if q:
        like = f"%{q}%"
        qs = qs.filter(or_(RepairJob.number.ilike(like),
                           RepairJob.symptom.ilike(like)))
    jobs = qs.order_by(getattr(RepairJob, "opened_at", RepairJob.id).desc()).all()

    # ---------- รอเปิดงานซ่อมจากใบเคลม ----------
    # set ของ (claim_id, claim_item_id) ที่มี RepairJob แล้ว (ทั้ง OPEN/DONE) -> ไม่ต้องแสดงใน pending
    existing_pairs = {
        (cid, iid)
        for cid, iid in db.session.query(RepairJob.claim_id, RepairJob.claim_item_id)
        .filter(RepairJob.claim_id.isnot(None), RepairJob.claim_item_id.isnot(None))
        .all()
    }

    # เคลมที่ต้องการแสดงใน pending (ส่งคำขอแล้ว/อนุมัติแล้ว)
    allowed_statuses = ["SUBMITTED", "APPROVED"]

    clms = (
        Claim.query
        .options(joinedload(Claim.items), joinedload(Claim.customer))
        .filter(Claim.status.in_(allowed_statuses))
        .order_by(Claim.date.desc(), Claim.number.desc())
        .all()
    )

    pending = []
    for c in clms:
        for it in (c.items or []):
            if (c.id, it.id) in existing_pairs:
                continue  # มีงานซ่อมแล้ว ไม่ต้องโชว์

            si = getattr(it, "sales_item", None)
            item_name = (getattr(si, "name", None) or getattr(it, "item_name", "") or "").strip()

            eq_suggest = _resolve_equipment_from_claim_item(it)  # ถ้ามีฟังก์ชันเดาอุปกรณ์

            pending.append({
                "claim_id": c.id,
                "claim_number": c.number,
                "claim_date": c.date,
                "customer_name": getattr(c.customer, "name", ""),
                "item_id": it.id,
                "item_name": item_name,
                "qty": getattr(it, "qty_claim", None) or 1,
                "eq_sku": getattr(eq_suggest, "sku", None),
                "eq_name": getattr(eq_suggest, "name", None),
            })

    # คีย์เวิร์ดค้นหาใน pending
    if q:
        needle = q.lower()
        def _hit(row):
            hay = " ".join(str(row.get(k, "")) for k in
                           ("claim_number", "customer_name", "item_name", "eq_sku", "eq_name")).lower()
            return needle in hay
        pending = [x for x in pending if _hit(x)]

    return render_template("repairs/list.html",
                           jobs=jobs,
                           pending=pending,
                           show=show)




@bp_repairs.route("/<int:jid>")
@login_required
@permission_required("repairs.view")
def view_(jid):
    job = RepairJob.query.get_or_404(jid)
    eq = Equipment.query.get(job.equipment_id)
    cl = Claim.query.get(job.claim_id) if job.claim_id else None

    spare_list = _load_spares()

    # debug ใน console
    print("SPARES COUNT:", len(spare_list))
    for p in spare_list:
        print("PART:", p.id, p.code, p.name, p.unit_price)

    return render_template(
        "repairs/view.html",
        job=job, eq=eq, cl=cl,
        spare_list=spare_list
    )



@bp_repairs.route("/<int:jid>/save", methods=["POST"])
@login_required
@permission_required("repairs.manage")
def save_(jid):
    """บันทึกอาการ + ค่าแรง (ยังไม่ปิดงาน)"""
    job = RepairJob.query.get_or_404(jid)
    job.symptom = request.form.get("symptom", "").strip()
    job.labor_cost = _dec(request.form.get("labor_cost", "0"))
    # คำนวณยอดรวม
    parts_total = sum((_dec(it.line_total) for it in job.items), Decimal("0"))
    job.parts_total = parts_total
    job.total_cost = parts_total + _dec(job.labor_cost)
    job.status = "IN_PROGRESS"
    db.session.commit()
    flash("บันทึกงานซ่อมแล้ว", "success")
    return redirect(url_for("repairs.view_", jid=jid))

@bp_repairs.route("/<int:jid>/add_part", methods=["POST"])
@login_required
@permission_required("repairs.manage")
def add_part(jid):
    """เพิ่มอะไหล่ลงงานซ่อม (เรียกจากปุ่มเลือกอะไหล่)"""
    job = RepairJob.query.get_or_404(jid)
    part_id = int(request.form["part_id"])
    qty = _dec(request.form.get("qty", "1"))

    sp = SparePart.query.get_or_404(part_id)
    unit = _dec(getattr(sp, "unit_price", getattr(sp, "unit_cost", 0)))
    line = qty * unit

    ri = RepairItem(
        job_id=job.id,
        part_id=sp.id,
        part_code=sp.code,      # ปรับชื่อฟิลด์ code ตามจริง
        part_name=sp.name,      # ปรับชื่อฟิลด์ name ตามจริง
        qty=qty,
        unit_price=unit,
        line_total=line,
    )
    db.session.add(ri)

    # อัปเดตรวมชั่วคราว
    job.parts_total = (job.parts_total or 0) + line
    job.total_cost = _dec(job.parts_total) + _dec(job.labor_cost or 0)
    db.session.commit()
    flash("เพิ่มอะไหล่แล้ว", "success")
    return redirect(url_for("repairs.view_", jid=jid))

@bp_repairs.route("/<int:jid>/remove_part/<int:item_id>", methods=["POST"])
@login_required
@permission_required("repairs.manage")
def remove_part(jid, item_id):
    job = RepairJob.query.get_or_404(jid)
    it = RepairItem.query.get_or_404(item_id)
    db.session.delete(it)
    db.session.flush()
    # คำนวณรวมใหม่
    parts_total = sum((_dec(x.line_total) for x in job.items), Decimal("0"))
    job.parts_total = parts_total
    job.total_cost = parts_total + _dec(job.labor_cost or 0)
    db.session.commit()
    flash("ลบอะไหล่แล้ว", "success")
    return redirect(url_for("repairs.view_", jid=jid))



@bp_repairs.route("/<int:jid>/close", methods=["POST"])
@login_required
@permission_required("repairs.manage")
def close_(jid):
    job = RepairJob.query.get_or_404(jid)
    eq  = Equipment.query.get_or_404(job.equipment_id)

    # --- กันพลาด: คำนวณยอดรวมล่าสุดก่อนปิด ---
    parts_total = _dec("0")
    for it in job.items:
        qty  = _dec(it.qty or 0)
        unit = _dec(it.unit_price or 0)
        it.line_total = qty * unit
        parts_total  += it.line_total
    job.parts_total = parts_total
    job.total_cost  = parts_total + _dec(job.labor_cost or 0)

    # --- หักสต็อกอะไหล่ (ไม่ให้ติดลบ) ---
    for it in job.items:
        sp = SparePart.query.get(it.part_id)
        if not sp:
            continue
        sp.stock_qty = _dec(sp.stock_qty or 0) - _dec(it.qty or 0)
        if sp.stock_qty < 0:
            sp.stock_qty = _dec("0")

    # --- เปลี่ยนสถานะอุปกรณ์กลับพร้อมให้เช่า ---
    prev = eq.status
    eq.status = "READY"

    # --- เขียน log อุปกรณ์ ---
    db.session.add(EquipmentLog(
        equipment_id=eq.id,
        action="REPAIR_DONE",
        note=f"ปิดงานซ่อม {job.number} (ค่าแรง {job.labor_cost} + อะไหล่ {job.parts_total} = {job.total_cost})",
        user_id=(current_user.id if current_user.is_authenticated else None),
    ))

    # --- อัปเดตสถานะงานซ่อม ---
    job.status = "DONE"
    job.closed_at = datetime.utcnow()


    db.session.commit()

    flash(
        f"ปิดงานซ่อมแล้ว (อุปกรณ์: {eq.sku} จาก {EQUIP_STATUS_THAI.get(prev, prev)} → {EQUIP_STATUS_THAI.get(eq.status, eq.status)})",
        "success",
    )
    return redirect(url_for("repairs.view_", jid=jid))


@bp_repairs.route("/open-from-claim/<int:cid>/<int:item_id>", methods=["POST"])
@login_required
@permission_required("repairs.manage")  # ถ้าสิทธิ์ของคุณชื่ออื่น เช่น "maintenance.create" ก็เปลี่ยนให้ตรง
def open_from_claim(cid: int, item_id: int):
    # 1) ถ้ามีงานของรายการนี้อยู่แล้ว ให้ไปหน้านั้นเลย
    existed = (
        RepairJob.query
        .filter_by(claim_id=cid, claim_item_id=item_id)
        .first()
    )
    if existed:
        flash(f"มีงานซ่อม {existed.number} อยู่แล้วสำหรับรายการเคลมนี้", "info")
        return redirect(url_for("repairs.view_", jid=existed.id))

    # 2) ดึงเคลม + item
    c = Claim.query.get_or_404(cid)
    claim_item = next((x for x in (c.items or []) if x.id == item_id), None)
    if not claim_item:
        flash("ไม่พบรายการเคลมที่เลือก", "danger")
        return redirect(url_for("repairs.list_", show="pending"))

    # 3) หาอุปกรณ์จากรายการเคลม
    eq = _resolve_equipment_from_claim_item(claim_item)

    # ✅ สำคัญ: ถ้าตาราง repair_jobs.equipment_id เป็น NOT NULL ต้องบังคับให้หาได้ก่อน
    if not eq:
        flash("ไม่พบอุปกรณ์ที่ผูกกับรายการเคลมนี้ กรุณาเลือกอุปกรณ์ก่อนเปิดงานซ่อม", "warning")
        return redirect(url_for("claim_view", claim_id=c.id))


    # 4) สร้างงานซ่อม (ใส่ equipment_id ให้แน่ชัด)
    job = RepairJob(
        number=_gen_running("RJ", RepairJob),
        equipment_id=eq.id,       # ต้องไม่ None
        claim_id=c.id,
        claim_item_id=claim_item.id,
        customer_id=c.customer_id if getattr(c, "customer_id", None) else None,
        status="OPEN",
        opened_at=datetime.utcnow(),
        labor_cost=_dec("0"),
    )
    db.session.add(job)
    db.session.flush()  # เอา job.id

    # 5) อัปเดตสถานะอุปกรณ์ + log
    prev = eq.status
    eq.status = "REPAIR"
    db.session.add(EquipmentLog(
        equipment_id=eq.id,
        action="REPAIR_OPEN",
        note=f"เปิดงานซ่อม {job.number} จากเคลม {c.number}",
        user_id=(current_user.id if current_user.is_authenticated else None),
    ))

    db.session.commit()
    flash(f"เปิดงานซ่อม {job.number} เรียบร้อย", "success")
    return redirect(url_for("repairs.view_", jid=job.id))


# ==================== Deliveries Blueprint (routes) ====================


bp_deliveries = Blueprint("deliveries", __name__, url_prefix="/deliveries")

def require_perm(code):
    return permission_required(code)  # ใช้ของเดิมคุณ

@bp_deliveries.route("/")
@require_perm("transport.access")
def list_docs():
    q = (
        DeliveryDoc.query
        .options(
            joinedload(DeliveryDoc.vehicle),
            joinedload(DeliveryDoc.driver),
        )
        .order_by(DeliveryDoc.created_at.desc(), DeliveryDoc.id.desc())
    )
    rows = q.all()
    return render_template(
        "deliveries/list.html",
        rows=rows,          # ✅ ชื่อตรงกับใน template
        total=len(rows),    # ✅ ใช้แสดง "ทั้งหมด X รายการ"
    )


@bp_deliveries.route("/vehicles")
@require_perm("transport.manage")
def vehicles():
    items = DeliveryVehicle.query.order_by(DeliveryVehicle.code).all()
    return render_template("deliveries/vehicles.html", items=items)

@bp_deliveries.route("/vehicles/new", methods=["POST"])
@require_perm("transport.manage")
def vehicle_create():
    code = request.form.get("code","").strip()
    name = request.form.get("name","").strip()
    plate = request.form.get("plate_no","").strip()
    if not code or not name:
        flash("กรุณากรอก Code และ Name", "warning")
        return redirect(url_for("deliveries.vehicles"))
    v = DeliveryVehicle(code=code, name=name, plate_no=plate)
    db.session.add(v)
    db.session.commit()
    flash("เพิ่มรถเรียบร้อย", "success")
    return redirect(url_for("deliveries.vehicles"))

@bp_deliveries.route("/drivers")
@require_perm("transport.manage")
def drivers():
    items = Driver.query.order_by(Driver.code).all()
    return render_template("deliveries/drivers.html", items=items)

@bp_deliveries.route("/drivers/new", methods=["POST"])
@require_perm("transport.manage")
def driver_create():
    code = request.form.get("code","").strip()
    name = request.form.get("full_name","").strip()
    phone = request.form.get("phone","").strip()
    if not code or not name:
        flash("กรุณากรอก Code และชื่อคนขับ", "warning")
        return redirect(url_for("deliveries.drivers"))
    d = Driver(code=code, full_name=name, phone=phone)
    db.session.add(d)
    db.session.commit()
    flash("เพิ่มคนขับเรียบร้อย", "success")
    return redirect(url_for("deliveries.drivers"))




# --------- Create from CLAIM (DLC) ----------
@bp_deliveries.route("/create-from-claim/<int:cid>", methods=["GET", "POST"])
@require_perm("transport.access")
def create_from_claim(cid):
    # โหลดใบเคลม + ลูกค้า + QU อ้างอิง + รายการเคลม
    claim = (
        Claim.query
        .options(
            joinedload(Claim.customer),
            joinedload(Claim.quote),
            joinedload(Claim.items).joinedload(ClaimItem.sales_item),
        )
        .get_or_404(cid)
    )

    existing = DeliveryDoc.query.filter_by(source_type="CLAIM", source_id=cid).first()
    if existing and request.method == "GET":
        flash("มีใบส่งสินค้าถูกสร้างจากใบเคลมนี้แล้ว", "info")
        return redirect(url_for("deliveries.view_doc", did=existing.id))

    default_name  = claim.customer.name if claim.customer else ""
    default_phone = claim.customer.phone if claim.customer and getattr(claim.customer, "phone", None) else ""
    default_addr  = claim.customer.address if claim.customer and getattr(claim.customer, "address", None) else ""
    default_delivery_date = date.today()

    if request.method == "POST":
        ship_to_name    = (request.form.get("ship_to_name") or "").strip()
        ship_to_phone   = (request.form.get("ship_to_phone") or "").strip()
        ship_to_address = (request.form.get("ship_to_address") or "").strip()
        ship_to_note    = (request.form.get("ship_to_note") or "").strip()

        delivery_date_str = (request.form.get("delivery_date") or "").strip()
        delivery_date = None
        if delivery_date_str:
            try:
                delivery_date = datetime.strptime(delivery_date_str, "%Y-%m-%d").date()
            except ValueError:
                flash("รูปแบบวันที่จัดส่งไม่ถูกต้อง", "warning")
                return redirect(request.url)
        else:
            delivery_date = default_delivery_date

        number = _gen_running("DLC", DeliveryDoc)
        doc = DeliveryDoc(
            number=number,
            d_type=DeliveryType.DLC,
            status=DeliveryStatus.PENDING,
            source_type="CLAIM",
            source_id=cid,
            ship_to_name=ship_to_name or default_name,
            ship_to_phone=ship_to_phone or default_phone,
            ship_to_address=ship_to_address or default_addr,
            ship_to_note=ship_to_note or None,
            delivery_date=delivery_date,  # 👈 ตรงนี้เหมือนกัน
        )
        db.session.add(doc)
        db.session.flush()

        for ci in claim.items:
            src = ci.sales_item
            d_item = DeliveryItem(
                doc_id=doc.id,
                source_item_id=ci.id,
                product_name=src.name if src else f"อุปกรณ์จากเคลม #{cid}",
                qty=ci.qty_claim or 0,
                unit="ชิ้น",
                note=None,
            )
            db.session.add(d_item)

        db.session.commit()
        flash("สร้างใบส่งสินค้าเคลมเรียบร้อย", "success")
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    return render_template(
        "deliveries/create_from_source.html",
        quotation=None,
        claim=claim,
        items=claim.items,
        is_claim=True,
        default_delivery_date=default_delivery_date,
        back_url=url_for("claims_list"),
        source_type="CLAIM",
        source_id=cid,
        d_type="DLC",
    )

# --------- View / assign route (จัดสายรถ) ----------
@bp_deliveries.route("/<int:did>")
@require_perm("transport.access")
def view_doc(did):
    doc = (DeliveryDoc.query
           .options(joinedload(DeliveryDoc.items),
                    joinedload(DeliveryDoc.vehicle),
                    joinedload(DeliveryDoc.driver))
           .get_or_404(did))
    vehicles = DeliveryVehicle.query.filter_by(is_active=True).all()
    drivers  = Driver.query.filter_by(is_active=True).all()
    return render_template("deliveries/view.html", doc=doc, vehicles=vehicles, drivers=drivers)

@bp_deliveries.route("/<int:did>/assign", methods=["POST"])
@require_perm("transport.manage")
def assign_route(did):
    doc = DeliveryDoc.query.get_or_404(did)
    doc.vehicle_id = request.form.get("vehicle_id") or None
    doc.driver_id = request.form.get("driver_id") or None
    sch_date = request.form.get("schedule_date")  # 'YYYY-MM-DD'
    sch_time = request.form.get("schedule_time")  # 'HH:MM'
    if sch_date:
        dt_str = sch_date + (f" {sch_time}" if sch_time else " 09:00")
        doc.schedule_at = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
    db.session.commit()
    flash("อัปเดตการจัดสายรถแล้ว", "success")
    return redirect(url_for("deliveries.view_doc", did=doc.id))

# --------- Update Status ----------
@bp_deliveries.route("/<int:did>/status", methods=["POST"])
@permission_required("transport.manage")
def update_status(did):
    doc = DeliveryDoc.query.get_or_404(did)

    # สถานะเดิม (Enum หรือ string)
    if doc.status is not None and hasattr(doc.status, "name"):
        old_status = doc.status.name.upper()
    else:
        old_status = (str(doc.status or "PENDING")).upper()

    # สถานะใหม่ + note จากฟอร์ม
    new_status = (request.form.get("status") or old_status).upper()
    status_note = (request.form.get("status_note") or "").strip()

    # เช็คว่าเป็นฟอร์ม "ยกเลิกการส่ง" จริง ๆ ไหม
    cancel_form = (new_status == "CANCELLED") and (
        "cancel_reason" in request.form or "cancel_note" in request.form
    )

    print(
        f"[DELIVERY_STATUS] did={did} {old_status} -> {new_status} "
        f"(cancel_form={cancel_form})"
    )

    # สถานะที่อนุญาต
    valid_statuses = {"PENDING", "ONGOING", "DONE", "CANCELLED"}
    if new_status not in valid_statuses:
        flash("สถานะไม่ถูกต้อง", "danger")
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    # นับจำนวนรูปก่อน/หลังส่ง
    photos_before = getattr(doc, "photos_before", []) or []
    photos_after = getattr(doc, "photos_after", []) or []
    before_count = len(list(photos_before))
    after_count = len(list(photos_after))

    # rule: จะเปลี่ยนเป็นกำลังจัดส่ง ต้องมีรูปก่อนส่ง >= 3
    if new_status == "ONGOING" and before_count < 3:
        flash(
            "ต้องอัปโหลดรูปสินค้าก่อนส่งอย่างน้อย 3 ภาพ (สูงสุด 10) "
            "ก่อนเปลี่ยนเป็น 'กำลังจัดส่ง'",
            "danger",
        )
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    # rule: จะเปลี่ยนเป็นจัดส่งสำเร็จ ต้องมีรูปหลังส่ง >= 3
    if new_status == "DONE" and after_count < 3:
        flash(
            "ต้องอัปโหลดรูปส่งเสร็จอย่างน้อย 3 ภาพ (สูงสุด 10) "
            "ก่อนเปลี่ยนเป็น 'จัดส่งสำเร็จ'",
            "danger",
        )
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    # rule: ห้ามข้ามจาก PENDING → DONE/CANCELLED ตรง ๆ
    if new_status in {"DONE", "CANCELLED"} and old_status == "PENDING":
        flash(
            "ให้เปลี่ยนเป็น 'กำลังจัดส่ง' ก่อน แล้วจึงเปลี่ยนเป็น 'จัดส่งสำเร็จ' หรือ 'ยกเลิก'",
            "danger",
        )
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    # -------------------------------------------------
    # จัดการ field เหตุผลการยกเลิก (ใช้ cancel_reason_code / cancel_note)
    # -------------------------------------------------
    if cancel_form:
        cancel_reason = (request.form.get("cancel_reason") or "").strip()
        cancel_note = (request.form.get("cancel_note") or "").strip()

        if hasattr(doc, "cancel_reason_code"):
            doc.cancel_reason_code = cancel_reason or None
        if hasattr(doc, "cancel_note"):
            doc.cancel_note = cancel_note or None
    else:
        # ถ้าไม่ได้อยู่สถานะ CANCELLED เคลียร์เหตุผลยกเลิกทิ้ง
        if new_status != "CANCELLED":
            if hasattr(doc, "cancel_reason_code"):
                doc.cancel_reason_code = None
            if hasattr(doc, "cancel_note"):
                doc.cancel_note = None

    # บันทึกสถานะหลัก
    doc.status = new_status

    # เขียนหมายเหตุลง internal_note (ถ้ามีกรอก)
    if status_note:
        base_note = doc.internal_note or ""
        if base_note:
            base_note += "\n"
        base_note += f"[{new_status}] {status_note}"
        doc.internal_note = base_note

    db.session.commit()
    flash("บันทึกสถานะใบส่งสินค้าแล้ว", "success")
    return redirect(url_for("deliveries.view_doc", did=doc.id))



@bp_deliveries.route("/<int:did>/reschedule", methods=["POST"])
@permission_required("transport.manage")
def reschedule(did):
    doc = DeliveryDoc.query.get_or_404(did)

    # อ่านวันที่ใหม่จากฟอร์ม
    date_str = (request.form.get("new_delivery_date") or "").strip()
    if not date_str:
        flash("กรุณาเลือกวันที่จัดส่งใหม่", "warning")
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    try:
        new_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        flash("รูปแบบวันที่จัดส่งไม่ถูกต้อง", "danger")
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    # อัปเดตวันที่จัดส่งใหม่
    doc.delivery_date = new_date

    # ✨ สำคัญ: เปลี่ยนสถานะกลับเป็นรอจัดส่ง
    doc.status = "PENDING"

    # ล้างเหตุผลยกเลิกเดิมออก (ใช้ชื่อฟิลด์ที่มีจริง)
    if hasattr(doc, "cancel_reason_code"):
        doc.cancel_reason_code = None
    if hasattr(doc, "cancel_note"):
        doc.cancel_note = None

    db.session.commit()
    flash("บันทึกวันนัดจัดส่งใหม่แล้ว", "success")
    return redirect(url_for("deliveries.view_doc", did=doc.id))




@bp_deliveries.route("/<int:did>/print")
@require_perm("transport.access")
def print_doc(did):
    # โหลดใบส่ง + รายการ + รถ + คนขับ
    doc = (
        DeliveryDoc.query
        .options(
            joinedload(DeliveryDoc.items),
            joinedload(DeliveryDoc.vehicle),
            joinedload(DeliveryDoc.driver),
        )
        .get_or_404(did)
    )

    quote = None       # ใบเสนอราคา
    claim = None       # ใบเคลม (ถ้ามี)
    original_dl = None # ใบส่งสินค้าเดิมจาก QU

    src_type = (doc.source_type or "").upper()

    if src_type == "QUOTATION":
        # DL ปกติ สร้างมาจากใบเสนอราคา
        quote = (
            SalesDoc.query
            .options(joinedload(SalesDoc.customer))
            .filter(SalesDoc.id == doc.source_id)
            .first()
        )

    elif src_type == "CLAIM":
        # DL เคลม: doc.source_id คือ claim.id
        claim = (
            Claim.query
            .options(
                joinedload(Claim.customer),
                joinedload(Claim.quote),
            )
            .filter(Claim.id == doc.source_id)
            .first()
        )

        if claim and claim.quote:
            # ใช้ใบ QU ต้นทางของเคลม
            quote = claim.quote

            # หาใบส่งสินค้าเดิมที่สร้างจาก QU นี้ (น่าจะมีแค่ใบเดียว)
            original_dl = (
                DeliveryDoc.query
                .filter_by(source_type="QUOTATION", source_id=claim.quote_id)
                .order_by(DeliveryDoc.id.asc())
                .first()
            )

    return render_template(
        "deliveries/print.html",
        doc=doc,
        quote=quote,
        claim=claim,
        original_dl=original_dl,
    )




# ---------- DELIVERY BLUEPRINT (สตับใช้งานได้ทันที) ----------





def permission_required(code):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            # ยังไม่ล็อกอิน → เด้งไปหน้า login
            if not current_user.is_authenticated:
                return login_manager.unauthorized()

            # ----- ให้ admin เป็น superuser ใช้ได้ทุกเมนู -----
            if getattr(current_user, "username", None) == "admin":
                return f(*args, **kwargs)

            # ตรวจสิทธิ์ปกติสำหรับ user อื่น
            if not user_has_perm(current_user, code):
                abort(403)

            return f(*args, **kwargs)
        return wrapped
    return decorator



@permission_required("transport.view")
def list_():
    flash("(สตับ) หน้ารายการใบส่งสินค้า — ยังไม่ได้ทำ UI list จริง", "info")
    return redirect(url_for("dashboard"))


@permission_required("transport.manage")
def create_from_quote(qid):
    # TODO: สร้างเอกสารขนส่งจากใบเสนอราคา qid
    flash(f"(สตับ) สร้างใบส่งสินค้าจาก QU #{qid} แล้ว (จำลอง)", "success")
    return redirect(url_for("qu_view", qid=qid))


@permission_required("transport.manage")
def create_from_claim(claim_id):
    # TODO: สร้างเอกสารขนส่งเคลมจากใบเคลม claim_id
    flash(f"(สตับ) สร้างใบส่งสินค้าเคลมจากเคลม #{claim_id} แล้ว (จำลอง)", "success")
    return redirect(url_for("claim_view", claim_id=claim_id))


@permission_required("transport.view")
def view(did):
    flash(f"(สตับ) เปิดใบส่งสินค้า DID={did} (ยังไม่มีหน้าจอจริง)", "info")
    return redirect(url_for("dashboard"))


# ==== Transport permissions seeding =========================================
def seed_transport_perms():
    """สร้างสิทธิ์งานขนส่ง + ผูกให้ role admin/supervisor (idempotent)"""
    # ---- helpers ----
    def _add_perm(code: str, name: str):
        p = Permission.query.filter_by(code=code).first()
        if not p:
            p = Permission(code=code, name=name)
            db.session.add(p)
        return p

    def _ensure_role(code: str, name: str | None = None):
        r = Role.query.filter_by(code=code).first()
        if not r:
            r = Role(code=code, name=name or code.title())
            db.session.add(r)
            db.session.flush()
        return r

    def _grant(role_code: str, perm_code: str):
        r = _ensure_role(role_code)
        p = Permission.query.filter_by(code=perm_code).first()
        if not p:
            return
        link = RolePermission.query.filter_by(role_id=r.id, perm_id=p.id).first()
        if not link:
            db.session.add(RolePermission(role_id=r.id, perm_id=p.id))

    # ---- seed perms & grants ----
    _add_perm("transport.view",   "ดูเมนู/รายการงานขนส่ง")
    _add_perm("transport.manage", "สร้าง/แก้ไขใบส่งสินค้าและจัดสายรถ")
    db.session.flush()

    for rc in ["admin", "supervisor"]:
        _ensure_role(rc)                  # สร้าง role code = 'admin'/'supervisor' ถ้ายังไม่มี
        _grant(rc, "transport.view")
        _grant(rc, "transport.manage")

    db.session.commit()

# ============================================================================



# ================== DELIVERY / TRANSPORT BLUEPRINT (PLACEHOLDER) ==================




# เมนูรายการใบส่งสินค้า

@permission_required("transport.view")
def list_():
    return render_template_string("""
    {% extends "base.html" %}{% block content %}
    <div class="container py-3">
      <h1 class="h5">รายการใบส่งสินค้า (placeholder)</h1>
      <p class="text-muted">หน้านี้เอาไว้ทดสอบเมนูก่อน เดี๋ยวค่อยทำตารางจริง</p>
    </div>
    {% endblock %}
    """)

# สร้างใบส่งสินค้า (ปกติ)

@permission_required("transport.manage")
def new_normal():
    return render_template_string("""
    {% extends "base.html" %}{% block content %}
    <div class="container py-3"><h1 class="h5">สร้างใบส่งสินค้า (ปกติ)</h1></div>
    {% endblock %}
    """)

# สร้างใบส่งสินค้าเคลม

@permission_required("transport.manage")
def new_claim():
    return render_template_string("""
    {% extends "base.html" %}{% block content %}
    <div class="container py-3"><h1 class="h5">ใบส่งสินค้าเคลม</h1></div>
    {% endblock %}
    """)

# จัดสายรถ / วางแผน

@permission_required("transport.manage")
def plan():
    return render_template_string("""
    {% extends "base.html" %}{% block content %}
    <div class="container py-3"><h1 class="h5">จัดสายรถ / วางแผน</h1></div>
    {% endblock %}
    """)

# รถขนส่ง

@permission_required("transport.manage")
def vehicles():
    # ตัวอย่างข้อมูล mock ให้หน้าไม่โล่ง (ภายหลังเปลี่ยนเป็น query จาก DB ได้)
    rows = [
        {"code":"TRK-01","plate":"1กก-1234 กทม","type":"กระบะ","capacity":"1.5 ตัน","status":"พร้อมใช้งาน"},
        {"code":"VAN-02","plate":"2ขข-5678 ปท.","type":"ตู้แห้ง","capacity":"12 คิว","status":"ใช้งานอยู่"},
    ]
    return render_template_string("""
    {% extends "base.html" %}
    {% block title %}รถขนส่ง{% endblock %}
    {% block content %}
    <div class="container py-3">
      <div class="d-flex justify-content-between align-items-center mb-3">
        <h1 class="h5 m-0"><i class="bi bi-truck me-2"></i>รถขนส่ง</h1>
        <a href="{{ url_for('delivery.vehicles_new') }}" class="btn btn-primary btn-sm">
          <i class="bi bi-plus-circle me-1"></i> เพิ่มรถขนส่ง
        </a>
      </div>

      <div class="card border-0 shadow-sm">
        <div class="table-responsive">
          <table class="table table-hover align-middle mb-0">
            <thead class="table-light">
              <tr>
                <th style="min-width:120px">รหัสรถ</th>
                <th style="min-width:160px">ทะเบียน</th>
                <th>ประเภท</th>
                <th style="min-width:120px">ความจุ</th>
                <th style="min-width:140px">สถานะ</th>
                <th class="text-end" style="width:120px"></th>
              </tr>
            </thead>
            <tbody>
              {% for r in rows %}
              <tr>
                <td class="fw-semibold">{{ r.code }}</td>
                <td>{{ r.plate }}</td>
                <td>{{ r.type }}</td>
                <td>{{ r.capacity }}</td>
                <td>
                  <span class="badge rounded-pill text-bg-success" 
                        style="--bs-badge-font-size:.78rem">{{ r.status }}</span>
                </td>
                <td class="text-end">
                  <div class="btn-group">
                    <a class="btn btn-sm btn-outline-primary" href="#"><i class="bi bi-pencil"></i> แก้ไข</a>
                    <a class="btn btn-sm btn-outline-danger" href="#"><i class="bi bi-trash"></i></a>
                  </div>
                </td>
              </tr>
              {% else %}
              <tr>
                <td colspan="6" class="text-center py-5 text-muted">
                  ยังไม่มีรถขนส่ง — กด “เพิ่มรถขนส่ง”
                </td>
              </tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
    </div>
    {% endblock %}
    """, rows=rows)


@permission_required("transport.manage")
def vehicles_new():
    # แบบฟอร์มตัวอย่าง รอเชื่อม DB จริง
    return render_template_string("""
    {% extends "base.html" %}{% block title %}เพิ่มรถขนส่ง{% endblock %}
    {% block content %}
    <div class="container py-3" style="max-width:720px">
      <h1 class="h5 mb-3"><i class="bi bi-plus-circle me-2"></i>เพิ่มรถขนส่ง</h1>
      <div class="card border-0 shadow-sm">
        <div class="card-body row g-3">
          <div class="col-md-4"><label class="form-label">รหัสรถ</label><input class="form-control" placeholder="เช่น TRK-01"></div>
          <div class="col-md-4"><label class="form-label">ทะเบียน</label><input class="form-control" placeholder="เช่น 1กก-1234 กทม"></div>
          <div class="col-md-4">
            <label class="form-label">ประเภท</label>
            <select class="form-select"><option>กระบะ</option><option>ตู้แห้ง</option><option>กระบะ 4 ประตู</option></select>
          </div>
          <div class="col-md-6"><label class="form-label">ความจุ</label><input class="form-control" placeholder="เช่น 1.5 ตัน / 12 คิว"></div>
          <div class="col-md-6">
            <label class="form-label">สถานะ</label>
            <select class="form-select"><option selected>พร้อมใช้งาน</option><option>ซ่อมบำรุง</option><option>ใช้งานอยู่</option></select>
          </div>
        </div>
        <div class="card-footer d-flex justify-content-between">
          <a href="{{ url_for('delivery.vehicles') }}" class="btn btn-outline-secondary">กลับ</a>
          <button class="btn btn-primary" disabled>บันทึก (ตัวอย่าง)</button>
        </div>
      </div>
    </div>
    {% endblock %}
    """)

# คนขับ

@permission_required("transport.manage")
def drivers():
    rows = [
        {"code":"DRV-01","name":"สมชาย พันธ์ดี","tel":"081-234-5678","license":"ชำนาญ 6 ล้อ"},
        {"code":"DRV-02","name":"วิชัย เดชาวุฒิ","tel":"089-222-1111","license":"ชำนาญ รถตู้"},
    ]
    return render_template_string("""
    {% extends "base.html" %}{% block title %}คนขับ{% endblock %}
    {% block content %}
    <div class="container py-3">
      <div class="d-flex justify-content-between align-items-center mb-3">
        <h1 class="h5 m-0"><i class="bi bi-person-vcard me-2"></i>คนขับ</h1>
        <a href="#" class="btn btn-primary btn-sm"><i class="bi bi-plus-circle me-1"></i> เพิ่มคนขับ</a>
      </div>
      <div class="card border-0 shadow-sm">
        <div class="table-responsive">
          <table class="table table-hover align-middle mb-0">
            <thead class="table-light">
              <tr><th>รหัส</th><th>ชื่อ</th><th>โทร</th><th>ความชำนาญ</th><th class="text-end" style="width:120px"></th></tr>
            </thead>
            <tbody>
              {% for r in rows %}
              <tr>
                <td class="fw-semibold">{{ r.code }}</td>
                <td>{{ r.name }}</td>
                <td>{{ r.tel }}</td>
                <td>{{ r.license }}</td>
                <td class="text-end">
                  <div class="btn-group">
                    <a class="btn btn-sm btn-outline-primary" href="#"><i class="bi bi-pencil"></i> แก้ไข</a>
                    <a class="btn btn-sm btn-outline-danger" href="#"><i class="bi bi-trash"></i></a>
                  </div>
                </td>
              </tr>
              {% else %}
              <tr><td colspan="5" class="text-center py-5 text-muted">ยังไม่มีข้อมูลคนขับ</td></tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
    </div>
    {% endblock %}
    """, rows=rows)

# เส้นทาง / โซน

@permission_required("transport.manage")
def zones():
    rows = [
        {"code":"ZN-A","name":"โซน A (ในเมือง)","desc":"รัศมี 10 กม.","stops":12},
        {"code":"ZN-B","name":"โซน B (ตะวันออก)","desc":"รามอินทรา–มีนบุรี","stops":8},
    ]
    return render_template_string("""
    {% extends "base.html" %}{% block title %}เส้นทาง / โซน{% endblock %}
    {% block content %}
    <div class="container py-3">
      <div class="d-flex justify-content-between align-items-center mb-3">
        <h1 class="h5 m-0"><i class="bi bi-geo-alt me-2"></i>เส้นทาง / โซน</h1>
        <a href="#" class="btn btn-primary btn-sm"><i class="bi bi-plus-circle me-1"></i> เพิ่มโซน</a>
      </div>
      <div class="card border-0 shadow-sm">
        <div class="table-responsive">
          <table class="table table-hover align-middle mb-0">
            <thead class="table-light">
              <tr><th>รหัส</th><th>ชื่อโซน</th><th>รายละเอียด</th><th class="text-center" style="width:120px">จำนวนจุด</th><th class="text-end" style="width:120px"></th></tr>
            </thead>
            <tbody>
              {% for r in rows %}
              <tr>
                <td class="fw-semibold">{{ r.code }}</td>
                <td>{{ r.name }}</td>
                <td class="text-muted">{{ r.desc }}</td>
                <td class="text-center">{{ r.stops }}</td>
                <td class="text-end">
                  <div class="btn-group">
                    <a class="btn btn-sm btn-outline-primary" href="#"><i class="bi bi-pencil"></i> แก้ไข</a>
                    <a class="btn btn-sm btn-outline-danger" href="#"><i class="bi bi-trash"></i></a>
                  </div>
                </td>
              </tr>
              {% else %}
              <tr><td colspan="5" class="text-center py-5 text-muted">ยังไม่มีข้อมูลโซน</td></tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
    </div>
    {% endblock %}
    """, rows=rows)


# ---------- Delivery: create from QU / Claim (ต้องประกาศก่อน register blueprint) ----------
@bp_deliveries.route("/create-from-quotation/<int:qid>", methods=["GET", "POST"])
@require_perm("transport.access")
def create_from_quotation(qid):
    # โหลดใบเสนอราคา (QU) พร้อมลูกค้าและรายการ
    quote = (
        SalesDoc.query
        .options(
            joinedload(SalesDoc.customer),
            joinedload(SalesDoc.items),
        )
        .filter(SalesDoc.id == qid, SalesDoc.doc_type == "QU")
        .first_or_404()
    )

    # ถ้ามีใบส่งจากใบนี้แล้ว และเป็นการเปิดหน้า GET ปกติ -> เด้งไปดูใบส่ง
    existing = DeliveryDoc.query.filter_by(source_type="QUOTATION", source_id=qid).first()
    if existing and request.method == "GET":
        flash("มีใบส่งสินค้าถูกสร้างจากใบเสนอราคานี้แล้ว", "info")
        return redirect(url_for("deliveries.view_doc", did=existing.id))

    # ค่า default จากใบเสนอราคา
    default_name  = quote.customer.name if quote.customer else ""
    default_phone = quote.customer.phone if quote.customer and getattr(quote.customer, "phone", None) else ""
    default_addr  = quote.customer.address if quote.customer and getattr(quote.customer, "address", None) else ""
    default_delivery_date = quote.doc_date if hasattr(quote, "doc_date") else date.today()

    if request.method == "POST":
        ship_to_name = (request.form.get("ship_to_name") or "").strip()
        ship_to_phone = (request.form.get("ship_to_phone") or "").strip()
        ship_to_address = (request.form.get("ship_to_address") or "").strip()
        ship_to_note = (request.form.get("ship_to_note") or "").strip()

        delivery_date_str = (request.form.get("delivery_date") or "").strip()
        delivery_date = None
        if delivery_date_str:
            try:
                delivery_date = datetime.strptime(delivery_date_str, "%Y-%m-%d").date()
            except ValueError:
                flash("รูปแบบวันที่จัดส่งไม่ถูกต้อง", "warning")
                return redirect(request.url)
        else:
            delivery_date = default_delivery_date

        number = _gen_running("DL", DeliveryDoc)
        doc = DeliveryDoc(
            number=number,
            d_type=DeliveryType.DL,
            status=DeliveryStatus.PENDING,
            source_type="QUOTATION",
            source_id=qid,
            ship_to_name=ship_to_name or default_name,
            ship_to_phone=ship_to_phone or default_phone,
            ship_to_address=ship_to_address or default_addr,
            ship_to_note=ship_to_note or None,
            delivery_date=delivery_date,   # 👈 ใช้ delivery_date ไม่ใช่ date
        )
        db.session.add(doc)
        db.session.flush()

        # คัดลอกรายการจากใบเสนอราคา มาเป็นรายการในใบส่งสินค้า
        for it in quote.items:
            d_item = DeliveryItem(
                doc_id=doc.id,
                source_item_id=it.id,
                product_name=it.name,
                qty=it.qty,
                unit="ชิ้น",
                note=None,
            )
            db.session.add(d_item)

        db.session.commit()
        flash("สร้างใบส่งสินค้าเรียบร้อย", "success")
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    return render_template(
        "deliveries/create_from_source.html",
        quotation=quote,
        items=quote.items,
        is_claim=False,
        default_delivery_date=default_delivery_date,
        back_url=url_for("qu_list"),
        source_type="QUOTATION",
        source_id=qid,
        d_type="DL",
    )



def create_from_qu(qid):
    """
    Wrapper endpoint สำหรับปุ่ม 'สร้างใบส่งสินค้า' จากใบเสนอราคา
    - ถ้ามี endpoint delivery.new อยู่ จะ redirect ไปพร้อม query string
    - ถ้าไม่มี (ยังไม่ได้ทำหน้า new) จะ fallback เป็น path ตรง
    """
    try:
        return redirect(url_for("delivery.new", from_="quote", qid=qid))
    except BuildError:
        return redirect(f"/delivery/new?from=quote&qid={qid}")


def create_from_claim(claim_id):
    """
    Wrapper endpoint สำหรับปุ่ม 'สร้างใบส่งสินค้าเคลม' จากใบเคลม
    - ถ้ามี endpoint delivery.new_claim อยู่ จะ redirect ไปพร้อม query string
    - ถ้าไม่มี (ยังไม่ได้ทำหน้า new-claim) จะ fallback เป็น path ตรง
    """
    try:
        return redirect(url_for("delivery.new_claim", claim_id=claim_id))
    except BuildError:
        return redirect(f"/delivery/new-claim?claim_id={claim_id}")
# -------------------------------------------------------------------------


# ลงทะเบียน blueprint


# debug: พิมพ์รายการเส้นทางที่เรามี
try:
    routes = [str(r) for r in app.url_map.iter_rules() if r.endpoint.startswith("delivery.")]
    print("DELIVERY ROUTES:", routes)
except Exception:
    pass
# ================================================================================ 


@bp_deliveries.route("/schedule")
@login_required
@permission_required("transport.manage")
def schedule_view():
    # วันที่เป้าหมาย (default = วันนี้)
    date_str = request.args.get("date")
    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()

    # รถที่ใช้งานอยู่
    vehicles = (
        DeliveryVehicle.query
        .filter(DeliveryVehicle.is_active == True)
        .order_by(DeliveryVehicle.name.asc())
        .all()
    )

    # คนขับที่ยัง active
    drivers = (
        Driver.query
        .filter(Driver.is_active == True)
        .order_by(Driver.full_name.asc())
        .all()
    )

    # ดึงใบส่งสินค้าที่นัดส่งในวันนั้น (ไม่เอาใบที่ยกเลิก)
    docs = (
    DeliveryDoc.query
    .options(
        joinedload(DeliveryDoc.vehicle),
        joinedload(DeliveryDoc.driver),
    )
    .filter(
        DeliveryDoc.d_type.in_([DeliveryType.DL, DeliveryType.DLC]),
        DeliveryDoc.status != DeliveryStatus.CANCELLED,
        DeliveryDoc.delivery_date == target_date,   # 👈 ใช้อันนี้
    )
    .order_by(DeliveryDoc.number.asc())
    .all()
)

    # แยกเป็นใบที่ยังไม่จัดรถ กับใบที่มีรถแล้ว
    docs_by_vehicle = {}
    unassigned_docs = []

    for d in docs:
        if d.vehicle_id:
            docs_by_vehicle.setdefault(d.vehicle_id, []).append(d)
        else:
            unassigned_docs.append(d)

    return render_template(
        "deliveries/schedule.html",
        target_date=target_date,
        vehicles=vehicles,
        drivers=drivers,
        unassigned_docs=unassigned_docs,
        docs_by_vehicle=docs_by_vehicle,
    )



@bp_deliveries.route("/<int:did>/assign", methods=["POST"])
@require_perm("transport.access")  # ถ้าอยากให้เฉพาะ role บางคนแก้ได้ เปลี่ยน permission ตามที่นายใช้
def assign_delivery(did):
    """อัปเดตรถ / คนขับ / วันที่จัดส่ง ของใบส่ง 1 ใบ"""
    doc = DeliveryDoc.query.get_or_404(did)

    vehicle_id = request.form.get("vehicle_id") or None
    driver_id = request.form.get("driver_id") or None
    date_str = request.form.get("delivery_date") or ""
    next_url = request.form.get("next") or url_for("deliveries.schedule_view")

    doc.vehicle_id = int(vehicle_id) if vehicle_id else None
    doc.driver_id = int(driver_id) if driver_id else None

    if date_str:
        try:
            doc.delivery_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    db.session.commit()
    flash("อัปเดตสายรถเรียบร้อย", "success")
    return redirect(next_url)



@bp_deliveries.route("/<int:did>/upload-before", methods=["POST"])
@require_perm("transport.manage")
def upload_before_photos(did):
    doc = DeliveryDoc.query.get_or_404(did)

    files = request.files.getlist("photos")
    existing = DeliveryPhoto.query.filter_by(doc_id=doc.id, kind="BEFORE").count()
    new_count = sum(1 for f in files if getattr(f, "filename", None))
    total = existing + new_count

    if total > 10:
        flash(f"รูปก่อนส่งได้ไม่เกิน 10 รูป (มีอยู่แล้ว {existing} รูป)", "warning")
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    saved = _save_delivery_photos(files, doc, "BEFORE")
    db.session.commit()

    total_after = existing + saved
    if total_after < 3:
        flash(f"อัปโหลดรูปก่อนส่งแล้ว (ตอนนี้มี {total_after} รูป) — ยังไม่ครบ 3 รูป ระบบจะยังไม่ให้เปลี่ยนเป็นกำลังจัดส่ง", "warning")
    else:
        flash(f"อัปโหลดรูปก่อนส่งแล้ว ({total_after} รูป)", "success")

    return redirect(url_for("deliveries.view_doc", did=doc.id))


@bp_deliveries.route("/<int:did>/upload-after", methods=["POST"])
@require_perm("transport.manage")
def upload_after_photos(did):
    doc = DeliveryDoc.query.get_or_404(did)

    files = request.files.getlist("photos")
    existing = DeliveryPhoto.query.filter_by(doc_id=doc.id, kind="AFTER").count()
    new_count = sum(1 for f in files if getattr(f, "filename", None))
    total = existing + new_count

    if total > 10:
        flash(f"รูปหลังส่งได้ไม่เกิน 10 รูป (มีอยู่แล้ว {existing} รูป)", "warning")
        return redirect(url_for("deliveries.view_doc", did=doc.id))

    saved = _save_delivery_photos(files, doc, "AFTER")
    db.session.commit()

    total_after = existing + saved
    if total_after < 3:
        flash(f"อัปโหลดรูปหลังส่งแล้ว (ตอนนี้มี {total_after} รูป) — ยังไม่ครบ 3 รูป ระบบจะยังไม่ให้เปลี่ยนเป็นจัดส่งสำเร็จ", "warning")
    else:
        flash(f"อัปโหลดรูปหลังส่งแล้ว ({total_after} รูป)", "success")

    return redirect(url_for("deliveries.view_doc", did=doc.id))


# ================== GIFT / LOYALTY ROUTES ==================

@app.route("/gifts")
@login_required
@permission_required("gifts.view")
def gifts_index():
    """
    หน้าแรกเมนูของขวัญ: สรุปตัวเลขรวม + ลิสต์แคมเปญทั้งหมด
    """
    campaigns = (
        GiftCampaign.query
        .order_by(GiftCampaign.period_start.desc())
        .all()
    )

    # สรุปตัวเลขรวมทั้งหมดจากทุกแคมเปญ
    total_qualified = GiftResult.query.count()
    total_given = GiftResult.query.filter_by(status="GIVEN").count()

    # หาว่าลูกค้าคนไหนผ่านเกณฑ์บ่อยที่สุด (กี่ครั้ง)
    top_row = (
        db.session.query(
            Customer.name.label("customer_name"),
            func.count(GiftResult.id).label("times"),
        )
        .join(GiftResult, GiftResult.customer_id == Customer.id)
        .group_by(Customer.id)
        .order_by(func.count(GiftResult.id).desc())
        .first()
    )
    top_times = top_row.times if top_row else 0
    top_customer_name = top_row.customer_name if top_row else None

    stats = {
        "total_qualified": total_qualified,
        "total_given": total_given,
        "top_times": top_times,
        "top_customer_name": top_customer_name,
    }

    return render_template(
        "gifts/index.html",
        campaigns=campaigns,
        stats=stats,
    )


@app.route("/gifts/new", methods=["GET", "POST"])
@login_required
@permission_required("gifts.manage")
def gifts_new():
    """
    สร้างแคมเปญของขวัญใหม่ + กำหนด tier (เกณฑ์) เบื้องต้น
    """
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        description = (request.form.get("description") or "").strip()
        period_start_str = request.form.get("period_start")
        period_end_str = request.form.get("period_end")
        cycle_months = int(request.form.get("cycle_months") or 4)
        anchor_month = int(request.form.get("anchor_month") or 1)

        if not name or not period_start_str or not period_end_str:
            flash("กรุณากรอกชื่อแคมเปญ และช่วงวันที่ให้ครบถ้วน", "danger")
            return redirect(url_for("gifts_new"))

        try:
            period_start = datetime.strptime(period_start_str, "%Y-%m-%d").date()
            period_end = datetime.strptime(period_end_str, "%Y-%m-%d").date()
        except ValueError:
            flash("รูปแบบวันที่ไม่ถูกต้อง", "danger")
            return redirect(url_for("gifts_new"))

        if period_end < period_start:
            flash("วันสิ้นสุดต้องไม่น้อยกว่าวันเริ่มต้น", "danger")
            return redirect(url_for("gifts_new"))

        campaign = GiftCampaign(
            name=name,
            description=description,
            period_start=period_start,
            period_end=period_end,
            cycle_months=cycle_months,
            anchor_month=anchor_month,
        )
        db.session.add(campaign)
        db.session.flush()  # ให้ได้ campaign.id ก่อน

                # อ่าน tier จากฟอร์ม (รองรับหลายรูปแบบของชื่อ field)
        # - tier_code_1 / tier_name_1 / tier_min_amount_1  (ตาม template)
        # - tier1_code / tier1_name / tier1_min           (ของเก่า)
        tier_indexes = set()
        for k in request.form.keys():
            m = re.match(r"tier_(?:code|name|min_amount)_(\d+)$", k)
            if m:
                tier_indexes.add(int(m.group(1)))
                continue
            m2 = re.match(r"tier(\d+)_(?:code|name|min|min_amount)$", k)
            if m2:
                tier_indexes.add(int(m2.group(1)))

        # เผื่อกรณี template ส่งมาเป็นแถวคงที่ A/B/C แต่ไม่มี key ตาม regex
        if not tier_indexes:
            tier_indexes = {1, 2, 3}

        saved_tiers = 0
        for idx in sorted(tier_indexes)[:20]:
            code = (request.form.get(f"tier_code_{idx}") or request.form.get(f"tier{idx}_code") or "").strip()
            tname = (request.form.get(f"tier_name_{idx}") or request.form.get(f"tier{idx}_name") or "").strip()
            min_amount_str = (
                request.form.get(f"tier_min_amount_{idx}")
                or request.form.get(f"tier{idx}_min_amount")
                or request.form.get(f"tier{idx}_min")
                or ""
            ).strip()

            # ถ้าแถวว่าง -> ข้าม
            if not code and not tname and not min_amount_str:
                continue

            # ต้องมีครบ
            if not code or not tname or min_amount_str == "":
                continue

            try:
                min_amount = Decimal(min_amount_str.replace(",", ""))
            except Exception:
                continue

            tier = GiftTier(
                campaign_id=campaign.id,
                code=code,
                name=tname,
                min_amount=min_amount,
                sort_order=idx,
            )
            db.session.add(tier)
            saved_tiers += 1

        db.session.commit()
        flash(f"สร้างแคมเปญของขวัญเรียบร้อย (บันทึก Tier = {saved_tiers})", "success")
        return redirect(url_for("gifts_campaign_view", cid=campaign.id))

    return render_template("gifts/new.html")


@app.route("/gifts/<int:cid>")
@login_required
@permission_required("gifts.view")
def gifts_campaign_view(cid):
    """
    ดูรายละเอียดแคมเปญ + รายชื่อลูกค้าที่ผ่านเกณฑ์ในแคมเปญนั้น
    """
    campaign = GiftCampaign.query.options(
        joinedload(GiftCampaign.tiers),
        joinedload(GiftCampaign.results).joinedload(GiftResult.customer),
    ).get_or_404(cid)

    # นับจำนวนลูกค้าที่ผ่านเกณฑ์ และจำนวนที่ให้ของขวัญแล้ว
    total_qualified = len(campaign.results)
    total_given = sum(1 for r in campaign.results if r.status == "GIVEN")

    return render_template(
        "gifts/campaign_view.html",
        campaign=campaign,
        total_qualified=total_qualified,
        total_given=total_given,
    )


@app.route("/gifts/<int:cid>/recalc", methods=["POST"])
@login_required
@permission_required("gifts.manage")
def gifts_campaign_recalc(cid):
    """
    กดปุ่มคำนวณลูกค้าที่ผ่านเกณฑ์ใหม่สำหรับแคมเปญ
    """
    campaign = GiftCampaign.query.get_or_404(cid)
    recalc_gift_results(campaign)
    flash("คำนวณลูกค้าที่ผ่านเกณฑ์ในแคมเปญนี้เรียบร้อยแล้ว", "success")
    return redirect(url_for("gifts_campaign_view", cid=cid))


@app.route("/gifts/result/<int:rid>/toggle", methods=["POST"])
@login_required
@permission_required("gifts.manage")
def gifts_toggle_result(rid):
    """
    สลับสถานะ ให้ของขวัญแล้ว / ยังไม่ให้ ให้กับลูกค้ารายหนึ่งในแคมเปญ
    """
    gr = GiftResult.query.get_or_404(rid)

    if gr.status == "GIVEN":
        gr.status = "PENDING"
        gr.given_at = None            # เคลียร์วันที่ให้ของขวัญ
        msg = "เปลี่ยนสถานะเป็น 'ยังไม่ให้ของขวัญ' แล้ว"
    else:
        gr.status = "GIVEN"
        gr.given_at = datetime.utcnow()  # บันทึกเวลาที่ให้ของขวัญ
        msg = "บันทึกว่า 'ให้ของขวัญแล้ว' เรียบร้อย"

    db.session.commit()
    flash(msg, "success")
    return redirect(url_for("gifts_campaign_view", cid=gr.campaign_id))


# ================== RETURN DOCS (ใบคืนสินค้า) ==================

@app.route("/sales/returns")
@login_required
@permission_required("sales.view")
def rn_list():
    """
    ลิสต์ใบคืนสินค้าทั้งหมด
    """
    docs = (
        ReturnDoc.query
        .order_by(ReturnDoc.date.desc(), ReturnDoc.number.desc())
        .all()
    )
    return render_template("sales/rn_list.html", docs=docs)


@app.route("/sales/returns/new")
@login_required
@permission_required("sales.manage")
def rn_new():
    """
    หน้าสร้างใบคืนสินค้า (ตอนนี้ทำเป็นสเต็ปถัดไป)
    ตอนนี้ให้ redirect กลับลิสต์ไปก่อน จะได้ไม่ขึ้น 404
    """
    flash("หน้าสร้างใบคืนสินค้า กำลังอยู่ระหว่างพัฒนา (Step ถัดไป)", "info")
    return redirect(url_for("rn_list"))


@app.route("/sales/returns/<int:rid>")
@login_required
@permission_required("sales.view")
def rn_view(rid):
    """
    ดูรายละเอียดใบคืนสินค้าแบบง่าย ๆ (เดี๋ยวค่อยทำหน้าสวยใน step ถัดไป)
    """
    doc = (
        ReturnDoc.query
        .options(
            joinedload(ReturnDoc.customer),
            joinedload(ReturnDoc.quote),
        joinedload(ReturnDoc.booking),
            joinedload(ReturnDoc.items),
        )
        .get_or_404(rid)
    )
    return render_template("sales/rn_view.html", doc=doc)


@app.route("/sales/returns/<int:rid>/print")
@login_required
@permission_required("sales.view")
def rn_print(rid):
    """
    ปริ้นใบคืนสินค้า (เดี๋ยวเราค่อยทำ template พิมพ์ทีหลัง)
    ตอนนี้ให้ใช้ template เปล่า ๆ ไปก่อน
    """
    doc = (
        ReturnDoc.query
        .options(
            joinedload(ReturnDoc.customer),
            joinedload(ReturnDoc.quote),
        joinedload(ReturnDoc.booking),
            joinedload(ReturnDoc.items),
        )
        .get_or_404(rid)
    )
    return render_template("sales/rn_print.html", doc=doc)

# ===== ใบคืนสินค้า (Returns) =====



@app.route("/returns/from-quote/<int:qid>")
@login_required
@permission_required("sales.manage")
def returns_from_quote(qid):
    """
    STEP ต่อไป: ใช้สร้างใบคืนสินค้าจริง ๆ จากใบเสนอราคา
    ตอนนี้ยังเป็น stub อยู่ แค่ redirect กลับไปที่หน้าเลือกลูกค้า
    """
    quote = SalesDoc.query.get_or_404(qid)
    flash("ฟังก์ชันสร้างใบคืนสินค้ายังไม่เปิดใช้งาน เดี๋ยวเราค่อยเติม logic ต่อ", "warning")
    return redirect(url_for("returns_new", customer_id=quote.customer_id))






# ลงทะเบียน blueprint (ถ้ายังไม่ได้)
app.register_blueprint(bp_repairs)
app.register_blueprint(bp_deliveries)
print("REPAIRS ROUTES:",
      [r.rule for r in app.url_map.iter_rules() if "repairs" in r.rule])
print("DELIVERY ROUTES:",
      [r.rule for r in app.url_map.iter_rules() if "deliveries" in r.rule])




# ==== seed default admin user ======================================
def seed_default_admin():
    from werkzeug.security import generate_password_hash

    admin = User.query.filter_by(username="admin").first()
    if admin:
        print("[seed] admin already exists")
        return

    admin = User(username="admin")

    # ถ้ามี role_code ให้ใช้เป็น admin
    if hasattr(User, "role_code"):
        admin.role_code = "admin"
    elif hasattr(User, "role"):
        admin.role = "admin"

    if hasattr(User, "full_name"):
        admin.full_name = "ผู้ดูแลระบบ"
    elif hasattr(User, "name"):
        admin.name = "ผู้ดูแลระบบ"

    if hasattr(User, "is_active"):
        admin.is_active = True

    admin.password_hash = generate_password_hash("admin123")

    db.session.add(admin)
    db.session.commit()
    print("[seed] created default admin user: admin / admin123")



# ==== run startup tasks (create tables + seed) =====================
def run_startup_tasks():
    """
    รันตอนแอปถูก import (เช่นบน Render)
    แต่จะข้าม create_all / seed ถ้าเป็นโหมด migration
    """
    from sqlalchemy.exc import OperationalError

    # 👉 ข้ามทันที ถ้าเป็นคำสั่ง flask db / alembic
    if os.environ.get("ST_SKIP_INIT") == "1" or _is_migration_command():
        print("[init] skip create_all/seed (migration mode)")
        return

    with app.app_context():
        # สร้างตารางทั้งหมด (เฉพาะตอนรันแอปจริง)
        try:
            db.create_all()
            print("[init] db.create_all completed")
        except OperationalError as e:
            print(f"[init] db.create_all failed: {e}")

        # seed ข้อมูลพื้นฐาน
        try:
            bootstrap()  # สร้างสิทธิ์/ตำแหน่ง/บริษัท/admin ฯลฯ
            print("[seed] bootstrap completed")
        except Exception as e:
            print(f"[seed] startup tasks failed: {e}")


# เรียกตอน import app ครั้งแรก (ทั้งตอน dev และบน Render)
run_startup_tasks()

@app.get("/sales/bookings")
@login_required
@permission_required("sales.view")
def bk_list():
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    qry = SalesDoc.query.options(joinedload(SalesDoc.customer)).filter_by(doc_type="BK")

    if start_d:
        qry = qry.filter(SalesDoc.date >= start_d)
    if end_d:
        qry = qry.filter(SalesDoc.date <= end_d)

    if q:
        like = f"%{q}%"
        qry = qry.join(Customer).filter(or_(SalesDoc.number.ilike(like), Customer.name.ilike(like)))

    docs = (
        qry
        .order_by(SalesDoc.date.desc(), SalesDoc.id.desc())
        .all()
    )
    return render_template("sales/bk_list.html", docs=docs, q=q, start=start, end=end)

@app.get("/sales/bookings/<int:doc_id>")
@login_required
@permission_required("sales.view")
def bk_view(doc_id):
    d: SalesDoc = (
        SalesDoc.query
        .options(
            joinedload(SalesDoc.customer),
            joinedload(SalesDoc.items),
        )
        .filter_by(doc_type="BK", id=doc_id)
        .first_or_404()
    )

    children = SalesDoc.query.filter_by(parent_id=d.id).all()
    children_map = { (c.doc_type or "").upper(): c for c in children }

    flow_ready = False
    try:
        flow_ready = _booking_flow_ready(d)
    except Exception:
        flow_ready = False

    return render_template(
        "sales/bk_view.html",
        d=d,
        children_map=children_map,
        flow_ready=flow_ready,
        page_title="ใบจอง"
    )






@app.route("/sales/bookings/<int:bid>/allocate", methods=["GET", "POST"])
@permission_required("sales.manage")
def bk_allocate(bid):
    """หน้าเลือกตัวอุปกรณ์จริงให้ใบจอง (BK)
    - QU: เก็บเป็นหมวดหมู่ (category_id / category_prefix)
    - BK: เลือก SKU อุปกรณ์จริงที่ READY เพื่อ "ล็อคของ" และเปลี่ยนเป็น RENTED
    """
    doc: SalesDoc = (
        SalesDoc.query
        .options(joinedload(SalesDoc.items), joinedload(SalesDoc.customer))
        .get_or_404(bid)
    )

    if (doc.doc_type or "").upper() != "BK":
        flash("เอกสารนี้ไม่ใช่ใบจอง (BK)", "warning")
        if "sales_doc_view" in current_app.view_functions:
            return redirect(url_for("sales_doc_view", doc_id=doc.id))
        return redirect(url_for("bk_view", doc_id=doc.id))

    import re as _re

    def _extract_prefix(title: str):
        # รองรับ [SPTE] หรือ [SPTE6901] อะไรที่อยู่ใน [] จะหยิบมา
        m = _re.search(r"\[([^\[\]]+?)\]", title or "")
        return m.group(1).strip() if m else ""

    def _normalize_cat_prefix(p: str) -> str:
        # เก็บเฉพาะ A-Z0-9 และทำให้เป็นตัวใหญ่
        p = (p or "").strip().upper()
        p = _re.sub(r"[^A-Z0-9]", "", p)
        return p

    def _resolve_category_for_item(it: SalesItem):
        """
        return dict:
          {
            'cat_id': int|None,
            'cat_prefix': str,    # Category.prefix_sku (เช่น SPTE/SPTA/E/A)
            'legacy_prefix': str  # it.category_prefix หรือ [] ในชื่อ
          }
        """
        cat_id = getattr(it, "category_id", None)
        legacy_prefix = (getattr(it, "category_prefix", "") or "").strip() or _extract_prefix(getattr(it, "name", ""))

        # 1) ถ้ามี category_id ใช้เลย (ดีที่สุด)
        if cat_id:
            c = Category.query.get(cat_id)
            return {
                "cat_id": cat_id,
                "cat_prefix": (c.prefix_sku or "").strip() if c else "",
                "legacy_prefix": legacy_prefix,
            }

        # 2) fallback: หา category จาก legacy_prefix (พยายาม normalize)
        lp = _normalize_cat_prefix(legacy_prefix)
        if lp:
            # เทียบตรง ๆ
            c = Category.query.filter_by(prefix_sku=lp).first()
            if not c:
                # เผื่อบางคนเก็บ prefix_sku เป็น "E" แต่ legacy เป็น "SPTE"
                if lp.startswith("SPT") and len(lp) >= 4:
                    c = Category.query.filter_by(prefix_sku=lp[3:]).first()
                elif len(lp) == 1:
                    c = Category.query.filter_by(prefix_sku="SPT" + lp).first()

            if c:
                return {"cat_id": c.id, "cat_prefix": (c.prefix_sku or "").strip(), "legacy_prefix": legacy_prefix}

        return {"cat_id": None, "cat_prefix": "", "legacy_prefix": legacy_prefix}

    # -------------------------
    # POST: save allocation
    # -------------------------
    if request.method == "POST":
        any_selected = False
        used_skus = set()

        # กันเลือก SKU ซ้ำในฟอร์มเดียวกัน
        for it in (doc.items or []):
            for sku in request.form.getlist(f"alloc_{it.id}"):
                sku = (sku or "").strip()
                if not sku:
                    continue
                if sku in used_skus:
                    flash(f"เลือกอุปกรณ์ซ้ำกันในฟอร์ม: {sku}", "danger")
                    return redirect(url_for("bk_allocate", bid=doc.id))
                used_skus.add(sku)

        # ทำทีละบรรทัด
        for it in (doc.items or []):
            chosen = [(x or "").strip() for x in request.form.getlist(f"alloc_{it.id}") if (x or "").strip()]
            qty_need = int(round(float(it.qty or 0))) if it.qty is not None else 0
            qty_need = max(qty_need, 1)

            if not chosen:
                # อนุญาตให้ยังไม่เลือกครบทุกบรรทัดได้
                continue

            any_selected = True

            if len(chosen) < qty_need:
                flash(f"รายการ '{it.name}' ต้องเลือกอย่างน้อย {qty_need} ตัว (ตอนนี้เลือก {len(chosen)} ตัว)", "danger")
                return redirect(url_for("bk_allocate", bid=doc.id))

            info = _resolve_category_for_item(it)
            cat_id = info["cat_id"]

            # ตรวจอุปกรณ์จริงว่าพร้อมให้เช่า + อยู่หมวดเดียวกัน
            eqs = Equipment.query.filter(Equipment.sku.in_(chosen)).all()
            eq_by_sku = {e.sku: e for e in eqs}

            for sku in chosen:
                eq = eq_by_sku.get(sku)
                if not eq:
                    flash(f"ไม่พบอุปกรณ์ SKU: {sku}", "danger")
                    return redirect(url_for("bk_allocate", bid=doc.id))
                if (eq.status or "").upper() != "READY":
                    flash(f"อุปกรณ์ {sku} ไม่ได้อยู่สถานะ READY (สถานะปัจจุบัน: {eq.status})", "danger")
                    return redirect(url_for("bk_allocate", bid=doc.id))
                if cat_id and eq.category_id != cat_id:
                    flash(f"อุปกรณ์ {sku} ไม่ตรงหมวดของรายการนี้", "danger")
                    return redirect(url_for("bk_allocate", bid=doc.id))

            # บันทึก SKU ลง BK item
            it.allocated_skus = ",".join(chosen)

            # sync หมวดให้ชัดเจน: ถ้า item ไม่มี category_id แต่ resolve ได้ -> ใส่กลับ
            if cat_id and not getattr(it, "category_id", None):
                it.category_id = cat_id

            # เก็บ category_prefix ให้เป็น prefix_sku ของ Category (เพื่อ UI/legacy)
            try:
                if cat_id:
                    c = Category.query.get(cat_id)
                    if c and c.prefix_sku:
                        it.category_prefix = (c.prefix_sku or "").strip()
            except Exception:
                pass

            # SYNC allocated_skus back to QU parent
            try:
                if doc.parent_id:
                    qu = SalesDoc.query.get(doc.parent_id)
                    if qu and (qu.doc_type or "").upper() == "QU" and getattr(it, "source_qu_item_id", None):
                        qu_it = SalesItem.query.get(it.source_qu_item_id)
                        if qu_it and qu_it.doc_id == qu.id:
                            qu_it.allocated_skus = it.allocated_skus
                            qu_it.category_id = it.category_id
                            qu_it.category_prefix = it.category_prefix
            except Exception:
                pass

            # ปล่อย reservation (ถ้ามี)
            try:
                _release_reservation_for_item(it.id)
            except Exception:
                pass

            # เปลี่ยนสถานะอุปกรณ์เป็น RENTED + log
            cust_name = doc.customer.name if getattr(doc, "customer", None) else ""
            for sku in chosen:
                eq = eq_by_sku[sku]
                eq.status = "RENTED"
                try:
                    _equip_log(eq, "RENT_OUT", f"จองจาก {doc.number} | ลูกค้า: {cust_name}")
                except Exception:
                    pass

        db.session.commit()
        flash("บันทึกการเลือกอุปกรณ์เรียบร้อยแล้ว" if any_selected else "ยังไม่ได้เลือกอุปกรณ์", "success" if any_selected else "info")
        return redirect(url_for("bk_view", doc_id=doc.id))

    # -------------------------
    # GET: build items_vm
    # -------------------------
    items_vm = []
    for it in (doc.items or []):
        info = _resolve_category_for_item(it)
        cat_id = info["cat_id"]
        qty_need = int(round(float(it.qty or 0))) if it.qty is not None else 0
        qty_need = max(qty_need, 1)

        q = Equipment.query.filter(Equipment.status == "READY")
        if cat_id:
            q = q.filter(Equipment.category_id == cat_id)

        eqs = q.order_by(Equipment.sku.asc()).limit(300).all()

        already = [s.strip() for s in (it.allocated_skus or "").split(",") if s.strip()]
        items_vm.append(
            {
                "item": it,
                "prefix": info["cat_prefix"] or info["legacy_prefix"],
                "qty_need": qty_need,
                "equipments": eqs,
                "already": already,
            }
        )

    return render_template("sales/bk_allocate.html", doc=doc, items_vm=items_vm)




@app.post("/sales/bookings/<int:bid>/approve")
@login_required
@permission_required("sales.manage")
def bk_approve(bid):
    d = (
        SalesDoc.query.options(
            joinedload(SalesDoc.items),
            joinedload(SalesDoc.customer),
        )
        .get_or_404(bid)
    )

    # กันกดซ้ำ
    if (d.status or "").upper() == "APPROVED":
        flash("ใบจองนี้อนุมัติแล้ว", "info")
        ct = SalesDoc.query.filter_by(parent_id=d.parent_id, doc_type='CT').first()
        if ct:
            return redirect(url_for('contract_view', cid=ct.id))
        return redirect(url_for('bk_view', doc_id=d.id))  # <-- ตรงนี้ใช้ doc_id=

    d.status = "APPROVED"

    # เปลี่ยนสถานะอุปกรณ์เป็น RENTED
    changed, missing = _update_equipment_from_quote(d, "RENTED")

    # ให้แน่ใจว่ามี BL/IV/RC ผูกกับ BK แล้ว
    _ensure_children_for_booking(d)

    db.session.commit()

    if missing:
        flash(f"อุปกรณ์บางตัวไม่มีในระบบ หรือไม่พร้อมให้เช่า: {', '.join(missing)}", "warning")
    else:
        flash("อนุมัติใบจองเรียบร้อยแล้ว", "success")

    ct = SalesDoc.query.filter_by(parent_id=d.parent_id, doc_type='CT').first()
    if ct:
        flash('ล็อคของแล้ว — ไปจัดการงวดในสัญญา/PO ใหญ่ต่อได้เลย', 'success')
        return redirect(url_for('contract_view', cid=ct.id))
    return redirect(url_for('bk_view', doc_id=d.id))  # <-- ตรงนี้ก็ใช้ doc_id=


@app.get("/sales/bookings/<int:doc_id>/print")
@login_required
@permission_required("sales.view")
def bk_print(doc_id):
    # โหลด BK + customer + รายการ
    bk: SalesDoc = (
        SalesDoc.query
        .options(
            joinedload(SalesDoc.customer),
            joinedload(SalesDoc.items),   # ← เอา .joinedload(SalesItem.equipment) ออก
        )
        .filter(
            SalesDoc.id == doc_id,
            SalesDoc.doc_type == "BK",
        )
        .first_or_404()
    )

    # โปรไฟล์บริษัท (ใช้ตัวเดียวกับ QU print)
    company = CompanyProfile.query.first()

    # map รูปภาพอุปกรณ์ (ใช้ helper เดิม ถ้ามี)
    img_map = _build_item_image_map(bk) if "_build_item_image_map" in globals() else {}

    # parent QU (ถ้ามี) ไว้ใช้แสดงข้อความอ้างอิง
    parent_qu = bk.parent if bk.parent and bk.parent.doc_type == "QU" else None

    return render_template(
        "sales/bk_print.html",
        d=bk,
        company=company,
        img_map=img_map,
        parent_qu=parent_qu,
    )




# ================== Export Excel (Reports) ==================
from io import BytesIO

def _excel_response(filename: str, headers: list[str], rows: list[list]):
    """สร้างไฟล์ .xlsx แล้วส่งกลับเป็น response"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)

    # ปรับความกว้างคอลัมน์แบบง่าย ๆ
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1] or ""))
        for rr in range(2, min(len(rows) + 2, 502)):  # จำกัดการวนเพื่อไม่ให้ช้า
            val = ws.cell(row=rr, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/export/purchases/po.xlsx")
@login_required
@permission_required("purchases.view")
def export_po_excel():
    """Export รายการใบสั่งซื้อ (PO) ตามช่วงวันที่ (po_date) และค้นหา (q)"""
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    qry = PurchaseOrder.query.options(joinedload(PurchaseOrder.supplier))

    if start_d:
        qry = qry.filter(PurchaseOrder.po_date >= start_d)
    if end_d:
        qry = qry.filter(PurchaseOrder.po_date <= end_d)

    if q:
        like = f"%{q}%"
        qry = qry.outerjoin(PurchaseOrder.supplier).filter(
            or_(PurchaseOrder.number.ilike(like), Supplier.name.ilike(like))
        )

    pos = qry.order_by(PurchaseOrder.id.desc()).all()

    headers = [
        "PO No", "Date", "Supplier", "Status",
        "Subtotal", "VAT", "Total",
        "Remark",
    ]
    rows = []
    for d in pos:
        supplier_name = ""
        try:
            supplier_name = d.supplier.name if getattr(d, "supplier", None) else ""
        except Exception:
            supplier_name = ""

        rows.append([
            getattr(d, "number", "") or "",
            getattr(d, "po_date", None) or "",
            supplier_name,
            getattr(d, "status", "") or "",
            float(getattr(d, "subtotal", 0.0) or 0.0),
            float(getattr(d, "vat", 0.0) or 0.0),
            float(getattr(d, "total", 0.0) or 0.0),
            getattr(d, "remark", "") or "",
        ])

    return _excel_response("purchase_orders.xlsx", headers, rows)

@app.get("/export/purchases/grn.xlsx")
@login_required
@permission_required("goods.receive")
def export_grn_excel():
    """Export รายการใบรับสินค้า (RC) ตามช่วงวันที่ (grn_date) และค้นหา (q)"""
    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    qry = GoodsReceipt.query.options(
        joinedload(GoodsReceipt.po).joinedload(PurchaseOrder.supplier),
        joinedload(GoodsReceipt.po),
    )

    if start_d:
        qry = qry.filter(GoodsReceipt.grn_date >= start_d)
    if end_d:
        qry = qry.filter(GoodsReceipt.grn_date <= end_d)

    if q:
        like = f"%{q}%"
        qry = qry.outerjoin(GoodsReceipt.po).outerjoin(PurchaseOrder.supplier).outerjoin(GoodsReceipt.po).filter(
            or_(
                GoodsReceipt.number.ilike(like),
                Supplier.name.ilike(like),
                PurchaseOrder.number.ilike(like),
            )
        )

    grns = qry.order_by(GoodsReceipt.id.desc()).all()

    headers = [
        "GRN No", "Date", "Supplier", "PO No",
        "Remark",
    ]
    rows = []
    for g in grns:
        supplier_name = ""
        try:
            supplier_name = g.supplier.name if getattr(g, "supplier", None) else ""
        except Exception:
            supplier_name = ""

        po_no = ""
        try:
            po_no = g.po.number if getattr(g, "po", None) else ""
        except Exception:
            po_no = ""

        rows.append([
            getattr(g, "number", "") or "",
            getattr(g, "grn_date", None) or "",
            supplier_name,
            po_no,
            getattr(g, "remark", "") or "",
        ])

    return _excel_response("goods_receipts.xlsx", headers, rows)

@app.get("/export/sales/<string:doc_type>.xlsx")
@login_required
@permission_required("sales.view")
def export_sales_doc_excel(doc_type: str):
    """Export เอกสารขายจากตาราง SalesDoc ตาม doc_type (เช่น QU/BK/BL/IV/RC/CT/...)"""
    dt = (doc_type or "").upper()
    docs = (
        SalesDoc.query
        .options(joinedload(SalesDoc.customer))
        .filter(SalesDoc.doc_type == dt)
        .order_by(SalesDoc.id.desc())
        .all()
    )
    headers = [
        "Doc Type", "Number", "Date", "Customer", "Status",
        "Subtotal", "VAT", "Total", "WHT", "Grand",
        "Project", "PO Customer", "Credit Days",
        "Billing Mode", "Installments", "Contract Start", "Contract End",
        "Remark",
    ]
    rows = []
    for d in docs:
        cust_name = ""
        try:
            cust_name = d.customer.name if getattr(d, "customer", None) else ""
        except Exception:
            cust_name = ""
        rows.append([
            getattr(d, "doc_type", "") or "",
            getattr(d, "number", "") or "",
            getattr(d, "date", None) or "",
            cust_name,
            getattr(d, "status", "") or "",
            float(getattr(d, "amount_subtotal", 0) or 0),
            float(getattr(d, "amount_vat", 0) or 0),
            float(getattr(d, "amount_total", 0) or 0),
            float(getattr(d, "amount_wht", 0) or 0),
            float(getattr(d, "amount_grand", 0) or 0),
            getattr(d, "project_name", "") or "",
            getattr(d, "po_customer", "") or "",
            int(getattr(d, "credit_days", 0) or 0),
            getattr(d, "billing_mode", "") or "",
            int(getattr(d, "installment_count", 0) or 0),
            getattr(d, "contract_start", None) or "",
            getattr(d, "contract_end", None) or "",
            getattr(d, "remark", "") or "",
        ])
    return _excel_response(f"sales_{dt.lower()}.xlsx", headers, rows)




# =========================
# Export Reports (Excel)
# =========================
def _parse_date_yyyy_mm_dd(s: str):
    """Parse YYYY-MM-DD -> date or None."""
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


@app.get("/export/sales_report/<doc_type>.xlsx")
@login_required
@permission_required("sales.view")
def export_sales_docs_report_excel(doc_type: str):
    """
    Export summary report of SalesDoc by doc_type with optional date range.
    Query params:
      - start=YYYY-MM-DD
      - end=YYYY-MM-DD
      - status=... (optional exact match, case-insensitive)
      - q=... (optional, search in number / customer name)
    """
    from datetime import datetime as _dt
    from sqlalchemy.orm import joinedload
    from openpyxl import Workbook

    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    status_q = (request.args.get("status") or "").strip()
    q = (request.args.get("q") or "").strip()

    start_d = _parse_date_yyyy_mm_dd(start_s)
    end_d = _parse_date_yyyy_mm_dd(end_s)
    if start_d and end_d and start_d > end_d:
        start_d, end_d = end_d, start_d
    if start_d and not end_d:
        end_d = start_d
    if end_d and not start_d:
        start_d = end_d

    # base query
    query = (
        SalesDoc.query
        .options(joinedload(SalesDoc.customer))
        .filter(SalesDoc.doc_type == (doc_type or "").upper())
    )

    # optional date filter (by document date)
    if start_d and end_d:
        query = query.filter(SalesDoc.date.between(start_d, end_d))

    # optional status filter
    if status_q:
        query = query.filter(func.upper(SalesDoc.status) == status_q.upper())

    # optional keyword search
    if q:
        query = query.outerjoin(Customer, SalesDoc.customer_id == Customer.id).filter(
            or_(
                SalesDoc.number.ilike(f"%{q}%"),
                Customer.name.ilike(f"%{q}%"),
            )
        )

    docs = query.order_by(SalesDoc.date.desc().nullslast(), SalesDoc.id.desc()).all()

    def _safe_float(x):
        try:
            return float(x or 0)
        except Exception:
            return 0.0

    def _doc_amount(d: "SalesDoc") -> float:
        for attr in ("amount_grand", "amount_total", "amount_subtotal"):
            if hasattr(d, attr):
                v = getattr(d, attr) or 0
                try:
                    return float(v)
                except Exception:
                    continue
        return 0.0

    wb = Workbook()
    ws = wb.active
    ws.title = f"{(doc_type or '').upper()}"

    ws.append([
        "วันที่",
        "เลขที่เอกสาร",
        "ประเภท",
        "ชื่อลูกค้า",
        "สถานะ",
        "ยอดก่อนภาษี",
        "VAT",
        "ยอดสุทธิ",
        "หมายเหตุ",
    ])

    for d in docs:
        cust_name = d.customer.name if getattr(d, "customer", None) else "-"
        ws.append([
            d.date.strftime("%Y-%m-%d") if d.date else "",
            d.number or "",
            d.doc_type or "",
            cust_name,
            d.status or "",
            _safe_float(getattr(d, "amount_subtotal", 0)),
            _safe_float(getattr(d, "amount_vat", 0)),
            _doc_amount(d),
            getattr(d, "remark", "") or "",
        ])

    # autosize
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 48)

    filename = f"{(doc_type or 'sales').upper()}_report"
    if start_d and end_d:
        filename += f"_{start_d.strftime('%Y%m%d')}_{end_d.strftime('%Y%m%d')}"
    return _xlsx_response(wb, filename=filename)


@app.get("/returns/export.xlsx")
@login_required
@permission_required("sales.manage")
def returns_export_excel():
    """Export รายการใบคืนสินค้า ตามช่วงวันที่ (ReturnDoc.date) + ค้นหา (q)

คอลัมน์: วันที่, เลขที่ใบคืน, เลขที่ใบเสนอราคา, ชื่อลูกค้า, ยอดคืนรวม, สถานะ, หมายเหตุ
ยอดคืนรวม = รวม (line_total/qty จาก QU item ตามหมวดหมู่) * จำนวนที่คืน
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    q = (request.args.get("q") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    start_d = _parse_date_yyyy_mm_dd(start)
    end_d = _parse_date_yyyy_mm_dd(end)

    query = ReturnDoc.query.options(
        joinedload(ReturnDoc.customer),
        joinedload(ReturnDoc.quote).joinedload(SalesDoc.items),
        joinedload(ReturnDoc.items).joinedload(ReturnItem.equipment),
    )

    # soft delete
    is_del_col = getattr(ReturnDoc, "is_deleted", None)
    if is_del_col is not None:
        query = query.filter(is_del_col.is_(False))

    if start_d:
        query = query.filter(ReturnDoc.date >= start_d)
    if end_d:
        query = query.filter(ReturnDoc.date <= end_d)

    if q:
        like = f"%{q}%"
        query = (
            query.outerjoin(ReturnDoc.customer)
            .outerjoin(ReturnDoc.quote)
            .filter(
                or_(
                    ReturnDoc.number.ilike(like),
                    SalesDoc.number.ilike(like),
                    Customer.name.ilike(like),
                )
            )
        )

    docs = query.order_by(ReturnDoc.date.desc(), ReturnDoc.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Returns"

    ws.append([
        "วันที่",
        "เลขที่ใบคืน",
        "เลขที่ใบเสนอราคา",
        "ชื่อลูกค้า",
        "ยอดคืนรวม",
        "สถานะ",
        "หมายเหตุ",
    ])

    def _remark_of(d):
        for k in ("remark", "note", "notes", "comment"):
            v = getattr(d, k, None)
            if v:
                return str(v)
        return ""

    for d in docs:
        cust_name = d.customer.name if getattr(d, "customer", None) else ""
        qu_no = d.quote.number if getattr(d, "quote", None) else ""

        # build unit_value per category from quote items: sum(line_total)/sum(qty)
        unit_by_cat = {}
        if getattr(d, "quote", None) and getattr(d.quote, "items", None):
            sums = {}
            qtys = {}
            for it in d.quote.items:
                cid = getattr(it, "category_id", None)
                if not cid:
                    continue
                lt = float(getattr(it, "line_total", 0.0) or 0.0)
                qit = float(getattr(it, "qty", 0.0) or 0.0)
                if qit <= 0:
                    continue
                sums[cid] = sums.get(cid, 0.0) + lt
                qtys[cid] = qtys.get(cid, 0.0) + qit
            for cid, total in sums.items():
                qsum = qtys.get(cid, 0.0) or 0.0
                unit_by_cat[cid] = (total / qsum) if qsum > 0 else 0.0

        total_return = 0.0
        for ri in (getattr(d, "items", None) or []):
            eq = getattr(ri, "equipment", None)
            cid = getattr(eq, "category_id", None) if eq else None
            unit = float(unit_by_cat.get(cid, 0.0) or 0.0)
            qty = float(getattr(ri, "qty", 0.0) or 0.0)
            total_return += unit * qty

        ws.append([
            getattr(d, "date", None) or "",
            getattr(d, "number", "") or "",
            qu_no,
            cust_name,
            float(total_return),
            getattr(d, "status", "") or "",
            _remark_of(d),
        ])

    # auto width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            try:
                max_len = max(max_len, len(str(v)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    return _xlsx_send(wb, "returns.xlsx")

@app.get("/sales/contracts/export.xlsx")
@login_required
@permission_required("sales.view")
def contracts_export_excel():
    """
    Export summary report for Contracts (SalesDoc doc_type=CT) by date range.
    Includes:
      - customer, PO ใหญ่ (po_customer), amount
      - contract_start/end, installment_count
      - paid_installments count
      - related document numbers from SalesInstallment.related_doc_id
    Query params:
      - start=YYYY-MM-DD
      - end=YYYY-MM-DD
      - q=... (optional)
      - status=... (optional exact match, case-insensitive)
    """
    from sqlalchemy.orm import joinedload
    from openpyxl import Workbook

    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    q = (request.args.get("q") or "").strip()
    status_q = (request.args.get("status") or "").strip()

    start_d = _parse_date_yyyy_mm_dd(start_s)
    end_d = _parse_date_yyyy_mm_dd(end_s)
    if start_d and end_d and start_d > end_d:
        start_d, end_d = end_d, start_d
    if start_d and not end_d:
        end_d = start_d
    if end_d and not start_d:
        start_d = end_d

    query = (
        SalesDoc.query
        .options(joinedload(SalesDoc.customer), joinedload(SalesDoc.installments))
        .filter(SalesDoc.doc_type == "CT")
    )

    if start_d and end_d:
        query = query.filter(SalesDoc.date.between(start_d, end_d))

    if status_q:
        query = query.filter(func.upper(SalesDoc.status) == status_q.upper())

    if q:
        query = query.outerjoin(Customer, SalesDoc.customer_id == Customer.id).filter(
            or_(
                SalesDoc.number.ilike(f"%{q}%"),
                Customer.name.ilike(f"%{q}%"),
                SalesDoc.po_customer.ilike(f"%{q}%"),
                SalesDoc.project_name.ilike(f"%{q}%"),
            )
        )

    contracts = query.order_by(SalesDoc.date.desc().nullslast(), SalesDoc.id.desc()).all()

    # map related doc ids -> number
    related_ids = set()
    for c in contracts:
        for inst in (getattr(c, "installments", None) or []):
            if getattr(inst, "related_doc_id", None):
                related_ids.add(inst.related_doc_id)

    related_map = {}
    if related_ids:
        rel_docs = SalesDoc.query.filter(SalesDoc.id.in_(list(related_ids))).all()
        related_map = {d.id: (d.number or "") for d in rel_docs}

    def _safe_float(x):
        try:
            return float(x or 0)
        except Exception:
            return 0.0

    def _doc_amount(d: "SalesDoc") -> float:
        for attr in ("amount_grand", "amount_total", "amount_subtotal"):
            if hasattr(d, attr):
                v = getattr(d, attr) or 0
                try:
                    return float(v)
                except Exception:
                    continue
        return 0.0

    def _is_paid(inst: "SalesInstallment") -> bool:
        st = (getattr(inst, "status", "") or "").upper()
        if st in ("PAID", "PAID_FULL", "DONE"):
            return True
        if getattr(inst, "paid_at", None):
            return True
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "CONTRACTS"

    ws.append([
        "วันที่สร้าง",
        "เลขที่สัญญา/PO ใหญ่",
        "ชื่อลูกค้า",
        "PO ลูกค้า",
        "ยอดสุทธิ",
        "เริ่มสัญญา",
        "สิ้นสุดสัญญา",
        "แบ่งงวด",
        "ชำระแล้ว(งวด)",
        "เลขที่เอกสารที่เกี่ยวข้อง",
        "สถานะ",
    ])

    for c in contracts:
        insts = list(getattr(c, "installments", None) or [])
        inst_count = int(getattr(c, "installment_count", 0) or len(insts) or 0)
        paid_cnt = sum(1 for it in insts if _is_paid(it))

        related_nums = []
        for it in insts:
            rid = getattr(it, "related_doc_id", None)
            if rid and related_map.get(rid):
                related_nums.append(related_map[rid])
        related_str = ", ".join([x for x in related_nums if x])

        ws.append([
            c.date.strftime("%Y-%m-%d") if c.date else "",
            c.number or "",
            (c.customer.name if getattr(c, "customer", None) else "-"),
            c.po_customer or "",
            _doc_amount(c),
            c.contract_start.strftime("%Y-%m-%d") if getattr(c, "contract_start", None) else "",
            c.contract_end.strftime("%Y-%m-%d") if getattr(c, "contract_end", None) else "",
            inst_count,
            paid_cnt,
            related_str,
            c.status or "",
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)

    filename = "contracts_report"
    if start_d and end_d:
        filename += f"_{start_d.strftime('%Y%m%d')}_{end_d.strftime('%Y%m%d')}"
    return _xlsx_response(wb, filename=filename)



# --- 403 Forbidden page ---
@app.errorhandler(403)
def forbidden(e):
    return render_template("errors/403.html"), 403

# ================== Main ==================


@app.get("/export/equipment_report.xlsx")
@login_required
@permission_required("equipment.view")
def export_equipment_report_excel():
    """
    Export equipment report with optional filters.
    Query params:
      - start=YYYY-MM-DD (filter by received_date >= start)
      - end=YYYY-MM-DD (filter by received_date <= end)
      - status=READY|RENTED|... (optional exact match)
      - q=... (optional, search in sku / name)
    Columns:
      - รหัสสินค้า (SKU เต็ม)
      - ชื่อ
      - หมวดหมู่
      - สถานะ
      - รับเข้า (received_date)
    """
    from openpyxl import Workbook
    from sqlalchemy.orm import joinedload

    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    status_q = (request.args.get("status") or "").strip().upper()
    q = (request.args.get("q") or "").strip()

    start_d = _parse_date_yyyy_mm_dd(start_s)
    end_d = _parse_date_yyyy_mm_dd(end_s)

    query = Equipment.query.options(joinedload(Equipment.category))

    if q:
        like = f"%{q}%"
        query = query.filter(or_(Equipment.sku.ilike(like), Equipment.name.ilike(like)))
    if status_q and status_q in EQUIP_STATUS:
        query = query.filter(Equipment.status == status_q)
    if start_d:
        query = query.filter(Equipment.received_date >= start_d)
    if end_d:
        query = query.filter(Equipment.received_date <= end_d)

    rows = query.order_by(Equipment.received_date.desc(), Equipment.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Report"

    headers = ["รหัสสินค้า", "ชื่อ", "หมวดหมู่", "สถานะ", "รับเข้า"]
    ws.append(headers)

    for e in rows:
        ws.append([
            e.sku or "",
            e.name or "",
            (e.category.name if getattr(e, "category", None) else ""),
            (e.status_th if hasattr(e, "status_th") else (e.status or "")),
            (e.received_date.strftime("%d/%m/%Y") if getattr(e, "received_date", None) else ""),
        ])

    # basic column width
    try:
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 14
    except Exception:
        pass

    return _xlsx_response(wb, "equipment_report.xlsx")



if __name__ == "__main__":
    bootstrap()
    app.run(host="127.0.0.1", port=8000, debug=True)
